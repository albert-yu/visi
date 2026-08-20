#!/usr/bin/env python3
"""Bindings/CLI equivalence.

The `visi_core` extension module and the `visi` CLI must be observationally
identical for every operation the fuzz harness drives. Nothing else checks
this: the two backends duplicate a little logic on purpose (`edit_chart`'s
clear-vs-set flags, `add_pivot_field`'s post-add subtotal mutation), and
without a test that duplication drifts silently and the fuzzer starts
measuring something other than what it reports.

Corpus: freshly generated workbooks at fixed seeds, so this is fully
self-contained on a clean checkout, plus any `fuzz_results/failures/*/source.xlsx`
lying around locally. That second source is opportunistic and usually empty --
`fuzz_results/` is gitignored, and only a genuine differential failure puts
anything there. Do not rely on it for coverage.

Comparison is over *parsed content*, never bytes: docProps/core.xml carries a
creation timestamp and chart/pivot ids come from a random `generate_unique_id`,
so two runs never produce identical archives.

    source fuzz/venv/bin/activate
    pytest fuzz/test_backend_parity.py
"""

import glob
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from chart_xlsx_reader import read_charts
from fuzz_chart import ChartFuzzGenerator
from fuzz_excel import ExcelFuzzGenerator, XLSXEvaluatedReader
from fuzz_pivot import DEST_CELL, DEST_RC, PIVOT_NAME, PivotFuzzGenerator
from fuzz_structural_edit import StructuralFuzzGenerator
from fuzz_vba import VbaGenerator
from fuzz_vba_parse_grammar import VbaGrammarGenerator
from visi_driver import (
    VisiChartDriver,
    VisiDriver,
    VisiPivotDriver,
    bindings_available,
)

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

pytestmark = pytest.mark.skipif(
    not bindings_available(),
    reason="build the bindings with `maturin develop -m visi-python/Cargo.toml --release`",
)


def _cli_binary_exists():
    from visi_driver import resolve_visi_binary

    p = resolve_visi_binary(None)
    return p is not None and os.path.exists(p)


requires_cli = pytest.mark.skipif(
    not _cli_binary_exists(), reason="no compiled visi binary; run `cargo build --release`"
)


def saved_failure_sources():
    return sorted(
        glob.glob(os.path.join(PROJECT_ROOT, "fuzz_results", "failures", "*", "source.xlsx"))
    )


# --------------------------------------------------------------------- generator coverage


def test_formula_text_function_dispatch_covers_listed_names():
    ExcelFuzzGenerator._check_text_function_generators()


def test_issue_94_high_value_generators_are_wired():
    gen = ExcelFuzzGenerator(seed=94)
    assert "CELL(" in gen.generate_range_info_formula("CELL", 5, 1, 3)
    assert "INFO(" in gen.generate_range_info_formula("INFO", 5, 1, 3)


def test_structural_edit_generator_tracks_dimensions():
    for seed in range(95, 115):
        dims = {sheet: [8, 8] for sheet in ["Sheet1", "Data"]}
        for edit in StructuralFuzzGenerator(seed).edits():
            axis = 0 if "row" in edit.kind else 1
            assert 1 <= edit.index <= dims[edit.sheet][axis]
            dims[edit.sheet][axis] += 1 if edit.kind.startswith("insert") else -1
            assert dims[edit.sheet][axis] >= 1


def test_vba_grammar_generator_covers_wide_syntax():
    modules = VbaGrammarGenerator(seed=97).project()
    sources = "\n".join(m.source for m in modules)
    assert "Private Type Point" in sources
    assert "Public Enum GeneratedColor" in sources
    assert "Public Property Get Answer" in sources
    assert "GoSub" in sources
    assert " _\n" in sources
    assert "first:=1" in sources


def test_vba_extended_host_surface_generates_stateful_object_model_cases():
    gen = VbaGenerator(seed=96, host_surface="extended")
    sources = "\n".join(gen.module(i) for i in range(1, 40))
    assert ".Interior." in sources or ".Font." in sources
    assert ".Rows(" in sources or ".Columns(" in sources
    assert ".ListObjects(" in sources


def test_chart_rich_shape_writes_dates_blanks_and_extra_columns(tmp_path):
    src = str(tmp_path / "chart.xlsx")
    ChartFuzzGenerator(seed=98, shape="rich").generate(src, num_rows=12)
    import openpyxl
    wb = openpyxl.load_workbook(src)
    ws = wb["Sheet1"]
    assert ws.max_column >= 4
    assert any(ws.cell(r, 2).value is None for r in range(2, 13))
    assert any(getattr(ws.cell(r, 1).value, "year", None) == 2026 for r in range(1, 13))


def test_pivot_rich_shape_expands_source_schema(tmp_path):
    src = str(tmp_path / "pivot.xlsx")
    config = PivotFuzzGenerator(seed=98, shape="rich").generate(src, num_rows=10, use_table=True)
    assert config["source_range"].startswith("A1:I")
    assert config["source_bounds"] == (0, 0, 10, 8)
    fields = {f["column"] for f in config["row_fields"] + config["col_fields"] + config["value_fields"]}
    assert fields <= set(PivotFuzzGenerator.RICH_COL_NAMES)


# --------------------------------------------------------------------- eval


@requires_cli
@pytest.mark.parametrize("seed", [48291, 1, 7, 12345, 999983])
def test_eval_parity_on_generated_workbooks(seed, tmp_path):
    src = str(tmp_path / "source.xlsx")
    ExcelFuzzGenerator(seed=seed).create_fuzz_workbook(src, num_rows=10, num_cols=5)
    _assert_eval_parity(src, tmp_path)


@requires_cli
@pytest.mark.skipif(not saved_failure_sources(), reason="no saved failures locally")
@pytest.mark.parametrize("src", saved_failure_sources() or [None])
def test_eval_parity_on_saved_failures(src, tmp_path):
    _assert_eval_parity(src, tmp_path)


def _assert_eval_parity(src, tmp_path):
    cli_out = str(tmp_path / "cli.xlsx")
    bnd_out = str(tmp_path / "bnd.xlsx")

    VisiDriver(backend="subprocess").run(src, cli_out)
    VisiDriver(backend="bindings").run(src, bnd_out)

    cli_cells = XLSXEvaluatedReader.read_evaluated_cells(cli_out)
    bnd_cells = XLSXEvaluatedReader.read_evaluated_cells(bnd_out)

    assert set(cli_cells) == set(bnd_cells), "different cells present"
    mismatches = {
        key: (cli_cells[key]["val"], bnd_cells[key]["val"])
        for key in cli_cells
        if cli_cells[key]["val"] != bnd_cells[key]["val"]
    }
    assert not mismatches, f"{len(mismatches)} value mismatch(es): {list(mismatches.items())[:5]}"


def test_run_returns_the_bytes_it_wrote(tmp_path):
    """The main loop parses the return value instead of re-reading the file."""
    src = str(tmp_path / "source.xlsx")
    out = str(tmp_path / "out.xlsx")
    ExcelFuzzGenerator(seed=4242).create_fuzz_workbook(src, num_rows=5, num_cols=3)

    data = VisiDriver(backend="bindings").run(src, out)
    with open(out, "rb") as f:
        assert data == f.read()
    assert XLSXEvaluatedReader.read_evaluated_cells_bytes(
        data
    ) == XLSXEvaluatedReader.read_evaluated_cells(out)


# ------------------------------------------------------------------- charts


@requires_cli
@pytest.mark.parametrize("seed", range(6))
def test_chart_parity(seed, tmp_path):
    src = str(tmp_path / "source.xlsx")
    range_str, add_config, edit_config = ChartFuzzGenerator(seed=seed).generate(
        src, num_rows=6
    )

    cli_out = str(tmp_path / "cli.xlsx")
    bnd_out = str(tmp_path / "bnd.xlsx")
    VisiChartDriver(backend="subprocess").run(
        src, range_str, add_config, edit_config, cli_out
    )
    VisiChartDriver(backend="bindings").run(
        src, range_str, add_config, edit_config, bnd_out
    )

    assert read_charts(cli_out) == read_charts(bnd_out)


# ------------------------------------------------------------------- pivots


@requires_cli
@pytest.mark.parametrize("seed", range(10))
@pytest.mark.parametrize("use_table", [False, True])
def test_pivot_parity(seed, use_table, tmp_path):
    src = str(tmp_path / "source.xlsx")
    config = PivotFuzzGenerator(seed=seed).generate(src, num_rows=8, use_table=use_table)

    cli_out = str(tmp_path / "cli.xlsx")
    bnd_out = str(tmp_path / "bnd.xlsx")
    VisiPivotDriver(backend="subprocess").run(
        src, config, cli_out, PIVOT_NAME, DEST_CELL, DEST_RC
    )
    VisiPivotDriver(backend="bindings").run(
        src, config, bnd_out, PIVOT_NAME, DEST_CELL, DEST_RC
    )

    cli_cells = XLSXEvaluatedReader.read_evaluated_cells(cli_out)
    bnd_cells = XLSXEvaluatedReader.read_evaluated_cells(bnd_out)
    assert cli_cells.keys() == bnd_cells.keys()
    for key in cli_cells:
        assert cli_cells[key]["val"] == bnd_cells[key]["val"], key


def test_empty_filter_selection_is_bindings_only(tmp_path):
    """`set_pivot_filter(name, col, [])` -- select nothing -- asserted here
    because no driver exercises it.

    `visi pivot filter` takes a non-empty comma list or --clear, with no verb
    for "select nothing"; only the bindings API can express the state at all.
    Neither backend of `VisiPivotDriver` applies it, though, and that is
    deliberate: real Excel refuses to hide a page field's last visible item, so
    an empty selection is a config the differential oracle cannot represent
    (see VisiPivotDriver.run). This test is therefore the only coverage of the
    engine behavior -- keep it even if the CLI grows the verb.
    """
    import visi_core

    wb = visi_core.Workbook()
    rows = [
        ("Region", "Amount"),
        ("East", "10"),
        ("West", "20"),
    ]
    for r, row in enumerate(rows):
        for c, val in enumerate(row):
            wb.set_cell(r, c, val)
    wb.evaluate()
    wb.add_pivot_from_range(
        "P", start_row=0, start_col=0, end_row=2, end_col=1, dest_row=0, dest_col=4
    )
    wb.add_pivot_field("P", "value", "Amount", agg="sum")
    wb.add_pivot_field("P", "filter", "Region")

    wb.set_pivot_filter("P", "Region", [])
    assert wb.pivots()[0]["filter_selections"]["Region"] == []

    wb.set_pivot_filter("P", "Region", None)
    assert wb.pivots()[0]["filter_selections"]["Region"] is None


# --------------------------------------------------------------- error text


@requires_cli
def test_binding_error_text_matches_the_cli(tmp_path):
    """str(exc) must equal the CLI's stderr minus its "Error: " prefix."""
    import subprocess

    import visi_core

    from visi_driver import resolve_visi_binary

    src = str(tmp_path / "source.xlsx")
    visi_core.Workbook().save(src)

    with pytest.raises(visi_core.NotFoundError) as exc:
        visi_core.Workbook.load(src).refresh_pivot("no-such-pivot")

    res = subprocess.run(
        [resolve_visi_binary(None), "pivot", "refresh", src, "--name", "no-such-pivot"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    assert res.returncode != 0
    cli_msg = res.stderr.strip().removeprefix("Error: ")
    assert str(exc.value) == cli_msg


# ------------------------------------------------------------------- macros


MACRO_SRC = 'Attribute VB_Name = "Mod1"\nPublic Sub Hello()\n    Range("A1").Value = 1\nEnd Sub\n'


@requires_cli
@pytest.mark.parametrize("kind,sheet", [("standard", None), ("document", "Sheet1")])
def test_macro_add_parity(kind, sheet, tmp_path):
    """`visi macro add` and `Workbook.add_macro` must produce the same module.

    The bindings duplicate the CLI's resolve-sheet-name-to-id step and its
    ThisWorkbook special case (visi-core takes a sheet id, not a name), so
    this is the same kind of mirrored logic as `edit_chart`'s flags.
    """
    import subprocess

    import visi_core

    from visi_driver import resolve_visi_binary

    src = str(tmp_path / "source.xlsx")
    bas = str(tmp_path / "mod1.bas")
    cli_out = str(tmp_path / "cli.xlsm")
    base = visi_core.Workbook()
    if sheet is not None:
        # Both backends resolve the sheet by name, so take it from the
        # workbook rather than assuming what an empty one calls its sheet.
        sheet = base.sheet_names[0]
    base.save(src)
    with open(bas, "w") as f:
        f.write(MACRO_SRC)

    cmd = [resolve_visi_binary(None), "macro", "add", src, "--name", "Mod1",
           "--kind", kind, "--source-file", bas, "--output", cli_out, "--quiet"]
    if sheet is not None:
        cmd += ["--sheet", sheet]
    res = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    assert res.returncode == 0, res.stderr

    wb = visi_core.Workbook.load(src)
    wb.add_macro("Mod1", MACRO_SRC, kind=kind, sheet=sheet)
    bindings_out = str(tmp_path / "bindings.xlsm")
    wb.save(bindings_out)

    def modules(path):
        return [
            (m["name"], m["kind"], m["source"], m["bound_sheet_id"] is not None)
            for m in visi_core.Workbook.load(path).macros()
        ]

    assert modules(cli_out) == modules(bindings_out)


@requires_cli
def test_macro_list_parity(tmp_path):
    """The dicts `macros()` returns must carry `macro list --json`'s keys."""
    import json
    import subprocess

    import visi_core

    from visi_driver import resolve_visi_binary

    path = str(tmp_path / "book.xlsm")
    wb = visi_core.Workbook()
    wb.add_macro("Mod1", MACRO_SRC)
    wb.add_macro("ThisWorkbook", 'Attribute VB_Name = "ThisWorkbook"\n', kind="document")
    wb.save(path)

    res = subprocess.run(
        [resolve_visi_binary(None), "macro", "list", path, "--json"],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
    )
    assert res.returncode == 0, res.stderr
    cli = json.loads(res.stdout)

    got = visi_core.Workbook.load(path).macros()
    assert [(m["name"], m["kind"], m["source_lines"]) for m in cli] == [
        (m["name"], m["kind"], m["source_lines"]) for m in got
    ]


# `Range("A1")` unqualified is the active sheet, and `C1` reads back what the
# formula in `B1` computed -- so this exercises a write, a recalculation and a
# worksheet-function call in four lines.
RUN_MACRO_SRC = (
    'Attribute VB_Name = "Runner"\n'
    "Public Function Go() As Variant\n"
    '    Range("A1").Value = 4\n'
    '    Range("B1").Formula = "=A1*2"\n'
    '    Range("C1").Value = Application.WorksheetFunction.Sum(Range("A1:B1"))\n'
    '    Go = Range("C1").Value\n'
    "End Function\n"
    "Public Function Peek() As Variant\n"
    "    Peek = ThisWorkbook.Worksheets.Count\n"
    "End Function\n"
)


@requires_cli
def test_macro_run_parity(tmp_path):
    """`visi macro run --output` and `Workbook.run_macro` + `save` must agree.

    Both the returned value and the *cells the macro wrote*: a run that
    reports the right number while saving the wrong workbook is the failure
    this whole phase is built to avoid, and it is invisible to a test that
    only compares return values.
    """
    import json
    import subprocess

    import visi_core

    from visi_driver import resolve_visi_binary

    src = str(tmp_path / "source.xlsm")
    cli_out = str(tmp_path / "cli.xlsm")
    bindings_out = str(tmp_path / "bindings.xlsm")

    base = visi_core.Workbook()
    base.add_macro("Runner", RUN_MACRO_SRC)
    base.save(src)

    res = subprocess.run(
        [resolve_visi_binary(None), "macro", "run", src, "--name", "Go",
         "--output", cli_out, "--json"],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
    )
    assert res.returncode == 0, res.stderr
    cli = json.loads(res.stdout)

    wb = visi_core.Workbook.load(src)
    type_name, value, mutated = wb.run_macro("Go")
    wb.save(bindings_out)

    assert (cli["type"], cli["value"], cli["mutated"]) == (type_name, value, mutated)
    assert mutated is True

    def cells(path):
        w = visi_core.Workbook.load(path)
        return [w.get_display(0, c) for c in range(3)]

    assert cells(cli_out) == cells(bindings_out)


@requires_cli
def test_macro_run_without_a_write_target_is_an_error_only_when_it_mutated(tmp_path):
    """The CLI refuses to discard a macro's writes; a read-only run is fine.

    The bindings have no equivalent refusal -- nothing there writes a file
    implicitly, so there is nothing to discard -- which is why this asserts
    the CLI's rule and only that the bindings agree about `mutated`.
    """
    import subprocess

    import visi_core

    from visi_driver import resolve_visi_binary

    src = str(tmp_path / "source.xlsm")
    base = visi_core.Workbook()
    base.add_macro("Runner", RUN_MACRO_SRC)
    base.save(src)

    def run(proc):
        return subprocess.run(
            [resolve_visi_binary(None), "macro", "run", src, "--name", proc],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
        )

    mutating = run("Go")
    assert mutating.returncode != 0
    assert "--output" in mutating.stderr

    reading = run("Peek")
    assert reading.returncode == 0, reading.stderr
    # The "a macro ran" notice is not suppressible, and is on stderr so it
    # cannot be confused with the value.
    assert "Running VBA procedure" in reading.stderr

    assert visi_core.Workbook.load(src).run_macro("Peek")[2] is False
    assert visi_core.Workbook.load(src).run_macro("Go")[2] is True
