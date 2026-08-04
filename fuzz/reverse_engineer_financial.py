#!/usr/bin/env python3
"""Reverse-engineers Excel's IRR/XIRR/RATE Newton-Raphson solver by grading a
grid of candidate algorithm variants against real Microsoft Excel.

`finance.rs`'s `rate`/`irr`/`xirr` are Newton-Raphson root finds and, per
`fuzz/README.md`, "a known residual source of differential fuzzer failures":
they can return `#NUM!` on inputs real Excel's own (undocumented) solver
converges on, and occasionally the reverse. This script narrows that gap
empirically rather than by guessing:

1. It generates cashflow/RATE inputs specifically chosen to sit near the
   convergence boundary (multi-sign-change cashflows with several plausible
   IRR roots, guesses swept across a wide range, RATE inputs that push the
   implied rate toward the -100% floor) -- the region generic random fuzzing
   rarely hits, since well-behaved random inputs mostly converge trivially.
2. It writes one formula per test case into a workbook and evaluates it in
   real Excel (ground truth) and in `visi` (current behavior) via the same
   AppleScript/CLI drivers `fuzz_excel.py` uses.
3. It also evaluates every case against a grid of candidate pure-Python
   Newton-Raphson variants (closed-form vs. numeric derivative, several
   iteration caps / tolerances / zero-guess-retry policies -- see
   `CANDIDATE_VARIANTS`), and reports which variant's `#NUM!`/converged-value
   boundary agrees with real Excel most often.

The closed-form derivative used by the "closed" variants is not guessed --
it's the standard OpenOffice-lineage TVM Newton-Raphson formulation used by
`formulajs` (a JS Excel-function reimplementation), translated to Python;
see the docstring on `newton_raphson_generic` for the derivation. Whether
*that* formula, plus which (eps, max_iter, retry) knobs, actually reproduces
Excel is exactly what this script measures rather than assumes.

Usage:
    # Smoke-test the pipeline without Excel (candidates vs. visi only):
    python3 fuzz/reverse_engineer_financial.py --driver mock

    # Full run against real Excel:
    python3 fuzz/reverse_engineer_financial.py \\
        --excel-path "/Applications/Microsoft Excel.app" --seed 1

Output: a per-function ranking of candidate variants by agreement rate with
Excel (best first), visi's own agreement rate as a baseline for comparison,
and the worst mismatches for the best-scoring candidate -- printed to stdout
and dumped as JSON under `fuzz_results/financial_reverse_engineering/`.
"""

import argparse
import itertools
import json
import math
import os
import random
import sys
import time

import openpyxl
import zipfile
import xml.etree.ElementTree as ET

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from fuzz_excel import ExcelDriver, VisiDriver, XLSXEvaluatedReader  # noqa: E402

NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
      "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
      "rel": "http://schemas.openxmlformats.org/package/2006/relationships"}


def map_sheet_names(xlsx_path):
    """`XLSXEvaluatedReader` keys cells by the internal `sheetN.xml`
    basename, not the sheet's display name -- both engines preserve sheet
    order (confirmed: sheet definition order in `xl/workbook.xml` matches
    `xl/_rels/workbook.xml.rels` target order for both rust_xlsxwriter's
    and Excel's own writer here), but relying on that silently would be
    fragile, so this reads the real name -> r:id -> target mapping."""
    mapping = {}
    if not os.path.exists(xlsx_path):
        return mapping
    with zipfile.ZipFile(xlsx_path, "r") as z:
        wb_root = ET.fromstring(z.read("xl/workbook.xml"))
        rels_root = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root.findall("rel:Relationship", NS)
        }
        for sheet in wb_root.findall(".//main:sheets/main:sheet", NS):
            name = sheet.attrib["name"]
            rid = sheet.attrib[f"{{{NS['r']}}}id"]
            target = rid_to_target.get(rid, "")
            basename = os.path.basename(target).replace(".xml", "")
            mapping[name] = basename
    return mapping

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)


def col_name(col_idx):
    """1-based column index -> A1 column letter (1 -> A, 27 -> AA)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


# -----------------------------------------------------------------------------
# Candidate Newton-Raphson algorithm variants
# -----------------------------------------------------------------------------
#
# Each variant is a dict of knobs fed into the generic solver below. `deriv`
# picks the closed-form TVM derivative (matching the formulajs/OpenOffice
# lineage) vs. a central-difference numeric derivative (matching visi's
# current `finance.rs::newton_raphson`, included as a baseline). `cap_error`
# controls what happens when the iteration budget runs out without
# converging: return `#NUM!` (`True`, what Excel's docs claim: "If IRR can't
# find a result... #NUM! is returned") or return the last iterate anyway
# (`False` -- what formulajs's own RATE actually does, despite its docstring
# implying otherwise; worth testing since it's a real discrepancy between
# documented and observed behavior in a widely-used reimplementation).
# `retry_zero` mirrors visi's existing fallback: retry once from a 0.0 guess
# if the caller's guess fails to converge.

CANDIDATE_VARIANTS = [
    {"name": name, "deriv": deriv, "eps": eps, "max_iter": max_iter,
     "cap_error": cap_error, "retry_zero": retry_zero,
     "step_halving": step_halving, "max_step": max_step}
    for deriv in ("closed", "numeric")
    for eps in (1e-6, 1e-7, 1e-10)
    for max_iter in (20, 50, 100, 200)
    for cap_error in (True, False)
    for retry_zero in (False, True)
    for step_halving in (False, True)
    for max_step in (None, 1.0)
    for name in [f"{deriv}/eps={eps:g}/iter={max_iter}/cap={cap_error}/retry0={retry_zero}/halve={step_halving}/ms={max_step}"]
]


def newton_raphson_generic(f, fprime_numeric, fprime_closed, guess, variant):
    """Shared Newton-Raphson core parameterized by `variant` (see
    CANDIDATE_VARIANTS). Includes reverse-engineered step halving, step capping,
    and domain boundary protection.
    """
    rate = guess
    eps = variant["eps"]
    max_iter = variant["max_iter"]
    use_closed = variant["deriv"] == "closed" and fprime_closed is not None
    step_halving = variant.get("step_halving", False)
    max_step = variant.get("max_step")
    allow_pos_transition = guess >= 0.0

    def is_bad(x):
        return isinstance(x, complex) or not math.isfinite(x)

    for _ in range(max_iter):
        if rate <= -0.9999:
            rate = -0.999999
        try:
            y = f(rate)
            if is_bad(y):
                return None
            if abs(y) < eps:
                if not allow_pos_transition and rate > 0.0:
                    return None
                return rate
            dy = fprime_closed(rate) if use_closed else fprime_numeric(rate)
            if is_bad(dy) or dy == 0:
                return None

            step = -y / dy
            if max_step is not None:
                if step > max_step:
                    step = max_step
                elif step < -max_step:
                    step = -max_step

            if step_halving:
                halvings = 0
                while rate + step <= -0.9999 and halvings < 50:
                    step /= 2.0
                    halvings += 1

            new_rate = rate + step
            if is_bad(new_rate) or new_rate <= -0.9999:
                return None

            if not allow_pos_transition and new_rate > 0.0:
                return None
        except (OverflowError, ValueError, ZeroDivisionError):
            return None
        if abs(new_rate - rate) < eps:
            return new_rate
        rate = new_rate

    return None if variant["cap_error"] else rate


def _numeric_deriv(f, rate):
    h = 1e-6 * (1.0 + abs(rate))
    return (f(rate + h) - f(rate - h)) / (2.0 * h)


def candidate_rate(nper, pmt, pv, fv, typ, guess, variant):
    typ = 1.0 if typ else 0.0

    def f(r):
        if abs(r) < 1e-10:
            return pv * (1 + nper * r) + pmt * (1 + r * typ) * nper + fv
        term = (1 + r) ** nper
        return pv * term + pmt * (1 / r + typ) * (term - 1) + fv

    def fprime_closed(r):
        if abs(r) < 1e-10:
            return pv * nper + pmt * typ * nper
        term = (1 + r) ** nper
        dterm = nper * (1 + r) ** (nper - 1)
        return pv * dterm + pmt * (1 / r + typ) * dterm - pmt / (r * r) * (term - 1)

    def fprime_numeric(r):
        return _numeric_deriv(f, r)

    result = newton_raphson_generic(f, fprime_numeric, fprime_closed, guess, variant)
    if result is None and variant["retry_zero"] and guess != 0.0:
        result = newton_raphson_generic(f, fprime_numeric, fprime_closed, 0.0, variant)
    return result


def candidate_irr(values, guess, variant):
    n = len(values)

    def f(r):
        rr = -0.999999999 if r <= -1.0 else r
        total = values[0]
        factor = 1.0
        base = 1.0 + rr
        for i in range(1, n):
            factor *= base
            total += values[i] / factor
        return total

    def fprime_closed(r):
        rr = -0.999999999 if r <= -1.0 else r
        base = 1.0 + rr
        total = 0.0
        factor = base
        for i in range(1, n):
            total += -i * values[i] / (factor * base)
            factor *= base
        return total

    def fprime_numeric(r):
        return _numeric_deriv(f, r)

    result = newton_raphson_generic(f, fprime_numeric, fprime_closed, guess, variant)
    if result is None and variant["retry_zero"] and guess != 0.0:
        result = newton_raphson_generic(f, fprime_numeric, fprime_closed, 0.0, variant)
    return result


def candidate_xirr(values, days, guess, variant):
    """`days[i]` = (date_i - date_0) in days (float), days[0] == 0."""
    n = len(values)
    fracs = [d / 365.0 for d in days]

    def f(r):
        rr = -0.999999999 if r <= -1.0 else r
        base = 1.0 + rr
        return sum(values[i] / (base ** fracs[i]) for i in range(n))

    def fprime_closed(r):
        rr = -0.999999999 if r <= -1.0 else r
        base = 1.0 + rr
        return sum(-fracs[i] * values[i] / (base ** (fracs[i] + 1)) for i in range(1, n))

    def fprime_numeric(r):
        return _numeric_deriv(f, r)

    result = newton_raphson_generic(f, fprime_numeric, fprime_closed, guess, variant)
    if result is None and variant["retry_zero"] and guess != 0.0:
        result = newton_raphson_generic(f, fprime_numeric, fprime_closed, 0.0, variant)
    return result


# -----------------------------------------------------------------------------
# Test case generation -- deliberately adversarial, targeting the
# convergence boundary rather than "typical" well-behaved inputs.
# -----------------------------------------------------------------------------

GUESS_SWEEP = [-0.99, -0.9, -0.5, -0.2, -0.05, 0.0, 0.05, 0.1, 0.3, 0.5, 1.0, 2.0, 5.0, 20.0]

MULTI_ROOT_CASHFLOWS = [
    [-1000, 300, -200, 900, -100, 400],       # two sign flips -> plausibly 2 real roots
    [-100, 500, -500, 500, -500, 500, -100],  # oscillating signs
    [1000, -3000, 2500],                       # positive-first, still requires + and -
    [-50, 200, -200, 200, -200, 200, -50],
    [-100000, 39000, 30000, 21000, 37000],     # textbook IRR example (unique root region)
    [-10, 21, -11],                             # classic Excel dual-root pathological case
    [-1, 100, -100, 100, -100, 1],
]

FLAT_NPV_CASHFLOWS = [
    # Long near-flat streams: NPV(r) barely changes with r near the guess,
    # so the derivative is tiny and Newton-Raphson can overshoot wildly.
    [-1000] + [10] * 40 + [600],
    [-5000] + [125] * 36,
]

EXTREME_MAGNITUDE_CASHFLOWS = [
    [-1e9, 3e8, 3e8, 3e8, 3e8],
    [-0.01, 0.003, 0.003, 0.003, 0.003],
    [-1e6, 1e6 + 1e-3],
]


def gen_irr_cases():
    cases = []
    for cf in MULTI_ROOT_CASHFLOWS + FLAT_NPV_CASHFLOWS + EXTREME_MAGNITUDE_CASHFLOWS:
        for guess in GUESS_SWEEP:
            cases.append({"kind": "irr", "values": cf, "guess": guess})
    return cases


def gen_rate_cases(rng):
    cases = []
    # Push the implied per-period rate toward the -100% floor: pmt large
    # relative to pv/nper (money "returned" per period approaches or
    # exceeds what pv could sustain at any positive rate).
    boundary_configs = [
        (nper, pmt, pv, fv, typ)
        for nper in (4, 12, 36, 120)
        for pmt in (-50, -200, -900, -2500)
        for pv in (1000, 5000)
        for fv in (0, -pv * 0.5)
        for typ in (0, 1)
    ]
    rng.shuffle(boundary_configs)
    for nper, pmt, pv, fv, typ in boundary_configs[:40]:
        for guess in (-0.9, -0.5, -0.1, 0.0, 0.1, 0.5, 2.0):
            cases.append({"kind": "rate", "nper": nper, "pmt": pmt, "pv": pv,
                          "fv": fv, "type": typ, "guess": guess})
    # A batch of realistic loans too, as a well-behaved-case control group.
    for _ in range(20):
        nper = rng.randint(6, 360)
        pv = round(rng.uniform(1000, 50000), 2)
        rate_true = rng.uniform(0.001, 0.02)
        pmt = -round(pv * rate_true / (1 - (1 + rate_true) ** -nper), 2)
        typ = rng.choice([0, 1])
        for guess in (0.1, 0.0, -0.05):
            cases.append({"kind": "rate", "nper": nper, "pmt": pmt, "pv": pv,
                          "fv": 0, "type": typ, "guess": guess})
    return cases


def gen_xirr_cases(rng):
    cases = []
    base_date_serial = 44927  # 2023-01-01 in Excel's 1900 date system
    for cf in MULTI_ROOT_CASHFLOWS:
        # Irregular, non-uniform date gaps (including out-of-order dates,
        # which XIRR explicitly permits per its docs).
        offsets = [0]
        for _ in range(len(cf) - 1):
            offsets.append(offsets[-1] + rng.randint(5, 400))
        if rng.random() < 0.5:
            # shuffle interior dates to test out-of-order handling, keep
            # offsets[0] == 0 as the anchor
            interior = offsets[1:]
            rng.shuffle(interior)
            offsets = [0] + interior
        dates = [base_date_serial + o for o in offsets]
        for guess in GUESS_SWEEP:
            cases.append({"kind": "xirr", "values": cf, "dates": dates, "guess": guess})
    return cases


# -----------------------------------------------------------------------------
# Workbook construction
# -----------------------------------------------------------------------------

def build_workbook(irr_cases, rate_cases, xirr_cases, path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("IRR")
    guess_col = 10  # column J; cashflow arrays here never exceed 8 entries
    for row, case in enumerate(irr_cases, start=1):
        values = case["values"]
        for i, v in enumerate(values):
            ws.cell(row=row, column=1 + i, value=float(v))
        ws.cell(row=row, column=guess_col, value=float(case["guess"]))
        last_col = col_name(len(values))
        formula_col = guess_col + 1
        ws.cell(row=row, column=formula_col,
                 value=f"=IRR(A{row}:{last_col}{row},{col_name(guess_col)}{row})")
        case["_cell"] = ("IRR", f"{col_name(formula_col)}{row}")

    ws = wb.create_sheet("RATE")
    for row, case in enumerate(rate_cases, start=1):
        ws.cell(row=row, column=1, value=float(case["nper"]))
        ws.cell(row=row, column=2, value=float(case["pmt"]))
        ws.cell(row=row, column=3, value=float(case["pv"]))
        ws.cell(row=row, column=4, value=float(case["fv"]))
        ws.cell(row=row, column=5, value=float(case["type"]))
        ws.cell(row=row, column=6, value=float(case["guess"]))
        ws.cell(row=row, column=7, value=f"=RATE(A{row},B{row},C{row},D{row},E{row},F{row})")
        case["_cell"] = ("RATE", f"G{row}")

    ws = wb.create_sheet("XIRR")
    dates_offset = 10  # values never exceed 8 entries; dates start at col J
    guess_col = 20
    for row, case in enumerate(xirr_cases, start=1):
        values = case["values"]
        dates = case["dates"]
        for i, v in enumerate(values):
            ws.cell(row=row, column=1 + i, value=float(v))
        for i, d in enumerate(dates):
            # Excel serial date: write as a plain number, XIRR accepts serials.
            ws.cell(row=row, column=dates_offset + i, value=float(d))
        ws.cell(row=row, column=guess_col, value=float(case["guess"]))
        vlast = col_name(len(values))
        dlast = col_name(dates_offset + len(dates) - 1)
        formula_col = guess_col + 1
        ws.cell(row=row, column=formula_col,
                 value=(f"=XIRR(A{row}:{vlast}{row},"
                        f"{col_name(dates_offset)}{row}:{dlast}{row},"
                        f"{col_name(guess_col)}{row})"))
        case["_cell"] = ("XIRR", f"{col_name(formula_col)}{row}")

    wb.save(path)


# -----------------------------------------------------------------------------
# Comparison
# -----------------------------------------------------------------------------

def read_cell(cells, sheet_name_map, sheet, ref):
    internal_name = sheet_name_map.get(sheet, sheet)
    entry = cells.get((internal_name, ref))
    if entry is None:
        return None
    val = entry["val"]
    if isinstance(val, str) and val.startswith("#"):
        return "ERROR"
    return val


def values_close(a, b):
    if a is None or b is None:
        return a is b
    if a == "ERROR" or b == "ERROR":
        return a == b
    try:
        return math.isclose(float(a), float(b), rel_tol=1e-3, abs_tol=1e-3)
    except (TypeError, ValueError):
        return a == b


def to_comparable(candidate_result):
    return "ERROR" if candidate_result is None else candidate_result


def run_candidates(case):
    out = {}
    for variant in CANDIDATE_VARIANTS:
        if case["kind"] == "irr":
            r = candidate_irr(case["values"], case["guess"], variant)
        elif case["kind"] == "rate":
            r = candidate_rate(case["nper"], case["pmt"], case["pv"], case["fv"],
                                case["type"], case["guess"], variant)
        else:
            d0 = case["dates"][0]
            days = [d - d0 for d in case["dates"]]
            r = candidate_xirr(case["values"], days, case["guess"], variant)
        out[variant["name"]] = to_comparable(r)
    return out


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--driver", choices=["auto", "applescript", "win32com", "mock"], default="auto")
    parser.add_argument("--excel-path", default=None)
    parser.add_argument("--binary", default=os.path.join(PROJECT_ROOT, "target", "release", "visi"))
    parser.add_argument("--seed", type=int, default=1)
    parser.add_argument("--out-dir", default=os.path.join(PROJECT_ROOT, "fuzz_results", "financial_reverse_engineering"))
    parser.add_argument("--top", type=int, default=8, help="how many top candidate variants to print per function")
    parser.add_argument("--mismatches", type=int, default=15, help="how many Excel-vs-visi mismatches to print per function")
    args = parser.parse_args()

    rng = random.Random(args.seed)
    os.makedirs(args.out_dir, exist_ok=True)

    irr_cases = gen_irr_cases()
    rate_cases = gen_rate_cases(rng)
    xirr_cases = gen_xirr_cases(rng)
    all_cases = irr_cases + rate_cases + xirr_cases
    print(f"Generated {len(irr_cases)} IRR, {len(rate_cases)} RATE, {len(xirr_cases)} XIRR cases.")

    source_path = os.path.join(args.out_dir, "source.xlsx")
    excel_out_path = os.path.join(args.out_dir, "excel_out.xlsx")
    visi_out_path = os.path.join(args.out_dir, "visi_out.xlsx")

    print("Building workbook...")
    build_workbook(irr_cases, rate_cases, xirr_cases, source_path)

    print("Evaluating with visi...")
    visi = VisiDriver(args.binary)
    visi.run(source_path, visi_out_path)
    visi_cells = XLSXEvaluatedReader.read_evaluated_cells(visi_out_path)
    visi_sheet_map = map_sheet_names(visi_out_path)

    excel_cells = {}
    excel_sheet_map = {}
    if args.driver != "mock":
        print("Evaluating with real Excel (this opens/recalculates/saves via AppleScript or COM)...")
        excel = ExcelDriver(excel_path=args.excel_path, driver_type=args.driver)
        t0 = time.time()
        excel.run(source_path, excel_out_path)
        print(f"  Excel round trip took {time.time() - t0:.1f}s")
        excel_cells = XLSXEvaluatedReader.read_evaluated_cells(excel_out_path)
        excel_sheet_map = map_sheet_names(excel_out_path)
    else:
        print("Mock mode: skipping real Excel, only scoring visi + candidates against each other's presence.")

    print("Computing candidate variants for every case (this is pure Python, may take a bit)...")
    for case in all_cases:
        sheet, ref = case["_cell"]
        case["visi"] = read_cell(visi_cells, visi_sheet_map, sheet, ref)
        case["excel"] = read_cell(excel_cells, excel_sheet_map, sheet, ref) if excel_cells else None
        case["candidates"] = run_candidates(case)

    report = {"seed": args.seed, "have_excel": bool(excel_cells), "cases": []}
    for case in all_cases:
        entry = dict(case)
        entry.pop("_cell", None)
        report["cases"].append(entry)

    report_path = os.path.join(args.out_dir, "report.json")
    with open(report_path, "w") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"Full per-case report written to {report_path}")

    if not excel_cells:
        print("\nNo Excel ground truth collected (mock mode) -- skipping scoring against Excel.")
        print("Re-run with --excel-path \"/Applications/Microsoft Excel.app\" for the real comparison.")
        return

    for kind, cases in (("IRR", irr_cases), ("RATE", rate_cases), ("XIRR", xirr_cases)):
        print(f"\n{'=' * 78}\n{kind}: {len(cases)} cases\n{'=' * 78}")

        visi_matches = sum(1 for c in cases if values_close(c["visi"], c["excel"]))
        print(f"visi (current finance.rs) agrees with Excel on {visi_matches}/{len(cases)} "
              f"({100 * visi_matches / len(cases):.1f}%)")

        scored = []
        for variant in CANDIDATE_VARIANTS:
            matches = sum(1 for c in cases if values_close(c["candidates"][variant["name"]], c["excel"]))
            scored.append((matches, variant["name"]))
        scored.sort(key=lambda t: -t[0])

        print(f"\nTop {args.top} candidate variants by agreement with Excel:")
        for matches, name in scored[: args.top]:
            print(f"  {matches:4d}/{len(cases)} ({100 * matches / len(cases):5.1f}%)  {name}")

        best_matches, best_name = scored[0]
        print(f"\nWorst mismatches for best candidate ({best_name}) -- Excel vs. that candidate:")
        shown = 0
        for c in cases:
            cand_val = c["candidates"][best_name]
            if not values_close(cand_val, c["excel"]) and shown < args.mismatches:
                shown += 1
                desc = {k: v for k, v in c.items() if k not in ("candidates", "visi")}
                print(f"  excel={c['excel']!r:>14}  candidate={cand_val!r:>14}  visi={c['visi']!r:>10}  {desc}")

    print(f"\nDone. Inspect {report_path} for the full per-case, per-variant data.")


if __name__ == "__main__":
    main()
