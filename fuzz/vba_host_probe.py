#!/usr/bin/env python3
"""
What does real Excel's VBA host object model actually do?
=========================================================
Phase 2 of `docs/vba-macro-support.md` binds the interpreter to a workbook,
and issue #57 lists a dozen behaviours that have to be *measured* rather than
taken from documentation -- the same rule that already caught a hand-probe
getting a `Variant` rule backwards in Phase 1.

Unlike `vba_expr_probe.py` this runs against a workbook with data in it (a
small grid plus a formula, a text cell, a boolean, and two date-formatted
cells, one of them fractional), because most of the open questions are about
what a *cell* reads back as. And unlike the fuzzer it compares nothing: it
asks Excel and prints the answer, because there is no second implementation
to disagree with yet.

    source fuzz/venv/bin/activate
    python fuzz/vba_host_probe.py

Cases run in order in one Excel session against one workbook, so every case
that writes is at the end, after every case that reads. A read case moved
below a write case silently measures the wrong thing.

Two traps inherited from `vba_expr_probe.py`, both of which have already
produced published-and-wrong conclusions in this project:

* **`CStr(Null)` is itself error 94**, so a case whose result may be `Null`
  has to be written as `IsNull(...)`.
* **A compile error hangs the AppleScript bridge** and, unlike a runtime
  error, is not catchable by the `On Error` wrapper. A case that never
  returns is a compile error, not a hang.
"""

import argparse
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: source fuzz/venv/bin/activate && pip install -r fuzz/requirements.txt")

try:
    import visi_core
except ImportError:
    sys.exit(
        "the visi_core bindings are required: "
        "maturin develop -m visi-python/Cargo.toml --release"
    )

from fuzz_vba import HARNESS_TEMPLATE, ExcelDriver

# `ws` is Sheet1 and `wb` ThisWorkbook in every case, so a case is one line.
PREAMBLE = [
    "Dim ws As Worksheet, wb As Workbook",
    "Set wb = ThisWorkbook",
    'Set ws = wb.Worksheets("Sheet1")',
    "Dim v, s, c",
]

# (label, expression). A `::`-separated prefix is setup, `\n`-separated for
# statements that cannot share a line (`With`, `For Each`).
READ_CASES = [
    # --- what the objects call themselves
    "TypeName(wb)",
    "TypeName(ws)",
    'TypeName(ws.Range("A1"))',
    'TypeName(ws.Range("A1:B2"))',
    "TypeName(ws.Cells)",
    "TypeName(wb.Worksheets)",
    "TypeName(wb.Sheets)",
    # --- addresses and shape
    'ws.Range("A1").Address',
    'ws.Range("A1:B2").Address',
    'ws.Range("A1").Address(False, False)',
    'CStr(ws.Range("A1:B2").Count)',
    'TypeName(ws.Range("A1:B2").Count)',
    "CStr(ws.Cells.Count)",
    "TypeName(ws.Cells.Count)",
    'ws.Range("A1").Row & "," & ws.Range("B2").Column',
    'ws.Range("B2:C4").Row & "," & ws.Range("B2:C4").Column',
    'ws.Cells(2, 3).Address',
    'ws.Range("A1", "B2").Address',
    'ws.Range(ws.Cells(1, 1), ws.Cells(2, 2)).Address',
    'ws.Range("B2").Offset(1, 1).Address',
    'ws.Range("B2").Offset(-1, 0).Address',
    'ws.Range("A1").Offset(-1, 0).Address',
    'ws.Range("A1:B2").Resize(3, 1).Address',
    'ws.Range("A1").Resize(0, 1).Address',
    # --- For Each order over a range, and over the sheets
    'For Each c In ws.Range("A1:B2")\\n s = s & c.Address(False, False) & " "\\nNext :: s',
    "For Each c In wb.Worksheets\\n s = s & c.Name & \" \"\\nNext :: s",
    # --- reading values, including the date question
    'TypeName(ws.Range("A1").Value)',
    'CStr(ws.Range("A1").Value)',
    'TypeName(ws.Range("C1").Value)',
    'CStr(ws.Range("C1").Value)',
    'TypeName(ws.Range("C1").Value2)',
    'CStr(ws.Range("C1").Value2)',
    'TypeName(ws.Range("C2").Value)',
    'CStr(ws.Range("C2").Value)',
    'CStr(ws.Range("C2").Value2)',
    'TypeName(ws.Range("C3").Value)',
    'CStr(ws.Range("C3").Value)',
    'ws.Range("C1").Text',
    'ws.Range("C2").Text',
    'ws.Range("A1").Text',
    'ws.Range("D1").Formula',
    'TypeName(ws.Range("D1").Value)',
    'CStr(ws.Range("D1").Value)',
    'ws.Range("D2").Formula',
    'TypeName(ws.Range("D2").Value)',
    'TypeName(ws.Range("E1").Value)',
    'TypeName(ws.Range("E2").Value)',
    'CStr(ws.Range("E2").Value)',
    'ws.Range("E2").Formula',
    'ws.Range("C1").NumberFormat',
    # C3 carries a date-and-time format but reads back as a plain Double,
    # unlike C1/C2 -- these two say whether Excel actually kept the format.
    'ws.Range("C3").NumberFormat',
    'ws.Range("C3").Text',
    # An error *in a cell* reads back as an error Variant, the same subtype
    # `Application.VLookup` returns on failure.
    'TypeName(ws.Range("F1").Value)',
    'CStr(CLng(ws.Range("F1").Value))',
    'ws.Range("F1").Text',
    'CStr(IsError(ws.Range("F1").Value))',
    'ws.Range("F1").Formula',
    'CStr(Application.WorksheetFunction.Sum(ws.Range("F1")))',
    # Unqualified, i.e. against the active sheet.
    'Range("A1").Address & "/" & Cells(2, 2).Address',
    'ws.Cells.Address',
    'TypeName(wb.Worksheets(1))',
    'wb.Worksheets("Sheet1").Range("B3").Address',
    # --- a multi-cell range read into a scalar
    # Bracketed because the bare form came back as `V()`, which is not a
    # thing `TypeName` returns -- worth knowing whether that is Excel or the
    # bridge eating characters.
    'v = ws.Range("A1:A3").Value :: "[" & TypeName(v) & "]"',
    'v = ws.Range("A1:A3").Value2 :: "[" & TypeName(v) & "]"',
    'v = ws.Range("A1:A3").Value :: CStr(Application.WorksheetFunction.Sum(v))',
    'v = ws.Range("A1:A3").Value :: CStr(UBound(v, 1)) & "," & CStr(UBound(v, 2))',
    'v = ws.Range("A1:A3").Value :: TypeName(v)',
    'v = ws.Range("A1:A3").Value :: CStr(v(2, 1))',
    'v = ws.Range("A1:A3") :: TypeName(v)',
    'v = ws.Range("A1:B2").Value :: CStr(v)',
    'v = ws.Range("A1").Value :: TypeName(v)',
    # --- errors
    'ws.Range("nope!!").Address',
    'ws.Range("").Address',
    'wb.Worksheets("nope").Name',
    'wb.Worksheets(5).Name',
    # `ws.Range("A1").Nonsense` and `Application.WorksheetFunction.Nonsense(1)`
    # are deliberately absent: `Range` and `WorksheetFunction` are early-bound,
    # so an unknown member is a *compile* error, which hangs the bridge and
    # cannot be trapped. Measured the hard way -- they took two whole batches
    # down with them.
    # --- Nothing and Is
    "Dim r As Range :: TypeName(r)",
    "Dim r As Range :: CStr(r Is Nothing)",
    'CStr(ws.Range("A1") Is ws.Range("A1"))',
    "CStr(ws Is wb.Worksheets(1))",
    "CStr(ws Is wb.Worksheets(2))",
    'CStr(ws.Range("A1") Is ws.Range("A2"))',
    "CStr(ws Is Nothing)",
    # --- With
    'With ws.Range("A2")\\n s = CStr(.Value) & "/" & .Address\\nEnd With :: s',
    # --- WorksheetFunction vs Application
    'CStr(Application.WorksheetFunction.Sum(ws.Range("A1:A3")))',
    'TypeName(Application.WorksheetFunction.Sum(ws.Range("A1:A3")))',
    'CStr(Application.WorksheetFunction.Sum(ws.Range("A1:E3")))',
    'CStr(Application.WorksheetFunction.Sum("1", 2))',
    'CStr(Application.WorksheetFunction.Sum(ws.Range("D2")))',
    'CStr(Application.WorksheetFunction.Sum(ws.Range("E1")))',
    'CStr(Application.WorksheetFunction.Count(ws.Range("A1:E3")))',
    'CStr(Application.WorksheetFunction.CountA(ws.Range("A1:E3")))',
    'CStr(Application.WorksheetFunction.Sum(ws.Range("C1")))',
    'Application.WorksheetFunction.Text(ws.Range("C1").Value, "yyyy-mm-dd")',
    'CStr(Application.WorksheetFunction.VLookup("zzz", ws.Range("A1:B3"), 2, False))',
    'TypeName(Application.VLookup("zzz", ws.Range("A1:B3"), 2, False))',
    'CStr(IsError(Application.VLookup("zzz", ws.Range("A1:B3"), 2, False)))',
    'v = Application.VLookup("zzz", ws.Range("A1:B3"), 2, False) :: CStr(CLng(v))',
    'CStr(Application.Sum(ws.Range("A1:A3")))',
    'CStr(Application.WorksheetFunction.Sum(ws.Range("A1"), 5))',
    # --- how an error Variant and a Date render, which the fuzz harness
    # compares through CStr and so must be exactly right
    'CStr(CVErr(2042))',
    'TypeName(CVErr(2042))',
    'CStr(IsError(CVErr(2042)))',
    'v = CVErr(2042) :: CStr(v & "")',
    'v = CVErr(2042) :: CStr(v + 1)',
    'v = CVErr(2042) :: CStr(v = 1)',
    'v = ws.Range("F1").Value :: CStr(v + 1)',
    'CStr(#6/22/2026#)',
    'TypeName(#6/22/2026#)',
    'CStr(#6/22/2026 12:00:00 PM#)',
    'CStr(CDate(0.5))',
    'CStr(CDate(46195))',
    'TypeName(#6/22/2026# + 1)',
    'CStr(#6/22/2026# + 1)',
    'CStr(#6/22/2026# - #6/21/2026#)',
    # object identity survives copying, which is what makes `Is` on a Range
    # mean something despite two fresh `Range()` calls never matching
    "Dim r As Range :: Set r = ws.Range(\"A1\") :: CStr(r Is r)",
    "Dim r As Range, q As Range :: Set r = ws.Range(\"A1\") :: Set q = r :: CStr(q Is r)",
    'Dim r As Range :: Set r = ws.Range("A1") :: CStr(r Is ws.Range("A1"))',
    # what WorksheetFunction does with an argument it cannot use
    'CStr(Application.WorksheetFunction.Sum(ws.Range("A1"), "x"))',
    'TypeName(Application.WorksheetFunction.Sum(ws.Range("A1:A3"), 1))',
    # --- names
    "wb.Name",
    "wb.Worksheets(1).Name",
    "CStr(wb.Worksheets.Count)",
    'CStr(wb.Worksheets("SHEET1").Name)',
]

# Everything that mutates. Ordered after every read, and each writes to its
# own scratch cell so one case cannot silently set up the next.
WRITE_CASES = [
    'ws.Range("G1").Value = 5 :: CStr(ws.Range("G1").Value)',
    'ws.Range("G2") = 7 :: CStr(ws.Range("G2").Value)',
    'ws.Range("G3").Formula = "=G1*2" :: CStr(ws.Range("G3").Value)',
    'ws.Range("G4").Value = "=G1*3" :: ws.Range("G4").Formula & "|" & CStr(ws.Range("G4").Value)',
    'ws.Range("G5").Value = #6/22/2026# :: TypeName(ws.Range("G5").Value) & "|" & ws.Range("G5").Text & "|" & ws.Range("G5").NumberFormat',
    'ws.Range("G6").Value = "6/22/2026" :: TypeName(ws.Range("G6").Value) & "|" & CStr(ws.Range("G6").Value2) & "|" & ws.Range("G6").NumberFormat',
    'ws.Range("G7").Value = "hello" :: TypeName(ws.Range("G7").Value) & "|" & ws.Range("G7").Formula',
    'ws.Range("G8:H9").Value = 3 :: CStr(ws.Range("H9").Value)',
    'ws.Range("A1").Value = 5 :: CStr(ws.Range("D1").Value)',
    'wb.Worksheets(2).Name = "Renamed" :: wb.Worksheets(2).Name',
    'ws.Range("I1").Value = ws.Range("A1:A3").Value :: TypeName(ws.Range("I1").Value) & "|" & CStr(ws.Range("I1").Value)',
]

CASES = READ_CASES + WRITE_CASES


def parse_case(text):
    setup, sep, expr = text.rpartition("::")
    if not sep:
        return [], text.strip()
    return [s.strip() for s in setup.split("\\n") if s.strip()], expr.strip()


def build_module(cases):
    parts = ['Attribute VB_Name = "H"']
    for i, (setup, expr) in enumerate(cases, start=1):
        body = "\n".join(f"    {s}" for s in PREAMBLE + setup)
        parts.append(f"Private Function Gen{i}()\n{body}\n    Gen{i} = {expr}\nEnd Function")
        parts.append(HARNESS_TEMPLATE.format(i=i))
    return "\n\n".join(parts) + "\n"


def build_workbook(path):
    """The grid every case reads.

    `C1`/`C2`/`C3` are the date question: an integral serial, a fractional
    one, and one carrying a time format. `D1` is a formula, `D2` text that
    looks like nothing else, `E1` a boolean and `E2` genuinely empty.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r, (a, b) in enumerate([(1, 10), (2, 20), (3, 30)], start=1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
    # 46195 is 2026-06-22 in the 1900 system.
    ws["C1"] = 46195
    ws["C1"].number_format = "m/d/yy"
    ws["C2"] = 46195.5
    ws["C2"].number_format = "m/d/yy"
    ws["C3"] = 46195.25
    ws["C3"].number_format = "m/d/yy h:mm"
    ws["D1"] = "=A1*2"
    ws["D2"] = "hi"
    ws["F1"] = "=1/0"
    ws["E1"] = True
    wb.create_sheet("Sheet2")
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--excel-path")
    ap.add_argument("--driver", choices=["auto", "applescript", "mock"], default="auto")
    ap.add_argument("--timeout", type=int, default=180)
    ap.add_argument("--batch", type=int, default=12,
                    help="Read cases per Excel round trip. Small on purpose: a "
                         "compile error costs the whole batch.")
    ap.add_argument("--keep", action="store_true", help="Keep the generated .xlsm")
    args = ap.parse_args()

    cases = [parse_case(c) for c in CASES]
    source = build_module(cases)

    driver = ExcelDriver(args.excel_path, args.driver, args.timeout)
    workdir = tempfile.mkdtemp(prefix="vba_host_probe_")
    base = os.path.join(workdir, "base.xlsx")
    xlsm = os.path.join(workdir, "probe.xlsm")
    build_workbook(base)
    wbk = visi_core.Workbook.load(base)
    wbk.add_macro("H", source)
    wbk.save(xlsm)
    if args.keep:
        print(f"workbook: {xlsm}\n", file=sys.stderr)

    results = {}
    # A round trip re-opens the workbook, so a batch boundary discards every
    # write the batch before it made. The reads are chunked small (one compile
    # error then loses a chunk rather than the run) and the writes go as a
    # single chunk, since several of them read back what an earlier one wrote.
    reads = list(range(1, len(READ_CASES) + 1))
    writes = list(range(len(READ_CASES) + 1, len(cases) + 1))
    for start in range(0, len(reads), args.batch):
        chunk = reads[start:start + args.batch]
        got = driver.run_batch(xlsm, chunk)
        if not got:
            print(f"cases {chunk[0]}-{chunk[-1]}: Excel returned nothing "
                  "(a compile error in one of them).", file=sys.stderr)
        results.update(got)
    got = driver.run_batch(xlsm, writes)
    if not got:
        print(f"cases {writes[0]}-{writes[-1]}: Excel returned nothing.",
              file=sys.stderr)
    results.update(got)

    width = min(max(len(c) for c in CASES), 78)
    for i, case in enumerate(CASES, start=1):
        print(f"{case[:width].ljust(width)}  {results.get(i, '-')}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
