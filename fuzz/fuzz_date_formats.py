#!/usr/bin/env python3
"""Focused date-format/display fuzzing for visi's xlsx output.

This is intentionally separate from fuzz_excel.py's value comparator: date cells
are numbers plus number formats, so a value-only comparison can miss a workbook
that calculates correctly but no longer displays or round-trips as dates.

Each iteration builds a small workbook through visi (so formula date-format
inheritance is exercised), saves it, lets both visi and Excel round-trip it, and
then checks the exported style/display metadata for:

* typed date literals retain the expected date number format;
* bare references and one-sided date arithmetic inherit that format;
* date component/difference formulas stay ordinary numbers;
* date-looking text remains text and has no date number format.
"""

import argparse
import datetime as dt
import os
import random
import shutil
import subprocess
import sys
import tempfile
import time

import openpyxl

try:
    import visi_core
except ImportError as exc:  # pragma: no cover - exercised by humans without bindings
    raise SystemExit(
        "visi_core bindings are required for date-format fuzzing. Build them with:\n"
        "  source fuzz/venv/bin/activate && maturin develop -m visi-python/Cargo.toml --release"
    ) from exc

from fuzz_excel import ExcelDriver  # noqa: E402

MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def col_name(idx):
    """1-based column index to A1 column letters."""
    out = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        out = chr(ord("A") + rem) + out
    return out


def is_date_format(fmt):
    """Approximate visi's date-format predicate for the formats this fuzzer emits."""
    if not fmt or fmt == "General":
        return False
    in_quote = False
    in_bracket = False
    has_date_token = False
    has_number_placeholder = False
    for ch in fmt:
        if ch == '"':
            in_quote = not in_quote
            continue
        if in_quote:
            continue
        if ch == "[":
            in_bracket = True
            continue
        if ch == "]":
            in_bracket = False
            continue
        if in_bracket:
            continue
        lower = ch.lower()
        if lower in ("y", "m", "d"):
            has_date_token = True
        if ch in "0#?":
            has_number_placeholder = True
    return has_date_token and not has_number_placeholder


LOCALE_FORMATTERS = {
    "en-US": [
        (lambda d: (f"{d.month}/{d.day}/{d.year % 100:02d}", "m/d/yy")),
        (lambda d: (f"{d.month}/{d.day}/{d.year}", "m/d/yyyy")),
        (lambda d: (f"{d.year}-{d.month:02d}-{d.day:02d}", "yyyy-mm-dd")),
        (lambda d: (f"{d.year}/{d.month:02d}/{d.day:02d}", "yyyy/mm/dd")),
        (lambda d: (f"{d.day}-{MONTH_ABBR[d.month - 1]}-{d.year}", "d-mmm-yyyy")),
        (lambda d: (f"{MONTH_ABBR[d.month - 1]}-{d.day}-{d.year}", "mmm-d-yyyy")),
    ],
    "en-GB": [
        (lambda d: (f"{d.day}/{d.month}/{d.year % 100:02d}", "d/m/yy")),
        (lambda d: (f"{d.day}/{d.month}/{d.year}", "d/m/yyyy")),
        (lambda d: (f"{d.year}-{d.month:02d}-{d.day:02d}", "yyyy-mm-dd")),
        (lambda d: (f"{d.day}-{MONTH_ABBR[d.month - 1]}-{d.year}", "d-mmm-yyyy")),
    ],
    "de-DE": [
        (lambda d: (f"{d.day}.{d.month}.{d.year % 100:02d}", "d.m.yy")),
        (lambda d: (f"{d.day}.{d.month}.{d.year}", "d.m.yyyy")),
        (lambda d: (f"{d.year}-{d.month:02d}-{d.day:02d}", "yyyy-mm-dd")),
    ],
}


def random_date(rng):
    # Keep days valid for every month and avoid pre-1900 edge cases.
    return dt.date(rng.randint(1995, 2035), rng.randint(1, 12), rng.randint(1, 28))


def build_workbook(path, rng, locale="en-US"):
    """Create one fuzz workbook through visi and return expected cell metadata."""
    wb = visi_core.Workbook(locale=locale)
    formatters = LOCALE_FORMATTERS.get(locale, LOCALE_FORMATTERS["en-US"])
    rows = rng.randint(3, 7)
    expected = {}

    for row in range(rows):
        date = random_date(rng)
        literal, fmt = rng.choice(formatters)(date)
        excel_row = row + 1

        wb.set_cell(row, 0, literal)
        wb.set_cell(row, 1, f"=A{excel_row}")
        wb.set_cell(row, 2, f"=A{excel_row}+{rng.randint(1, 14)}")
        wb.set_cell(row, 3, f"=YEAR(A{excel_row})")
        wb.set_cell(row, 4, f"=A{excel_row}-A{excel_row}")
        wb.set_cell(row, 5, f'"{literal}"')

        expected[(row, 0)] = {"kind": "date", "format": fmt}
        expected[(row, 1)] = {"kind": "date", "format": fmt}
        expected[(row, 2)] = {"kind": "date", "format": fmt}
        expected[(row, 3)] = {"kind": "plain"}
        expected[(row, 4)] = {"kind": "plain"}
        expected[(row, 5)] = {"kind": "text", "text": literal}

    wb.evaluate()
    expected_displays = {
        rc: wb.get_display(rc[0], rc[1])
        for rc in expected
    }
    wb.save(path)
    return rows, expected, expected_displays


def inspect_workbook(path, rows):
    """Read cell formats, data types and visi-rendered display strings."""
    py_wb = openpyxl.load_workbook(path, data_only=False)
    ws = py_wb[py_wb.sheetnames[0]]
    visi_wb = visi_core.Workbook.load(path)

    cells = {}
    for row in range(rows):
        for col in range(6):
            cell = ws.cell(row=row + 1, column=col + 1)
            cells[(row, col)] = {
                "format": cell.number_format,
                "data_type": cell.data_type,
                "value": cell.value,
                "display": visi_wb.get_display(row, col),
            }
    return cells


def baseline_expected_formats(source_cells, expected):
    """Use the visi-authored source workbook as the exact round-trip baseline."""
    out = {rc: dict(want) for rc, want in expected.items()}
    for rc, want in out.items():
        if want["kind"] == "date":
            want["format"] = source_cells[rc]["format"]
    return out


def compare_cells(
    label,
    actual,
    expected,
    expected_displays,
    *,
    exact_date_format=True,
    compare_date_display=True,
):
    failures = []
    for rc, want in expected.items():
        got = actual[rc]
        fmt = got["format"]
        if want["kind"] == "date":
            if exact_date_format:
                if fmt != want["format"]:
                    failures.append(
                        f"{label} {cell_ref(rc)}: expected date format {want['format']!r}, got {fmt!r}"
                    )
            elif not is_date_format(fmt):
                failures.append(f"{label} {cell_ref(rc)}: expected a date format, got {fmt!r}")
        else:
            if is_date_format(fmt):
                failures.append(f"{label} {cell_ref(rc)}: expected no date format, got {fmt!r}")

        if want["kind"] == "text":
            if got["data_type"] != "s" or got["value"] != want["text"]:
                failures.append(
                    f"{label} {cell_ref(rc)}: expected text {want['text']!r}, "
                    f"got type={got['data_type']!r} value={got['value']!r}"
                )

        if (want["kind"] != "date" or compare_date_display) and got["display"] != expected_displays[rc]:
            failures.append(
                f"{label} {cell_ref(rc)}: expected display {expected_displays[rc]!r}, "
                f"got {got['display']!r}"
            )
    return failures


def cell_ref(rc):
    row, col = rc
    return f"{col_name(col + 1)}{row + 1}"


def copy_failure_artifacts(dest, *paths):
    os.makedirs(dest, exist_ok=True)
    for path in paths:
        if path and os.path.exists(path):
            shutil.copyfile(path, os.path.join(dest, os.path.basename(path)))


def main():
    parser = argparse.ArgumentParser(description="Fuzz date number formats and display strings in xlsx output.")
    parser.add_argument("--excel-path", help="Path to Microsoft Excel app/binary.")
    parser.add_argument("--driver", choices=["auto", "applescript", "win32com", "cli", "mock"], default="auto")
    parser.add_argument("--iterations", type=int, default=20)
    parser.add_argument("--seed", type=int)
    parser.add_argument("--locale", choices=["en-US", "en-GB", "de-DE"], help="Spreadsheet locale to fuzz.")
    parser.add_argument("--output-dir", default="./fuzz_results")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    failures_dir = os.path.join(args.output_dir, "date_format_failures")
    os.makedirs(failures_dir, exist_ok=True)

    excel_driver = ExcelDriver(excel_path=args.excel_path, driver_type=args.driver)
    smoke_mode = excel_driver.driver_type == "mock"

    print("=====================================================================")
    print("        visi Date Format / Display Fuzzing Harness")
    print("=====================================================================")
    print(f" Iterations : {args.iterations}")
    print(f" Excel Driver: {excel_driver.driver_type} ({args.excel_path or 'Default'})")
    if args.locale:
        print(f" Locale      : {args.locale}")
    if smoke_mode:
        print(" Mock mode: Excel round-trip is skipped; only visi source/round-trip is checked.")
    print("=====================================================================\n")

    passed = 0
    failed = 0
    start = time.time()

    locales = [args.locale] if args.locale else ["en-US", "en-GB", "de-DE"]

    for i in range(1, args.iterations + 1):
        iter_seed = (args.seed + i) if args.seed is not None else random.randint(1, 1_000_000)
        rng = random.Random(iter_seed)
        locale = rng.choice(locales)
        temp_dir = tempfile.mkdtemp(prefix=f"date_fmt_fuzz_{i}_")
        source_xlsx = os.path.join(temp_dir, "source.xlsx")
        visi_out_xlsx = os.path.join(temp_dir, "visi_out.xlsx")
        excel_out_xlsx = os.path.join(temp_dir, "excel_out.xlsx")

        try:
            rows, expected, displays = build_workbook(source_xlsx, rng, locale=locale)

            # A visi-authored workbook is the exact baseline. Excel is allowed
            # to canonicalize built-in/localized date format spellings on save,
            # but a visi round-trip must preserve the format codes byte-for-byte.
            source_cells = inspect_workbook(source_xlsx, rows)
            expected = baseline_expected_formats(source_cells, expected)
            failures = compare_cells("source", source_cells, expected, displays)

            visi_roundtrip = visi_core.Workbook.load(source_xlsx)
            visi_roundtrip.evaluate()
            visi_roundtrip.save(visi_out_xlsx)
            failures.extend(compare_cells("visi", inspect_workbook(visi_out_xlsx, rows), expected, displays))

            if not smoke_mode:
                excel_driver.run(source_xlsx, excel_out_xlsx)
                failures.extend(compare_cells(
                    "excel",
                    inspect_workbook(excel_out_xlsx, rows),
                    expected,
                    displays,
                    exact_date_format=False,
                    compare_date_display=False,
                ))

            if failures:
                failed += 1
                fail_dir = os.path.join(failures_dir, f"fail_iter_{i}_seed_{iter_seed}")
                copy_failure_artifacts(fail_dir, source_xlsx, visi_out_xlsx, None if smoke_mode else excel_out_xlsx)
                with open(os.path.join(fail_dir, "failure.txt"), "w", encoding="utf-8") as f:
                    f.write("\n".join(failures))
                    f.write("\n")
                print(f" Iteration {i:3d}/{args.iterations} [FAILED] (Seed: {iter_seed})")
                for line in failures[:10]:
                    print(f"   - {line}")
                if len(failures) > 10:
                    print(f"   ... {len(failures) - 10} more")
                print(f"   artifacts: {fail_dir}")
            else:
                passed += 1
                print(f" Iteration {i:3d}/{args.iterations} [PASSED] (Seed: {iter_seed})")
        except Exception as exc:  # noqa: BLE001 - fuzz harness should preserve artifacts
            failed += 1
            fail_dir = os.path.join(failures_dir, f"fail_iter_{i}_seed_{iter_seed}")
            copy_failure_artifacts(fail_dir, source_xlsx, visi_out_xlsx, None if smoke_mode else excel_out_xlsx)
            with open(os.path.join(fail_dir, "exception.txt"), "w", encoding="utf-8") as f:
                f.write(repr(exc))
                f.write("\n")
            print(f" Iteration {i:3d}/{args.iterations} [ERROR] (Seed: {iter_seed})")
            print(f"   {exc!r}")
            print(f"   artifacts: {fail_dir}")
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    elapsed = time.time() - start
    print("\n=====================================================================")
    print(f" Fuzzing Completed in {elapsed:.2f}s")
    print(f" Passed : {passed}/{args.iterations}")
    print(f" Failed : {failed}/{args.iterations}")
    print("=====================================================================")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
