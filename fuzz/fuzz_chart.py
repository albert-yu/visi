#!/usr/bin/env python3
"""
visi vs. Microsoft Excel Differential Fuzzing: Charts
=========================================================
Generates a random source data grid, builds a matching chart in both `visi`
(via the `visi chart add`/`visi chart edit` CLI) and real Microsoft Excel (by
driving Excel's own chart object model via AppleScript/COM automation --
like pivot tables, Excel must *construct* a chart object; there's no
"write XML, then recalculate" shortcut the way plain formulas work, which is
why this lives in its own script rather than reusing `fuzz_excel.py`'s
"generate formulas -> recalculate" flow), then compares the two engines'
resulting chart structure (type, category/value ranges, title, axis labels,
legend) via `chart_xlsx_reader.read_charts` -- `fuzz_excel.py`'s
`XLSXEvaluatedReader`/`DifferentialComparator` only understand cell values
and are structurally blind to charts.

Every fuzz iteration exercises `visi chart add` followed by `visi chart
edit` (not just add), so the edit path gets differential coverage against
real Excel too.

IMPORTANT -- unlike pivot tables, chart creation via AppleScript works
*natively* against real Excel; no VBA-macro-template workaround is needed
here. `make new chart object at end of chart objects of <sheet>` mirrors
pivot's `make new pivot cache`'s "-50 Parameter error", but the working
syntax is `make new chart object at <sheet>` (skip "at end of chart objects
of" entirely) -- confirmed via a manual spike, not documented in Excel.sdef
or any Microsoft reference found. Once the chart object exists, Excel's
`chart wizard` command (the AppleScript exposure of VBA's
`Chart.ChartWizard`) reliably sets source data, chart type, title, axis
titles, and legend in one call, for every chart type visi supports except
Pie -- see ExcelChartDriver's docstring for that caveat.

Usage:
    python3 fuzz/fuzz_chart.py --driver mock --iterations 5
    python3 fuzz/fuzz_chart.py --excel-path "/Applications/Microsoft Excel.app" --iterations 10 --seed 1
"""

import argparse
import json
import os
import random
import shutil
import subprocess
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from chart_xlsx_reader import read_charts  # noqa: E402
from fuzz_excel import SMOKE_BANNER, smoke_check  # noqa: E402

# -----------------------------------------------------------------------------
# 1. Source workbook + chart configuration generator
# -----------------------------------------------------------------------------


class ChartFuzzGenerator:
    """Generates a random source data grid plus a matching chart
    configuration (as plain dicts, not XML) that both `VisiChartDriver` and
    `ExcelChartDriver` build a real chart from.

    Each iteration produces two configs: `add_config` (what `visi chart add`
    -- and, on the Excel side, the initial `chart wizard` call -- creates
    the chart with) and `edit_config` (the *final* target state after a
    follow-up `visi chart edit` call). `edit_config` always specifies every
    field explicitly (never "leave unchanged") since the fuzzer's job is to
    check the resulting xlsx structure is correct, not to re-test
    `edit_chart`'s tri-state leave/set/clear semantics -- that's already
    covered by the Rust-level unit test in `visi/tests/cli_tests.rs`. On the
    Excel side this final state is reached with a single `chart wizard`
    call rather than mimicking a two-step history, since Excel has no
    separate "edit" concept to exercise -- only the resulting xlsx structure
    is compared.
    """

    CHART_TYPES = ["column", "bar", "line", "pie", "scatter", "area"]
    # Pie and Area are excluded here deliberately, not as an oversight:
    # Excel's AppleScript `chart wizard` command rejects `category title`/
    # `value title` parameters for Pie charts outright (Parameter error -50
    # -- pie charts have no axes to title), and Area chart axis titles were
    # found not to read back through openpyxl's `_charts` in a manual spike
    # against this Excel/openpyxl version (unexplained further -- see
    # fuzz/README.md's Chart Fuzzing section). Restricting xlabel/ylabel
    # generation to the four types confirmed to round-trip cleanly keeps the
    # fuzzer meaningful instead of chasing an openpyxl reader quirk.
    AXIS_LABEL_TYPES = ["column", "bar", "line", "scatter"]

    TITLES = ["Sales", "Revenue by Region", "Q3 Results", None]
    AXIS_LABELS = ["Category", "Amount", "Units", None]

    def __init__(self, seed=None):
        if seed is not None:
            random.seed(seed)

    def _random_config(self, chart_type=None):
        chart_type = chart_type or random.choice(self.CHART_TYPES)
        allows_axis_labels = chart_type in self.AXIS_LABEL_TYPES
        return {
            "chart_type": chart_type,
            "title": random.choice(self.TITLES),
            "xlabel": random.choice(self.AXIS_LABELS) if allows_axis_labels else None,
            "ylabel": random.choice(self.AXIS_LABELS) if allows_axis_labels else None,
            "show_legend": random.choice([True, False]),
        }

    def generate(self, source_path, num_rows):
        """Writes a small source workbook (one category column, one numeric
        column) to `source_path` via openpyxl, and returns
        `(range_str, add_config, edit_config)`.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i in range(1, num_rows + 1):
            ws.cell(row=i, column=1, value=f"Cat{i}")
            ws.cell(row=i, column=2, value=random.randint(1, 1000))
        wb.save(source_path)

        range_str = f"Sheet1!A1:B{num_rows}"
        add_config = self._random_config()
        # Pick a distinct chart type for the edit step often enough to
        # meaningfully exercise `chart edit --chart-type`, not just
        # relabeling the same type.
        other_types = [t for t in self.CHART_TYPES if t != add_config["chart_type"]]
        edit_type = random.choice(other_types) if random.random() < 0.7 else add_config["chart_type"]
        edit_config = self._random_config(chart_type=edit_type)
        return range_str, add_config, edit_config


# -----------------------------------------------------------------------------
# 2. Drivers
# -----------------------------------------------------------------------------


# `VisiChartDriver` moved to visi_driver.py, which drives visi either through
# the in-process `visi_core` bindings or through the `visi chart` CLI.
from visi_driver import (  # noqa: E402,F401
    VisiChartDriver,
    add_backend_arg,
    bindings_hint,
)


class ExcelChartDriver:
    """Drives Microsoft Excel's own chart object model to build a chart
    matching `edit_config` (the final target state -- see
    `ChartFuzzGenerator`'s docstring for why Excel doesn't need to mimic the
    add-then-edit history).

    The win32com path (Windows) uses the standard VBA object model
    (`ChartObjects.Add` + `Chart.SetSourceData`/`.ChartType`/etc.) directly.

    The AppleScript path (macOS) creates the chart with `make new chart
    object at <sheet>` -- NOT `make new chart object at end of chart
    objects of <sheet>`, which fails with the same generic "Parameter error
    (-50)" `fuzz_pivot.py`'s `make new pivot cache` hit; this alternate form
    was found to work via manual trial (undocumented in Excel.sdef) -- then
    configures it with the `chart wizard` command, which reliably sets
    source data/type/title/axis titles/legend in one call for every chart
    type except Pie: passing `category title`/`value title` to `chart
    wizard` for a Pie gallery type also raises -50 (pie charts have no
    axes), so those parameters are omitted whenever the target type is Pie.
    """

    def __init__(self, excel_path=None, driver_type="auto"):
        self.excel_path = excel_path
        self.driver_type = driver_type
        if driver_type == "auto":
            if sys.platform == "darwin":
                self.driver_type = "applescript"
            elif sys.platform == "win32":
                self.driver_type = "win32com"
            else:
                self.driver_type = "mock"

    def run(self, source_file, range_str, edit_config, output_file):
        shutil.copyfile(source_file, output_file)
        abs_output = os.path.abspath(output_file)

        if self.driver_type == "mock":
            print("[ExcelChartDriver Warning] Running in mock mode (Excel not invoked).")
            return
        elif self.driver_type == "applescript":
            self._run_applescript(abs_output, range_str, edit_config)
        elif self.driver_type == "win32com":
            self._run_win32com(abs_output, range_str, edit_config)
        else:
            raise RuntimeError(f"Unsupported chart driver type: {self.driver_type}")

    # -- AppleScript (macOS) --------------------------------------------

    GALLERY_NAMES = {
        "column": "column clustered",
        "bar": "bar clustered",
        "line": "line chart",
        "pie": "pie chart",
        "scatter": "xyscatter",
        "area": "area chart",
    }

    def _applescript_str(self, s):
        escaped = s.replace("\\", "\\\\").replace('"', '\\"')
        return f'"{escaped}"'

    def _build_applescript(self, abs_output, range_str, config):
        app_name = self.excel_path if self.excel_path else "Microsoft Excel"
        if app_name.endswith(".app"):
            app_name = os.path.splitext(os.path.basename(app_name))[0]

        gallery = self.GALLERY_NAMES[config["chart_type"]]
        # `range_str` looks like "Sheet1!A1:B10"; `chart wizard`'s `source`
        # parameter wants a bare range object relative to the active sheet.
        cell_range = range_str.split("!", 1)[1]

        wizard_parts = [f'source (range {self._applescript_str(cell_range)} of ws)', f"gallery {gallery}"]
        wizard_parts.append(f'has legend {"true" if config["show_legend"] else "false"}')
        if config["title"]:
            wizard_parts.append(f'title {self._applescript_str(config["title"])}')
        # Pie charts have no axes -- passing category/value title raises a
        # generic Parameter error (-50), so these are omitted entirely for
        # Pie regardless of what the (always-None, per the generator) config
        # says.
        if config["chart_type"] != "pie":
            if config["xlabel"]:
                wizard_parts.append(f'category title {self._applescript_str(config["xlabel"])}')
            if config["ylabel"]:
                wizard_parts.append(f'value title {self._applescript_str(config["ylabel"])}')

        lines = [
            f'tell application "{app_name}"',
            "    set display alerts to false",
            "    try",
            "        close workbooks saving no",
            "    end try",
            "    try",
            f'        set targetFile to POSIX file "{abs_output}"',
            "        open targetFile",
            "        set wb to active workbook",
            "        set ws to active sheet of wb",
            "        set co to make new chart object at ws",
            "        set ch to chart of co",
            "        chart wizard ch " + " ".join(wizard_parts),
            "        save wb",
            "        close wb saving no",
            "    on error errText number errNum",
            "        try",
            "            close workbooks saving no",
            "        end try",
            "        error errText number errNum",
            "    end try",
            "end tell",
        ]
        return "\n".join(lines)

    def _run_applescript(self, abs_output, range_str, config):
        script = self._build_applescript(abs_output, range_str, config)
        res = None
        for attempt in range(5):
            time.sleep(0.5)
            try:
                res = subprocess.run(
                    ["osascript", "-e", script],
                    stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=20,
                )
                if res.returncode == 0:
                    break
            except subprocess.TimeoutExpired:
                subprocess.run(["killall", "Microsoft Excel"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                time.sleep(1.0)
        if res is not None and res.returncode != 0:
            raise RuntimeError(f"Excel chart AppleScript failed:\nSTDERR: {res.stderr}\nScript:\n{script}")

    # -- win32com (Windows) -----------------------------------------------

    WIN32COM_CHART_TYPE = {
        "column": "xlColumnClustered",
        "bar": "xlBarClustered",
        "line": "xlLine",
        "pie": "xlPie",
        "scatter": "xlXYScatter",
        "area": "xlArea",
    }

    def _run_win32com(self, abs_output, range_str, config):
        try:
            import win32com.client
            from win32com.client import constants as c
        except ImportError:
            raise RuntimeError("pywin32 (win32com) is required for Excel automation on Windows.")

        cell_range = range_str.split("!", 1)[1]

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(abs_output)
            ws = wb.Sheets("Sheet1")
            co = ws.ChartObjects().Add(Left=200, Top=0, Width=300, Height=200)
            chart = co.Chart
            chart.SetSourceData(Source=ws.Range(cell_range))
            chart.ChartType = getattr(c, self.WIN32COM_CHART_TYPE[config["chart_type"]])
            chart.HasLegend = config["show_legend"]
            if config["title"]:
                chart.HasTitle = True
                chart.ChartTitle.Text = config["title"]
            else:
                chart.HasTitle = False
            # Pie charts raise an error on `chart.Axes(...)` (no axes
            # exist), matching the AppleScript `chart wizard` gap above.
            if config["chart_type"] != "pie":
                if config["xlabel"]:
                    cat_axis = chart.Axes(c.xlCategory)
                    cat_axis.HasTitle = True
                    cat_axis.AxisTitle.Text = config["xlabel"]
                if config["ylabel"]:
                    val_axis = chart.Axes(c.xlValue)
                    val_axis.HasTitle = True
                    val_axis.AxisTitle.Text = config["ylabel"]
            wb.Save()
            wb.Close()
        finally:
            excel.Quit()


# -----------------------------------------------------------------------------
# 3. Comparator
# -----------------------------------------------------------------------------


class ChartComparator:
    """Extracts chart structure from `visi_out.xlsx` and `excel_out.xlsx`
    via `chart_xlsx_reader.read_charts` and diffs type/ranges/title/
    xlabel/ylabel/legend. Assumes exactly one chart per file -- a
    deliberate fuzz-scope limitation that avoids chart-matching/ordering
    ambiguity, matching visi's own single-series-per-chart model.
    """

    def _normalize_range(self, ref):
        """Strips `$` and surrounding single-quotes around a sheet name,
        and lowercases the sheet-name portion, so e.g. an
        openpyxl/Excel-authored `'Sheet1'!$B$2:$B$4` compares equal to
        visi's `Sheet1!$B$1:$B$4`. New logic -- `fuzz_excel.py`'s
        comparator never needed range normalization for plain cell-value
        comparison."""
        if ref is None:
            return None
        if "!" not in ref:
            return ref.replace("$", "")
        sheet, cell_range = ref.split("!", 1)
        sheet = sheet.strip("'").lower()
        return f"{sheet}!{cell_range.replace('$', '')}"

    def compare(self, visi_xlsx_path, excel_xlsx_path):
        """Returns (is_match, mismatches), the same shape as
        `DifferentialComparator.compare` in `fuzz_excel.py`, for a similar
        reporting loop in `main()`."""
        visi_charts = read_charts(visi_xlsx_path)
        excel_charts = read_charts(excel_xlsx_path)
        mismatches = []

        if len(visi_charts) != 1 or len(excel_charts) != 1:
            mismatches.append({
                "field": "chart_count",
                "visi": len(visi_charts),
                "excel": len(excel_charts),
            })
            return False, mismatches

        v = visi_charts[0]
        e = excel_charts[0]

        fields = ["chart_type", "title", "xlabel", "ylabel", "show_legend"]
        for field in fields:
            if v[field] != e[field]:
                mismatches.append({"field": field, "visi": v[field], "excel": e[field]})

        for field in ["cat_range", "val_range"]:
            v_norm = self._normalize_range(v[field])
            e_norm = self._normalize_range(e[field])
            if v_norm != e_norm:
                mismatches.append({"field": field, "visi": v[field], "excel": e[field]})

        return len(mismatches) == 0, mismatches


# -----------------------------------------------------------------------------
# 4. Main fuzzing loop
# -----------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Differential fuzzing test harness for visi vs Microsoft Excel charts."
    )
    parser.add_argument("--excel-path", help="Path to Microsoft Excel binary or application bundle.")
    parser.add_argument(
        "--driver", choices=["auto", "applescript", "win32com", "mock"], default="auto",
        help="Excel execution driver.",
    )
    parser.add_argument("--visi-path", default="./target/release/visi", help="Path to compiled visi binary (used by the subprocess backend).")
    add_backend_arg(parser)
    parser.add_argument("--iterations", type=int, default=10, help="Number of fuzz iterations to run.")
    parser.add_argument("--rows", type=int, default=8, help="Max source data rows per iteration.")
    parser.add_argument("--seed", type=int, default=None, help="Random seed for reproducible fuzzing.")
    parser.add_argument("--output-dir", default="./fuzz_results", help="Directory to store test outputs and failure artifacts.")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    failures_dir = os.path.join(args.output_dir, "failures")
    os.makedirs(failures_dir, exist_ok=True)

    visi_driver = VisiChartDriver(binary_path=args.visi_path, backend=args.backend)
    if args.backend == "auto" and visi_driver.backend != "bindings":
        print(bindings_hint(), file=sys.stderr)
    excel_driver = ExcelChartDriver(excel_path=args.excel_path, driver_type=args.driver)
    comparator = ChartComparator()
    smoke_mode = excel_driver.driver_type == "mock"

    print("=====================================================================")
    print("         visi vs. Microsoft Excel Chart Differential Fuzzer          ")
    print("=====================================================================")
    print(f" Iterations  : {args.iterations}")
    print(f" Max rows    : {args.rows}")
    print(f" Visi        : {visi_driver.describe()}")
    print(f" Excel Driver: {excel_driver.driver_type} ({args.excel_path or 'Default'})")
    if smoke_mode:
        print(f" {SMOKE_BANNER}")
    print("=====================================================================\n")

    passed_count = 0
    failed_count = 0
    start_time = time.time()

    for i in range(1, args.iterations + 1):
        iter_seed = (args.seed + i) if args.seed is not None else random.randint(1, 1000000)
        generator = ChartFuzzGenerator(seed=iter_seed)

        temp_dir = tempfile.mkdtemp(prefix=f"fuzz_chart_iter_{i}_")
        source_xlsx = os.path.join(temp_dir, "source.xlsx")
        visi_out_xlsx = os.path.join(temp_dir, "visi_out.xlsx")
        excel_out_xlsx = os.path.join(temp_dir, "excel_out.xlsx")

        try:
            num_rows = random.randint(2, max(2, args.rows))
            range_str, add_config, edit_config = generator.generate(source_xlsx, num_rows=num_rows)

            visi_driver.run(source_xlsx, range_str, add_config, edit_config, visi_out_xlsx)

            if smoke_mode:
                ok, reason = smoke_check(read_charts(visi_out_xlsx), what="charts")
                if ok:
                    passed_count += 1
                    print(
                        f" Iteration {i:3d}/{args.iterations} [OK] (Seed: {iter_seed},"
                        f" type: {edit_config['chart_type']})"
                    )
                else:
                    failed_count += 1
                    print(f"\n Iteration {i:3d}/{args.iterations} [FAILED] (Seed: {iter_seed})")
                    print(f"   {reason}")
                    fail_case_dir = os.path.join(
                        failures_dir, f"chart_smoke_iter_{i}_seed_{iter_seed}"
                    )
                    shutil.copytree(temp_dir, fail_case_dir, dirs_exist_ok=True)
                    print(f"   Saved reproducing files to: {fail_case_dir}\n")
                continue

            excel_driver.run(source_xlsx, range_str, edit_config, excel_out_xlsx)

            is_match, mismatches = comparator.compare(visi_out_xlsx, excel_out_xlsx)

            if is_match:
                passed_count += 1
                print(f" Iteration {i:3d}/{args.iterations} [PASSED] (Seed: {iter_seed}, type: {edit_config['chart_type']})")
            else:
                failed_count += 1
                print(f"\n Iteration {i:3d}/{args.iterations} [FAILED] (Seed: {iter_seed})")
                print(f"   Found {len(mismatches)} field mismatch(es):")
                for m in mismatches[:5]:
                    print(f"   - {m['field']}: visi={m['visi']} | Excel={m['excel']}")

                fail_case_dir = os.path.join(failures_dir, f"chart_fail_iter_{i}_seed_{iter_seed}")
                shutil.copytree(temp_dir, fail_case_dir, dirs_exist_ok=True)
                print(f"   Saved failure reproducing files to: {fail_case_dir}\n")

        except Exception as err:
            failed_count += 1
            print(f"\n Iteration {i:3d}/{args.iterations} [ERROR]: {err}")
            fail_case_dir = os.path.join(failures_dir, f"chart_error_iter_{i}_seed_{iter_seed}")
            if os.path.exists(temp_dir):
                shutil.copytree(temp_dir, fail_case_dir, dirs_exist_ok=True)

        finally:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)

    duration = time.time() - start_time
    print("\n=====================================================================")
    print(f" {'Smoke test' if smoke_mode else 'Fuzzing'} Completed in {duration:.2f}s")
    if smoke_mode:
        print(f" Ran    : {passed_count}/{args.iterations} without a crash")
    else:
        print(f" Passed : {passed_count}/{args.iterations}")
    print(f" Failed : {failed_count}/{args.iterations}")
    print("=====================================================================")

    if failed_count > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
