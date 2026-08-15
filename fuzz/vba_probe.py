#!/usr/bin/env python3
"""
VBA execution probe: can real Excel run a macro that only `visi` authored?
==========================================================================
Not a fuzzer. This is the feasibility check underpinning the test plan in
`docs/vba-macro-support.md` -- a fixed, deterministic set of assertions about
what Excel's AppleScript bridge will and will not do with a VBA module that
was injected into `vbaProject.bin` by `visi macro add`, with Excel never
involved in authoring it.

It exists because that pipeline is the load-bearing assumption of any future
`fuzz_vba.py`, and it is not obvious that it works: Excel for Mac's
AppleScript dictionary exposes no VBProject object, so there is no automation
path that puts a macro *into* a workbook. `visi` writing the module at the
file-format level is the only reason a VBA differential fuzzer is possible on
macOS at all. Re-run this before trusting that plan after any change to
`vba_xlsx.rs` / `vba_synth.rs`, or after an Excel update.

The four checks, and why each one matters to the harness design:

    author-and-run   visi-authored module loads, runs, and its cell writes
                     survive Excel's own save (and are still readable by
                     visi afterwards)
    return-value     `run VB macro` hands a typed return value straight back
                     to AppleScript, so results need not be routed through
                     cells and a file read
    trapped-error    a runtime error under `On Error GoTo` comes back as
                     structured text, making Err.Number comparable against
                     the interpreter's
    wrapper          an error raised inside a *called* procedure still
                     reaches the caller's handler -- this is what lets the
                     harness wrap generated code safely

There is a fifth behaviour, deliberately NOT asserted here because asserting
it means hanging for the timeout on every run: an *untrapped* runtime error
pops a modal dialog that `set display alerts to false` does not suppress, the
osascript call never returns, and Excel must be SIGKILLed. That is why the
`wrapper` check exists. Pass --demo-hang to see it, at the cost of a stall.

Usage:
    python3 fuzz/vba_probe.py
    python3 fuzz/vba_probe.py --visi target/release/visi --keep
"""

import argparse
import os
import shutil
import subprocess
import sys
import tempfile

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: source fuzz/venv/bin/activate && pip install -r fuzz/requirements.txt")

EXCEL_APP = "Microsoft Excel"
OSASCRIPT_TIMEOUT = 60


# -- VBA sources ---------------------------------------------------------

PROBE_BAS = '''Attribute VB_Name = "VisiProbe"
Public Sub RunProbe()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Sheet1")
    Dim i As Long
    Dim total As Double
    total = 0
    For i = 1 To 5
        total = total + i * i
    Next i
    ws.Range("C1").Value = total
    ws.Range("C2").Value = "hello " & CStr(Len("abcd"))
    ws.Range("C3").Value = ws.Range("A1").Value + ws.Range("A2").Value
    ThisWorkbook.Save
End Sub
'''

# `Harness` is the pattern a generated-VBA fuzzer must use verbatim: the
# generated procedure is called from inside an `On Error GoTo`, so a runtime
# error anywhere down the call stack returns as data instead of stalling the
# automation bridge on a modal dialog.
HARNESS_BAS = '''Attribute VB_Name = "VisiHarness"
Public Function Harness(ByVal which As String) As String
    On Error GoTo Failed
    Dim r As Variant
    Select Case which
        Case "double"
            r = Doubler("21")
        Case "divzero"
            r = DivZero()
        Case "typemismatch"
            r = TypeMismatch()
    End Select
    Harness = "OK|" & TypeName(r) & "|" & CStr(r)
    Exit Function
Failed:
    Harness = "ERR|" & CStr(Err.Number) & "|" & Err.Description
End Function

Public Function Doubler(ByVal src As String) As Variant
    Doubler = CDbl(src) * 2
End Function

Public Function DivZero() As Variant
    Dim x As Double
    x = 1 / 0
    DivZero = x
End Function

Public Function TypeMismatch() As Variant
    TypeMismatch = CLng("not a number")
End Function
'''

# Only reached with --demo-hang: no handler anywhere, so Excel goes modal.
HANG_BAS = '''Attribute VB_Name = "VisiHang"
Public Function Unhandled() As String
    Unhandled = CStr(CLng("not a number"))
End Function
'''


# -- helpers -------------------------------------------------------------

def run_osascript(script, timeout=OSASCRIPT_TIMEOUT):
    """Returns (ok, output). ok=False with output='<timeout>' means Excel went
    modal -- the caller is responsible for restarting it."""
    try:
        res = subprocess.run(
            ["osascript", "-e", script],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            text=True, timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        return False, "<timeout>"
    if res.returncode != 0:
        return False, res.stderr.strip()
    return True, res.stdout.strip()


def restart_excel():
    """SIGKILL by PID, not `killall` alone -- Excel can intercept SIGTERM to
    run its own quit handshake and stay listed as running (see
    fuzz_pivot.py::_restart_excel)."""
    subprocess.run(["killall", EXCEL_APP], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    subprocess.run(["sleep", "1"])
    pgrep = subprocess.run(["pgrep", "-x", EXCEL_APP], stdout=subprocess.PIPE, text=True)
    for pid in pgrep.stdout.split():
        subprocess.run(["kill", "-9", pid], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    subprocess.run(["sleep", "1"])


def excel_script(path, body):
    return "\n".join([
        f'tell application "{EXCEL_APP}"',
        "    set display alerts to false",
        "    try",
        "        close workbooks saving no",
        "    end try",
        f'    open POSIX file "{path}"',
        "    set wb to active workbook",
        body,
        "    close wb saving no",
        "    return theResult",
        "end tell",
    ])


def make_base_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    wb.save(path)


def visi_macro_add(visi, base, name, source, out):
    bas = out + ".bas"
    with open(bas, "w") as f:
        f.write(source)
    res = subprocess.run(
        [visi, "macro", "add", base, "--name", name, "--kind", "standard",
         "--source-file", bas, "--output", out],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
    )
    if res.returncode != 0:
        raise RuntimeError(f"visi macro add failed: {res.stderr.strip()}")


# -- checks --------------------------------------------------------------

def check_author_and_run(visi, workdir, results):
    """visi-authored module loads, runs, mutates cells, and the mutations
    survive Excel's save -- readable by both openpyxl and visi."""
    base = os.path.join(workdir, "base.xlsx")
    xlsm = os.path.join(workdir, "probe.xlsm")
    make_base_workbook(base)
    visi_macro_add(visi, base, "VisiProbe", PROBE_BAS, xlsm)

    ok, out = run_osascript(excel_script(
        xlsm, '    run VB macro "RunProbe"\n    set theResult to "ran"'))
    if not ok:
        results.append(("author-and-run", False, f"AppleScript failed: {out}"))
        return

    wb = openpyxl.load_workbook(xlsm, data_only=True)
    ws = wb["Sheet1"]
    got = [ws.cell(row=r, column=3).value for r in (1, 2, 3)]
    want = [55, "hello 4", 3]
    if got != want:
        results.append(("author-and-run", False, f"cells {got!r} != {want!r}"))
        return

    listing = subprocess.run([visi, "macro", "list", xlsm],
                            stdout=subprocess.PIPE, text=True).stdout
    if "VisiProbe" not in listing:
        results.append(("author-and-run", False,
                        "visi no longer lists the module after Excel saved the file"))
        return
    results.append(("author-and-run", True, f"C1:C3 = {got!r}; module survives Excel's save"))


def check_macro_behaviours(visi, workdir, results):
    """Return values, trapped errors, and error propagation out of a called
    procedure -- all three in one Excel session, since the session is the
    expensive part."""
    base = os.path.join(workdir, "base2.xlsx")
    xlsm = os.path.join(workdir, "harness.xlsm")
    make_base_workbook(base)
    visi_macro_add(visi, base, "VisiHarness", HARNESS_BAS, xlsm)

    body = "\n".join([
        '    set r1 to run VB macro "Harness" arg1 "double"',
        '    set r2 to run VB macro "Harness" arg1 "divzero"',
        '    set r3 to run VB macro "Harness" arg1 "typemismatch"',
        '    set theResult to r1 & " ;; " & r2 & " ;; " & r3',
    ])
    ok, out = run_osascript(excel_script(xlsm, body))
    if not ok:
        for name in ("return-value", "trapped-error", "wrapper"):
            results.append((name, False, f"AppleScript failed: {out}"))
        return

    parts = [p.strip() for p in out.split(";;")]
    if len(parts) != 3:
        for name in ("return-value", "trapped-error", "wrapper"):
            results.append((name, False, f"unparseable result: {out!r}"))
        return
    r1, r2, r3 = parts

    results.append(("return-value", r1 == "OK|Double|42",
                    f"run VB macro returned {r1!r} (want 'OK|Double|42')"))
    results.append(("trapped-error", r2.startswith("ERR|11|"),
                    f"1/0 under On Error returned {r2!r} (want 'ERR|11|...')"))
    results.append(("wrapper", r3.startswith("ERR|13|"),
                    f"error inside a called proc returned {r3!r} (want 'ERR|13|...')"))


def demo_hang(visi, workdir, results):
    """Deliberately reproduce the modal-dialog hang, then clean up after it."""
    base = os.path.join(workdir, "base3.xlsx")
    xlsm = os.path.join(workdir, "hang.xlsm")
    make_base_workbook(base)
    visi_macro_add(visi, base, "VisiHang", HANG_BAS, xlsm)

    print(f"  (stalling up to {OSASCRIPT_TIMEOUT}s on purpose...)", flush=True)
    ok, out = run_osascript(excel_script(
        xlsm, '    set theResult to run VB macro "Unhandled"'))
    restart_excel()
    results.append(("untrapped-error-hangs", (not ok) and out == "<timeout>",
                    f"untrapped error {'hung as expected' if out == '<timeout>' else f'gave {out!r}'}"))


# -- main ----------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--visi", default="target/release/visi", help="path to the visi binary")
    ap.add_argument("--keep", action="store_true", help="keep the generated workbooks")
    ap.add_argument("--demo-hang", action="store_true",
                    help="also reproduce the untrapped-error hang (stalls, then restarts Excel)")
    args = ap.parse_args()

    if sys.platform != "darwin":
        sys.exit("This probe drives Excel through AppleScript and is macOS-only.")
    visi = shutil.which(args.visi) or os.path.abspath(args.visi)
    if not os.path.exists(visi):
        sys.exit(f"visi binary not found at {args.visi!r} -- build it with `cargo build --release`")

    workdir = tempfile.mkdtemp(prefix="vba_probe_")
    results = []
    try:
        check_author_and_run(visi, workdir, results)
        check_macro_behaviours(visi, workdir, results)
        if args.demo_hang:
            demo_hang(visi, workdir, results)
    finally:
        if args.keep:
            print(f"\nworkbooks kept in {workdir}")
        else:
            shutil.rmtree(workdir, ignore_errors=True)

    print()
    failed = 0
    for name, ok, detail in results:
        print(f"  [{'PASS' if ok else 'FAIL'}] {name:24} {detail}")
        failed += 0 if ok else 1
    print()
    if failed:
        print(f"{failed} of {len(results)} checks failed -- the VBA fuzz plan's assumptions "
              f"no longer hold; see docs/vba-macro-support.md")
    else:
        print(f"all {len(results)} checks passed -- Excel runs visi-authored macros end to end")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
