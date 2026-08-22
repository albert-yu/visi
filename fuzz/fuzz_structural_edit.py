#!/usr/bin/env python3
"""Differential fuzzer for structural edits against real Microsoft Excel.

This harness generates a small multi-sheet workbook with values and formulas,
applies a random sequence of row/column insert/delete operations in both Excel
and visi, recalculates/saves, then compares both the rewritten formula text and
cached values.

Examples:

    # Windows, real Excel oracle
    PYTHONIOENCODING=utf-8 python fuzz/fuzz_structural_edit.py --driver win32com --iterations 20

    # macOS, real Excel oracle
    python fuzz/fuzz_structural_edit.py --excel-path "/Applications/Microsoft Excel.app" --iterations 20

    # Harness smoke test only; does not invoke Excel
    python fuzz/fuzz_structural_edit.py --driver mock --iterations 2
"""

import argparse
import os
import random
import shutil
import subprocess
import sys
import tempfile
import time
from dataclasses import dataclass

import openpyxl

from fuzz_excel import DifferentialComparator, ExcelDriver, XLSXEvaluatedReader
from visi_driver import CLI_TIMEOUT_SECONDS

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SHEETS = ["Sheet1", "Data"]
ROWS = 8
COLS = 8
MAX_EDITS = 5


@dataclass
class Edit:
    kind: str
    sheet: str
    index: int

    def label(self):
        axis = "row" if "row" in self.kind else "col"
        verb = "insert" if self.kind.startswith("insert") else "delete"
        return f"{self.sheet}:{verb}_{axis}@{self.index}"


def col_name(idx):
    return openpyxl.utils.get_column_letter(idx)


def run(cmd):
    res = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=CLI_TIMEOUT_SECONDS,
    )
    if res.returncode != 0:
        raise RuntimeError(
            f"command failed ({res.returncode}): {' '.join(cmd)}\nSTDOUT:{res.stdout}\nSTDERR:{res.stderr}"
        )
    return res.stdout


class StructuralFuzzGenerator:
    def __init__(self, seed):
        self.seed = seed
        self.rng = random.Random(seed)

    def value(self):
        choice = self.rng.choice(["int", "float", "text", "blank"])
        if choice == "blank":
            return None
        if choice == "text":
            return self.rng.choice(["alpha", "bravo", "x", "", "42"])
        if choice == "float":
            return round(self.rng.uniform(-20, 20), 3)
        return self.rng.randint(-20, 20)

    def cell_ref(self, abs_ok=True, sheet=None):
        row = self.rng.randint(1, ROWS)
        col = self.rng.randint(1, COLS)
        col_s = col_name(col)
        if abs_ok and self.rng.random() < 0.35:
            col_s = "$" + col_s
        row_s = str(row)
        if abs_ok and self.rng.random() < 0.35:
            row_s = "$" + row_s
        ref = f"{col_s}{row_s}"
        if sheet:
            ref = f"{sheet}!{ref}"
        return ref

    def range_ref(self, whole=False, sheet=None):
        if whole:
            if self.rng.random() < 0.5:
                c1 = self.rng.randint(1, COLS)
                c2 = self.rng.randint(c1, COLS)
                ref = f"{col_name(c1)}:{col_name(c2)}"
            else:
                r1 = self.rng.randint(1, ROWS)
                r2 = self.rng.randint(r1, ROWS)
                ref = f"{r1}:{r2}"
        else:
            r1 = self.rng.randint(1, ROWS)
            r2 = self.rng.randint(r1, ROWS)
            c1 = self.rng.randint(1, COLS)
            c2 = self.rng.randint(c1, COLS)
            ref = f"{col_name(c1)}{r1}:{col_name(c2)}{r2}"
        if sheet:
            ref = f"{sheet}!{ref}"
        return ref

    def formula(self, current_sheet):
        other = "Data" if current_sheet == "Sheet1" else "Sheet1"
        style = self.rng.choice([
            "ref",
            "abs_ref",
            "sum_range",
            "sum_whole",
            "binary",
            "cross_ref",
            "cross_range",
            "cross_whole",
        ])
        if style == "ref":
            return f"={self.cell_ref(abs_ok=False)}"
        if style == "abs_ref":
            return f"={self.cell_ref(abs_ok=True)}"
        if style == "sum_range":
            return f"=SUM({self.range_ref()})"
        if style == "sum_whole":
            return f"=SUM({self.range_ref(whole=True)})"
        if style == "binary":
            op = self.rng.choice(["+", "-", "*"])
            return f"={self.cell_ref()} {op} {self.cell_ref()}"
        if style == "cross_ref":
            return f"={self.cell_ref(sheet=other)}"
        if style == "cross_whole":
            return f"=SUM({self.range_ref(whole=True, sheet=other)})"
        return f"=SUM({self.range_ref(sheet=other)})"

    def workbook(self, path):
        wb = openpyxl.Workbook()
        wb.active.title = SHEETS[0]
        wb.create_sheet(SHEETS[1])
        for ws in wb.worksheets:
            for r in range(1, ROWS + 1):
                for c in range(1, COLS + 1):
                    # Keep formulas on the first sheet. The rewrite compiler's
                    # unqualified-reference representation is anchored there,
                    # while explicit Data! references still exercise cross-sheet
                    # structural edits.
                    if ws.title == "Sheet1" and self.rng.random() < 0.30:
                        ws.cell(r, c, self.formula(ws.title))
                    else:
                        ws.cell(r, c, self.value())
        wb.save(path)

    def edits(self):
        count = self.rng.randint(1, MAX_EDITS)
        edits = []
        dims = {sheet: [ROWS, COLS] for sheet in SHEETS}
        for _ in range(count):
            sheet = self.rng.choice(SHEETS)
            possible = ["insert_row", "insert_col"]
            if dims[sheet][0] > 1:
                possible.append("delete_row")
            if dims[sheet][1] > 1:
                possible.append("delete_col")
            kind = self.rng.choice(possible)
            axis = 0 if "row" in kind else 1
            limit = dims[sheet][axis]
            index = self.rng.randint(1, limit)
            edits.append(Edit(kind, sheet, index))
            dims[sheet][axis] += 1 if kind.startswith("insert") else -1
        return edits


class VisiStructuralDriver:
    def __init__(self, binary_path=None):
        self.binary_path = self._resolve_binary(binary_path)
        if not self.binary_path or not os.path.exists(self.binary_path):
            raise RuntimeError("visi binary not found; run cargo build or pass --visi-path")

    def _resolve_binary(self, binary_path):
        if binary_path and os.path.exists(binary_path):
            return binary_path
        candidates = []
        if binary_path:
            candidates.extend([binary_path, binary_path + ".exe"])
        for flavor in ("release", "debug"):
            base = os.path.join(REPO, "target", flavor, "visi")
            candidates.extend([base, base + ".exe"])
        existing = [p for p in candidates if os.path.exists(p)]
        return max(existing, key=os.path.getmtime) if existing else binary_path

    def run(self, source, edits, output):
        shutil.copyfile(source, output)
        for edit in edits:
            axis = "row" if "row" in edit.kind else "col"
            verb = "insert" if edit.kind.startswith("insert") else "delete"
            idx = str(edit.index if axis == "row" else col_name(edit.index))
            run([
                self.binary_path,
                axis,
                verb,
                output,
                "--sheet",
                edit.sheet,
                "--index",
                idx,
                "--in-place",
                "--quiet",
            ])


class ExcelStructuralDriver:
    def __init__(self, excel_path=None, driver_type="auto"):
        self.inner = ExcelDriver(excel_path=excel_path, driver_type=driver_type)

    def describe(self):
        return self.inner.driver_type

    def run(self, source, edits, output):
        shutil.copyfile(source, output)
        abs_output = os.path.abspath(output)
        if self.inner.driver_type == "mock":
            print("[ExcelStructuralDriver Warning] Running in mock mode (Excel not invoked).")
            return
        if self.inner.driver_type == "win32com":
            self._run_win32com(abs_output, edits)
        elif self.inner.driver_type == "applescript":
            self._run_applescript(abs_output, edits)
        else:
            raise RuntimeError(f"driver {self.inner.driver_type!r} is not supported for structural edits")

    def _run_win32com(self, path, edits):
        try:
            import win32com.client
        except ImportError:
            raise RuntimeError("pywin32 (win32com) is required for Excel automation on Windows.")
        last_err = None
        for attempt in range(5):
            if attempt > 0:
                subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                time.sleep(1.0)
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            try:
                wb = excel.Workbooks.Open(path)
                for edit in edits:
                    ws = wb.Worksheets(edit.sheet)
                    if edit.kind == "insert_row":
                        ws.Rows(edit.index).Insert()
                    elif edit.kind == "delete_row":
                        ws.Rows(edit.index).Delete()
                    elif edit.kind == "insert_col":
                        ws.Columns(edit.index).Insert()
                    else:
                        ws.Columns(edit.index).Delete()
                excel.CalculateFullRebuild()
                wb.Save()
                wb.Close(False)
                last_err = None
                break
            except Exception as exc:
                last_err = exc
            finally:
                excel.Quit()
        if last_err is not None:
            raise last_err

    def _run_applescript(self, path, edits):
        app_name = self.inner.excel_path if self.inner.excel_path else "Microsoft Excel"
        if app_name.endswith(".app"):
            app_name = os.path.splitext(os.path.basename(app_name))[0]
        edit_lines = []
        for edit in edits:
            if edit.kind == "insert_row":
                line = f'insert into range (entire row of range "A{edit.index}" of worksheet "{edit.sheet}" of active workbook) shift shift down'
            elif edit.kind == "delete_row":
                line = f'delete range (entire row of range "A{edit.index}" of worksheet "{edit.sheet}" of active workbook) shift shift up'
            elif edit.kind == "insert_col":
                col = col_name(edit.index)
                line = f'insert into range (entire column of range "{col}1" of worksheet "{edit.sheet}" of active workbook) shift shift to right'
            else:
                col = col_name(edit.index)
                line = f'delete range (entire column of range "{col}1" of worksheet "{edit.sheet}" of active workbook) shift shift to left'
            edit_lines.append(line)
        script = f'''
        tell application "{app_name}"
            set display alerts to false
            try
                close workbooks saving no
            end try
            try
                open POSIX file "{path}"
                {chr(10).join(edit_lines)}
                calculate full
                save active workbook
                close active workbook saving no
            on error errText number errNum
                try
                    close workbooks saving no
                end try
                error errText number errNum
            end try
        end tell
        '''
        res = subprocess.run(["osascript", "-e", script], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=60)
        if res.returncode != 0:
            raise RuntimeError(f"Excel AppleScript failed:\nSTDERR: {res.stderr}")


def read_formulas(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    out[(ws.title, cell.coordinate)] = value
    return out


def normalize_formula_text(formula):
    if formula is None:
        return None
    text = "".join(str(formula).split()).upper()
    # Excel preserves the sheet prefix when a cross-sheet reference is deleted
    # (`Data!#REF!`); visi serializes the same invalid reference as plain
    # `#REF!`. They evaluate the same, and this harness is aimed at movement
    # bugs rather than that cosmetic spelling difference.
    for sheet in SHEETS:
        text = text.replace(f"{sheet.upper()}!#REF!", "#REF!")
    return text


def formula_mismatches(visi_path, excel_path):
    vf = read_formulas(visi_path)
    ef = read_formulas(excel_path)
    mismatches = []
    for key in sorted(set(vf) | set(ef)):
        if normalize_formula_text(vf.get(key)) != normalize_formula_text(ef.get(key)):
            mismatches.append({
                "key": key,
                "reason": "Formula text mismatch",
                "visi": vf.get(key),
                "excel": ef.get(key),
                "formula": vf.get(key) or ef.get(key),
            })
    return mismatches


def compare_values(visi_path, excel_path, strict_error_class=False):
    visi_cells = XLSXEvaluatedReader.read_evaluated_cells(visi_path)
    excel_cells = XLSXEvaluatedReader.read_evaluated_cells(excel_path)
    comp = DifferentialComparator(strict_error_class=strict_error_class)
    ok, mismatches = comp.compare(visi_cells, excel_cells)
    # After some structural edits, Excel occasionally saves rewritten formula
    # text without a cached <v>. Formula text is this harness's primary signal;
    # compare values only where the oracle actually wrote one.
    mismatches = [
        m for m in mismatches
        if not (m.get("excel") is None and m.get("formula"))
    ]
    return not mismatches, mismatches, comp.error_class_only


def save_failure(work_dir, output_dir, iteration, seed):
    fail_dir = os.path.join(output_dir, "failures", f"structural_fail_iter_{iteration}_seed_{seed}")
    os.makedirs(fail_dir, exist_ok=True)
    for name in ["source.xlsx", "visi_out.xlsx", "excel_out.xlsx", "edits.txt"]:
        src = os.path.join(work_dir, name)
        if os.path.exists(src):
            shutil.copyfile(src, os.path.join(fail_dir, name))
    return fail_dir


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--iterations", type=int, default=20)
    parser.add_argument("--seed", type=int, default=None, help="Run one deterministic iteration seed")
    parser.add_argument("--driver", choices=["auto", "applescript", "win32com", "mock"], default="auto")
    parser.add_argument("--excel-path", default=None)
    parser.add_argument("--visi-path", default=None)
    parser.add_argument("--output-dir", default="fuzz_results")
    parser.add_argument("--strict-error-class", action="store_true")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    visi = VisiStructuralDriver(args.visi_path)
    excel = ExcelStructuralDriver(args.excel_path, args.driver)

    print("=" * 69)
    print("        visi vs. Microsoft Excel Structural Edit Fuzzer")
    print("=" * 69)
    print(f" Iterations  : {args.iterations}")
    print(f" Visi        : {visi.binary_path}")
    print(f" Excel Driver: {excel.describe()}")
    print("=" * 69)
    print()

    failed = 0
    tolerated = 0
    seeds = [args.seed] if args.seed is not None else [random.randint(1, 1_000_000) for _ in range(args.iterations)]
    for i, seed in enumerate(seeds, 1):
        with tempfile.TemporaryDirectory() as td:
            source = os.path.join(td, "source.xlsx")
            visi_out = os.path.join(td, "visi_out.xlsx")
            excel_out = os.path.join(td, "excel_out.xlsx")
            edits_txt = os.path.join(td, "edits.txt")
            gen = StructuralFuzzGenerator(seed)
            gen.workbook(source)
            edits = gen.edits()
            with open(edits_txt, "w", encoding="utf-8") as f:
                for edit in edits:
                    f.write(edit.label() + "\n")
            try:
                visi.run(source, edits, visi_out)
                excel.run(source, edits, excel_out)
                f_mismatches = [] if excel.inner.driver_type == "mock" else formula_mismatches(visi_out, excel_out)
                values_ok, v_mismatches, error_class_only = (True, [], 0)
                if excel.inner.driver_type != "mock":
                    values_ok, v_mismatches, error_class_only = compare_values(visi_out, excel_out, args.strict_error_class)
                tolerated += error_class_only
                mismatches = f_mismatches + v_mismatches
                if mismatches:
                    failed += 1
                    fail_dir = save_failure(td, args.output_dir, i, seed)
                    print(f" Iteration {i:3d}/{len(seeds)} [FAILED] (Seed: {seed})")
                    print(f"   Edits: {', '.join(e.label() for e in edits)}")
                    print(f"   Artifacts: {fail_dir}")
                    for m in mismatches[:10]:
                        print(f"   - {m['reason']} at {m['key']}: visi={m['visi']} | Excel={m['excel']} (Formula: {m.get('formula')})")
                else:
                    print(f" Iteration {i:3d}/{len(seeds)} [PASSED] (Seed: {seed})")
            except Exception as exc:
                failed += 1
                fail_dir = save_failure(td, args.output_dir, i, seed)
                print(f" Iteration {i:3d}/{len(seeds)} [ERROR] (Seed: {seed})")
                print(f"   Edits: {', '.join(e.label() for e in edits)}")
                print(f"   Artifacts: {fail_dir}")
                print(f"   {type(exc).__name__}: {exc}")

    print()
    print("=" * 69)
    print(" Structural edit fuzzing completed")
    print(f" Passed : {len(seeds) - failed}/{len(seeds)}")
    print(f" Failed : {failed}/{len(seeds)}")
    if tolerated:
        print(f" Tolerated value error-class divergences: {tolerated}")
    print("=" * 69)
    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
