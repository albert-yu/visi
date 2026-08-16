#!/usr/bin/env python3
"""
Does a held VBA `Range` follow a row or column insert/delete?
============================================================
Issue #58 left this open: visi's `Range` is a `(sheet_id, rect)` *value*
(`vba/host.rs`), so a row inserted above one silently makes it point somewhere
else, while Excel's `Range` objects track the edit. Matching Excel means
interning ranges -- but "tracks the edit" is not a specification, and the
interesting cases are the ones nobody remembers:

* What is `r.Address` after the single row `r` pointed at is **deleted**?
* Does a multi-row `r` *grow* when a row is inserted inside it, the way a
  formula's range reference does?
* Does inserting at `r`'s first row move it or grow it?
* Does a range survive having *all* of its rows deleted, and as what?

This asks Excel and prints the answer. It compares nothing -- visi does not
implement the tracking yet, which is the point.

    source fuzz/venv/bin/activate
    maturin develop -m visi-python/Cargo.toml --release
    python fuzz/vba_range_tracking_probe.py

**Every case runs in its own Excel round trip** (`--batch 1`, and do not raise
it). Unlike `vba_host_probe.py`'s read cases, every case here mutates the
sheet's shape, so two cases sharing a session would measure each other rather
than Excel. A round trip re-opens the workbook, which is what resets the grid.

The traps inherited from the other probes still apply: a **compile** error
hangs the AppleScript bridge and is not catchable by the `On Error` wrapper,
so a case that never returns is a compile error rather than a slow run --
`killall "Microsoft Excel"` and read the generated source.
"""

import argparse
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: source fuzz/venv/bin/activate && pip install -r fuzz/requirements.txt")

try:
    import visi_core
except ImportError:
    sys.exit(
        "the visi_core bindings are required: "
        "maturin develop -m visi-python/Cargo.toml --release"
    )

from fuzz_vba import HARNESS_TEMPLATE, ExcelDriver

PREAMBLE = [
    "Dim ws As Worksheet, wb As Workbook",
    "Set wb = ThisWorkbook",
    'Set ws = wb.Worksheets("Sheet1")',
    "Dim r As Range, q As Range, s",
]

# A1:A10 hold 1..10 and B1:B10 hold 101..110, so a tracked range can be asked
# what it now *reads* as well as where it now points -- an address that looks
# right over the wrong data would otherwise pass unnoticed.
CASES = [
    # --- the baseline: does tracking happen at all
    ('insert above a single cell',
     'Set r = ws.Range("A5") :: ws.Rows(1).Insert :: r.Address'),
    ('insert above a single cell, value',
     'Set r = ws.Range("A5") :: ws.Rows(1).Insert :: CStr(r.Value)'),
    ('insert below a single cell',
     'Set r = ws.Range("A5") :: ws.Rows(9).Insert :: r.Address'),

    # --- span growth, the rule that differs between "move" and "grow"
    ('insert at a span first row',
     'Set r = ws.Range("A5:A7") :: ws.Rows(5).Insert :: r.Address'),
    ('insert inside a span',
     'Set r = ws.Range("A5:A7") :: ws.Rows(6).Insert :: r.Address'),
    ('insert below a span',
     'Set r = ws.Range("A5:A7") :: ws.Rows(8).Insert :: r.Address'),

    # --- deletion, including the cases with no obvious right answer
    ('delete a row above a single cell',
     'Set r = ws.Range("A5") :: ws.Rows(1).Delete :: r.Address'),
    ('delete the single row a cell points at',
     'Set r = ws.Range("A5") :: ws.Rows(5).Delete :: r.Address'),
    ('delete the single row a cell points at, value',
     'Set r = ws.Range("A5") :: ws.Rows(5).Delete :: CStr(r.Value)'),
    ('delete a row inside a span',
     'Set r = ws.Range("A5:A7") :: ws.Rows(6).Delete :: r.Address'),
    ('delete a span first row',
     'Set r = ws.Range("A5:A7") :: ws.Rows(5).Delete :: r.Address'),
    ('delete every row of a span',
     'Set r = ws.Range("A5:A7") :: ws.Rows("5:7").Delete :: r.Address'),
    ('delete every row of a span, value',
     'Set r = ws.Range("A5:A7") :: ws.Rows("5:7").Delete :: CStr(r.Cells(1, 1).Value)'),

    # --- columns
    ('insert a column left of a cell',
     'Set r = ws.Range("C5") :: ws.Columns(1).Insert :: r.Address'),
    ('insert a column inside a span',
     'Set r = ws.Range("A5:C5") :: ws.Columns(2).Insert :: r.Address'),
    ('delete the column a cell points at',
     'Set r = ws.Range("C5") :: ws.Columns(3).Delete :: r.Address'),

    # --- an edit somewhere else must not move it
    ('insert on another sheet',
     'Set r = ws.Range("A5") :: wb.Worksheets("Sheet2").Rows(1).Insert :: r.Address'),

    # --- identity survives tracking, i.e. it is the same object that moved
    ('identity survives an edit',
     'Set r = ws.Range("A5") :: Set q = r :: ws.Rows(1).Insert :: CStr(q Is r) & "/" & q.Address'),
    ('a copy taken before the edit tracks too',
     'Set r = ws.Range("A5") :: Set q = r :: ws.Rows(1).Insert :: r.Address & "/" & q.Address'),

    # --- the same edit spelled through a Range rather than through Rows()
    ('EntireRow.Insert',
     'Set r = ws.Range("A5") :: ws.Range("A1").EntireRow.Insert :: r.Address'),
    ('EntireColumn.Insert',
     'Set r = ws.Range("C5") :: ws.Range("A1").EntireColumn.Insert :: r.Address'),
    ('EntireRow.Delete',
     'Set r = ws.Range("A5") :: ws.Range("A5").EntireRow.Delete :: r.Address'),

    # --- what the edit expressions themselves are, for the surface we build
    ('Rows(n) address', 'ws.Rows(3).Address'),
    ('Columns(n) address', 'ws.Columns(3).Address'),
    ('Range EntireRow address', 'ws.Range("B5").EntireRow.Address'),
    ('Range EntireColumn address', 'ws.Range("B5").EntireColumn.Address'),
    ('Rows(n) TypeName', 'TypeName(ws.Rows(3))'),
    # Inconclusive about tracking, and kept only so nobody re-derives it as
    # evidence: a bare `.Insert` on a partial range lets Excel pick the shift
    # direction from the range's shape, and for a tall narrow one it shifts
    # *right*, so `A5` not moving says nothing about whether it would have
    # tracked a downward shift.
    ('a partial-range Insert picks its own direction',
     'Set r = ws.Range("A5") :: ws.Range("A2:A3").Insert :: r.Address'),

    # --- what a range whose cells were deleted actually raises.
    # The bare `HARNESS_TEMPLATE` only reports `Err.Number`, and the numbers
    # these come back with are not ones anybody should implement from; these
    # cases trap inside the case and report the description too.
    ('dead range, Address err',
     'Set r = ws.Range("A5") :: ws.Rows(5).Delete :: On Error Resume Next\\n'
     's = r.Address\\n'
     's = "[" & CStr(Err.Number) & "] " & Err.Description\\n'
     'On Error GoTo 0 :: s'),
    ('dead range, Value err',
     'Set r = ws.Range("A5") :: ws.Rows(5).Delete :: On Error Resume Next\\n'
     's = CStr(r.Value)\\n'
     's = "[" & CStr(Err.Number) & "] " & Err.Description\\n'
     'On Error GoTo 0 :: s'),
    ('dead range, is it Nothing',
     'Set r = ws.Range("A5") :: ws.Rows(5).Delete :: CStr(r Is Nothing)'),
    ('dead range, TypeName',
     'Set r = ws.Range("A5") :: ws.Rows(5).Delete :: TypeName(r)'),
    ('dead span, Address err',
     'Set r = ws.Range("A5:A7") :: ws.Rows("5:7").Delete :: On Error Resume Next\\n'
     's = r.Address\\n'
     's = "[" & CStr(Err.Number) & "] " & Err.Description\\n'
     'On Error GoTo 0 :: s'),
    ('dead column range, Address err',
     'Set r = ws.Range("C5") :: ws.Columns(3).Delete :: On Error Resume Next\\n'
     's = r.Address\\n'
     's = "[" & CStr(Err.Number) & "] " & Err.Description\\n'
     'On Error GoTo 0 :: s'),
]


def parse_case(text):
    setup, sep, expr = text.rpartition("::")
    if not sep:
        return [], text.strip()
    return [s.strip() for s in setup.split("\\n") if s.strip()], expr.strip()


def build_module(cases):
    parts = ['Attribute VB_Name = "R"']
    for i, (setup, expr) in enumerate(cases, start=1):
        body = "\n".join(f"    {s}" for s in PREAMBLE + setup)
        parts.append(f"Private Function Gen{i}()\n{body}\n    Gen{i} = {expr}\nEnd Function")
        parts.append(HARNESS_TEMPLATE.format(i=i))
    return "\n\n".join(parts) + "\n"


def build_workbook(path):
    """1..10 down column A and 101..110 down column B, plus a second sheet.

    Distinct values per row so a tracked range can be asked what it reads,
    not only where it points.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, 11):
        ws.cell(row=r, column=1, value=r)
        ws.cell(row=r, column=2, value=100 + r)
        ws.cell(row=r, column=3, value=200 + r)
    wb.create_sheet("Sheet2")
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--excel-path")
    ap.add_argument("--driver", choices=["auto", "applescript", "mock"], default="auto")
    ap.add_argument("--timeout", type=int, default=180)
    ap.add_argument("-k", "--filter", default="", help="only cases whose label contains this")
    ap.add_argument("--keep", action="store_true", help="Keep the generated .xlsm")
    args = ap.parse_args()

    selected = [c for c in CASES if args.filter in c[0]]
    cases = [parse_case(expr) for _, expr in selected]
    source = build_module(cases)

    driver = ExcelDriver(args.excel_path, args.driver, args.timeout)
    workdir = tempfile.mkdtemp(prefix="vba_range_tracking_")
    base = os.path.join(workdir, "base.xlsx")
    xlsm = os.path.join(workdir, "probe.xlsm")
    build_workbook(base)
    wbk = visi_core.Workbook.load(base)
    wbk.add_macro("R", source)
    wbk.save(xlsm)
    if args.keep:
        print(f"workbook: {xlsm}\n", file=sys.stderr)

    label_width = min(max(len(label) for label, _ in selected), 44)
    expr_width = min(max(len(expr) for _, expr in selected), 62)

    # One case per round trip: every case here changes the sheet's shape, so
    # sharing a session would have them measure each other.
    for i, (label, expr) in enumerate(selected, start=1):
        got = driver.run_batch(xlsm, [i])
        if not got:
            answer = "<nothing -- compile error>"
        else:
            answer = got.get(i, "<missing>")
        print(f"{label:<{label_width}}  {expr:<{expr_width}}  {answer}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
