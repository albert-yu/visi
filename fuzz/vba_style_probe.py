#!/usr/bin/env python3
"""
What does Excel's VBA style surface actually do?
===============================================
Phase 3 of `docs/vba-macro-support.md` maps `.Interior.Color`, `.Font.*` and
`.NumberFormat` onto `CellStyle`. The colour conversion is notable
because VBA's `Color` is a **BGR** `Long` while `CellStyle.bg_color` is an `"#RRGGBB"` string -- so
`&HFF0000` is *blue*, not red, and getting it wrong produces a file that opens
fine and is the wrong colour.

Two channels, because the VBA one cannot catch a consistent-but-wrong
convention:

* **`--ask`** runs expressions and prints what Excel's object model returns.
  If visi and Excel both round-trip `&HFF0000` back to `16711680`, that says
  nothing about which colour it painted.
* **`--paint`** has Excel *save* the workbook after setting colours, then
  reads the real `fgColor`/`font color` ARGB out of the xlsx with openpyxl.
  This is the channel that settles BGR, and it is independent of both
  implementations.

    source fuzz/venv/bin/activate
    maturin develop -m visi-python/Cargo.toml --release
    python fuzz/vba_style_probe.py              # both channels
    python fuzz/vba_style_probe.py --paint      # just the decisive one

Same traps as the other VBA probes: a **compile** error hangs the AppleScript
bridge and is not catchable by `On Error`, so a batch that returns nothing is
a compile error rather than a slow run.
"""

import argparse
import os
import subprocess
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: source fuzz/venv/bin/activate && pip install -r fuzz/requirements.txt")

try:
    import visi_core
except ImportError:
    sys.exit(
        "the visi_core bindings are required: "
        "maturin develop -m visi-python/Cargo.toml --release"
    )

from fuzz_vba import HARNESS_TEMPLATE, ExcelDriver

PREAMBLE = [
    "Dim ws As Worksheet, wb As Workbook",
    "Set wb = ThisWorkbook",
    'Set ws = wb.Worksheets("Sheet1")',
    "Dim v, s, c",
]

# Reads first, then writes -- a batch boundary re-opens the workbook, so a
# read moved below a write silently measures the wrong thing.
READ_CASES = [
    # --- what RGB() composes, which is the whole question
    "CStr(RGB(255, 0, 0))",
    "CStr(RGB(0, 255, 0))",
    "CStr(RGB(0, 0, 255))",
    "CStr(RGB(1, 2, 3))",
    "TypeName(RGB(1, 2, 3))",
    # out-of-range components, since a macro can pass anything
    "CStr(RGB(300, 0, 0))",
    "CStr(RGB(-1, 0, 0))",
    "CStr(RGB(1.6, 0, 0))",
    # --- the objects themselves
    'TypeName(ws.Range("A1").Interior)',
    'TypeName(ws.Range("A1").Font)',
    # --- defaults on an untouched cell
    'CStr(ws.Range("A1").Interior.Color)',
    'CStr(ws.Range("A1").Interior.ColorIndex)',
    'CStr(ws.Range("A1").Font.Bold)',
    'CStr(ws.Range("A1").Font.Italic)',
    'CStr(ws.Range("A1").Font.Size)',
    'TypeName(ws.Range("A1").Font.Size)',
    'ws.Range("A1").Font.Name',
    'CStr(ws.Range("A1").Font.Color)',
    'CStr(ws.Range("A1").Font.ColorIndex)',
    'ws.Range("A1").NumberFormat',
    # C1 is a date-formatted serial in the fixture.
    'ws.Range("C1").NumberFormat',
    'CStr(ws.Range("C1").Value2)',
    # --- a multi-cell read, where the cells disagree
    'CStr(ws.Range("A1:A2").Font.Bold)',
    'CStr(IsNull(ws.Range("A1:A2").Interior.Color))',
]

WRITE_CASES = [
    # --- setting a fill, read back through the object model
    'ws.Range("G1").Interior.Color = RGB(255, 0, 0) :: CStr(ws.Range("G1").Interior.Color)',
    'ws.Range("G2").Interior.Color = &HFF0000 :: CStr(ws.Range("G2").Interior.Color)',
    'ws.Range("G3").Interior.Color = RGB(1, 2, 3) :: CStr(ws.Range("G3").Interior.Color)',
    # does setting Color move ColorIndex, and vice versa
    'ws.Range("G4").Interior.Color = RGB(255, 0, 0) :: CStr(ws.Range("G4").Interior.ColorIndex)',
    'ws.Range("G5").Interior.ColorIndex = 3 :: CStr(ws.Range("G5").Interior.Color)',
    # clearing
    'ws.Range("G6").Interior.Color = RGB(255, 0, 0)\\nws.Range("G6").Interior.ColorIndex = -4142 :: CStr(ws.Range("G6").Interior.Color)',
    # --- font
    'ws.Range("H1").Font.Bold = True :: CStr(ws.Range("H1").Font.Bold)',
    'ws.Range("H2").Font.Italic = True :: CStr(ws.Range("H2").Font.Italic)',
    'ws.Range("H3").Font.Size = 14 :: CStr(ws.Range("H3").Font.Size)',
    'ws.Range("H4").Font.Name = "Courier New" :: ws.Range("H4").Font.Name',
    'ws.Range("H5").Font.Color = RGB(0, 0, 255) :: CStr(ws.Range("H5").Font.Color)',
    # --- number format and date formatting
    'ws.Range("I1").Value = 46195\\nws.Range("I1").NumberFormat = "m/d/yy" :: ws.Range("I1").Text',
    'ws.Range("I2").Value = 46195\\nws.Range("I2").NumberFormat = "m/d/yy"\\nws.Range("I2").NumberFormat = "General" :: ws.Range("I2").Text & "|" & CStr(ws.Range("I2").Value2)',
    # setting a format on the *existing* date cell: does the serial survive
    'ws.Range("C1").NumberFormat = "General" :: CStr(ws.Range("C1").Value2) & "|" & TypeName(ws.Range("C1").Value)',
    # --- a whole range at once
    'ws.Range("J1:J3").Interior.Color = RGB(0, 255, 0) :: CStr(ws.Range("J3").Interior.Color)',
    'ws.Range("J1:J3").Font.Bold = True :: CStr(ws.Range("J2").Font.Bold)',
    # --- reading a range whose cells *disagree*. Written here rather than
    # read off the fixture, because a fixture built by openpyxl leaves it
    # ambiguous whether a `False` means "Excel says False" or "the fixture
    # never applied the bold".
    'ws.Range("K1").Font.Bold = True :: CStr(IsNull(ws.Range("K1:K2").Font.Bold))',
    'ws.Range("K3").Font.Bold = True\\nws.Range("K4").Font.Bold = True :: CStr(ws.Range("K3:K4").Font.Bold)',
    'ws.Range("L1").Interior.Color = RGB(255, 0, 0) :: CStr(IsNull(ws.Range("L1:L2").Interior.Color))',
    'ws.Range("M1").Font.Size = 10.5 :: CStr(ws.Range("M1").Font.Size)',
    'ws.Range("M2").Font.Size = 10.5 :: TypeName(ws.Range("M2").Font.Size)',
]

# `ColorIndex` is a palette slot, not a colour. Rather than decide whether to
# support it from documentation, ask Excel what each slot actually is: set the
# index, read the `Color` back. Slot 1..56 plus `xlNone`.
PALETTE_CASES = [
    f'ws.Range("A{i}").Interior.ColorIndex = {i} :: CStr(ws.Range("A{i}").Interior.Color)'
    for i in range(1, 57)
]

# The `--paint` channel: a macro that only writes, after which Excel saves and
# openpyxl reads the actual ARGB. `expect` is what the colour *should* be if
# `Color` is BGR, and is what the run checks against.
PAINT = [
    ("A1", "Interior", "RGB(255, 0, 0)", "FFFF0000"),
    ("A2", "Interior", "&HFF0000", "FF0000FF"),
    ("A3", "Interior", "RGB(1, 2, 3)", "FF010203"),
    ("A4", "Interior", "RGB(0, 128, 0)", "FF008000"),
    ("B1", "Font", "RGB(255, 0, 0)", "FFFF0000"),
    ("B2", "Font", "&HFF0000", "FF0000FF"),
]


def parse_case(text):
    setup, sep, expr = text.rpartition("::")
    if not sep:
        return [], text.strip()
    return [s.strip() for s in setup.split("\\n") if s.strip()], expr.strip()


def build_module(cases):
    parts = ['Attribute VB_Name = "S"']
    for i, (setup, expr) in enumerate(cases, start=1):
        body = "\n".join(f"    {s}" for s in PREAMBLE + setup)
        parts.append(f"Private Function Gen{i}()\n{body}\n    Gen{i} = {expr}\nEnd Function")
        parts.append(HARNESS_TEMPLATE.format(i=i))
    return "\n\n".join(parts) + "\n"


def build_paint_module():
    lines = [
        'Attribute VB_Name = "S"',
        "Public Sub Paint()",
        "    Dim ws As Worksheet",
        '    Set ws = ThisWorkbook.Worksheets("Sheet1")',
    ]
    for addr, kind, expr, _ in PAINT:
        lines.append(f'    ws.Range("{addr}").{kind}.Color = {expr}')
    lines.append("End Sub")
    return "\n".join(lines) + "\n"


def build_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A2"].font = openpyxl.styles.Font(bold=True)
    # 46195 is 2026-06-22 in the 1900 system.
    ws["C1"] = 46195
    ws["C1"].number_format = "m/d/yy"
    wb.save(path)


def run_and_save(driver, xlsm, out_path):
    """Runs `Paint`, then has Excel save the workbook so openpyxl can read it."""
    script = "\n".join([
        f'tell application "{driver.app_name()}"',
        "    set display alerts to false",
        "    try",
        "        close workbooks saving no",
        "    end try",
        f'    open POSIX file "{os.path.abspath(xlsm)}"',
        "    set wb to active workbook",
        '    run VB macro "Paint"',
        f'    save wb in POSIX file "{os.path.abspath(out_path)}"',
        "    close wb saving no",
        '    return "saved"',
        "end tell",
    ])
    try:
        res = subprocess.run(["osascript", "-e", script], stdout=subprocess.PIPE,
                             stderr=subprocess.PIPE, text=True, timeout=driver.timeout)
    except subprocess.TimeoutExpired:
        driver.restart()
        raise RuntimeError("Excel did not respond to the paint run")
    if res.returncode != 0:
        raise RuntimeError(f"AppleScript failed: {res.stderr.strip()}")
    return res.stdout.strip()


def paint_channel(driver, workdir):
    base = os.path.join(workdir, "paint_base.xlsx")
    xlsm = os.path.join(workdir, "paint.xlsm")
    saved = os.path.join(workdir, "painted.xlsm")
    build_workbook(base)
    wbk = visi_core.Workbook.load(base)
    wbk.add_macro("S", build_paint_module())
    wbk.save(xlsm)

    run_and_save(driver, xlsm, saved)

    wb = openpyxl.load_workbook(saved)
    ws = wb["Sheet1"]
    print("\n--- what Excel actually painted (read back with openpyxl) ---")
    print(f"{'cell':<6}{'set to':<18}{'kind':<10}{'in the file':<14}{'BGR predicts':<14}")
    mismatches = 0
    for addr, kind, expr, expect in PAINT:
        cell = ws[addr]
        if kind == "Interior":
            got = getattr(cell.fill.fgColor, "rgb", None)
        else:
            got = getattr(cell.font.color, "rgb", None)
        got = str(got) if got else "<none>"
        flag = "" if got.upper() == expect else "   <-- differs"
        if flag:
            mismatches += 1
        print(f"{addr:<6}{expr:<18}{kind:<10}{got:<14}{expect:<14}{flag}")
    print(
        f"\n{len(PAINT) - mismatches}/{len(PAINT)} match the BGR reading of VBA's Color"
    )
    return mismatches


def ask_channel(driver, workdir, batch):
    cases = [parse_case(c) for c in READ_CASES + WRITE_CASES]
    base = os.path.join(workdir, "ask_base.xlsx")
    xlsm = os.path.join(workdir, "ask.xlsm")
    build_workbook(base)
    wbk = visi_core.Workbook.load(base)
    wbk.add_macro("S", build_module(cases))
    wbk.save(xlsm)

    results = {}
    reads = list(range(1, len(READ_CASES) + 1))
    writes = list(range(len(READ_CASES) + 1, len(cases) + 1))
    for start in range(0, len(reads), batch):
        chunk = reads[start:start + batch]
        got = driver.run_batch(xlsm, chunk)
        if not got:
            print(f"cases {chunk[0]}-{chunk[-1]}: Excel returned nothing "
                  "(a compile error in one of them).", file=sys.stderr)
        results.update(got)
    for start in range(0, len(writes), batch):
        chunk = writes[start:start + batch]
        got = driver.run_batch(xlsm, chunk)
        if not got:
            print(f"cases {chunk[0]}-{chunk[-1]}: Excel returned nothing.", file=sys.stderr)
        results.update(got)

    all_cases = READ_CASES + WRITE_CASES
    width = min(max(len(c) for c in all_cases), 86)
    print("--- what Excel's object model reports ---")
    for i, text in enumerate(all_cases, start=1):
        print(f"{text:<{width}}  {results.get(i, '<missing>')}")


def palette_channel(driver, workdir, batch):
    """What colour each of Excel's 56 `ColorIndex` slots actually is."""
    cases = [parse_case(c) for c in PALETTE_CASES]
    base = os.path.join(workdir, "pal_base.xlsx")
    xlsm = os.path.join(workdir, "pal.xlsm")
    build_workbook(base)
    wbk = visi_core.Workbook.load(base)
    wbk.add_macro("S", build_module(cases))
    wbk.save(xlsm)

    results = {}
    idx = list(range(1, len(cases) + 1))
    for start in range(0, len(idx), batch):
        chunk = idx[start:start + batch]
        got = driver.run_batch(xlsm, chunk)
        if not got:
            print(f"cases {chunk[0]}-{chunk[-1]}: Excel returned nothing.", file=sys.stderr)
        results.update(got)

    print("--- ColorIndex slot -> Color (BGR Long) -> #RRGGBB ---")
    out = []
    for i in range(1, len(cases) + 1):
        raw = results.get(i, "")
        value = raw.split("|")[-1] if "|" in raw else raw
        try:
            bgr = int(value)
        except ValueError:
            print(f"{i:>3}  {raw}")
            continue
        r, g, b = bgr & 0xFF, (bgr >> 8) & 0xFF, (bgr >> 16) & 0xFF
        out.append((i, bgr, f"#{r:02X}{g:02X}{b:02X}"))
        print(f"{i:>3}  {bgr:>9}  #{r:02X}{g:02X}{b:02X}")
    print("\nas a Rust table:")
    print("    " + ", ".join(f'"{hexcode}"' for _, _, hexcode in out))


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--excel-path")
    ap.add_argument("--driver", choices=["auto", "applescript", "mock"], default="auto")
    ap.add_argument("--timeout", type=int, default=180)
    ap.add_argument("--batch", type=int, default=10)
    ap.add_argument("--ask", action="store_true", help="only the object-model channel")
    ap.add_argument("--paint", action="store_true", help="only the saved-file channel")
    ap.add_argument("--palette", action="store_true", help="only the ColorIndex palette")
    args = ap.parse_args()

    driver = ExcelDriver(args.excel_path, args.driver, args.timeout)
    workdir = tempfile.mkdtemp(prefix="vba_style_probe_")

    if args.palette:
        palette_channel(driver, workdir, args.batch)
        return 0

    both = not (args.ask or args.paint)
    if args.ask or both:
        ask_channel(driver, workdir, args.batch)
    if args.paint or both:
        return 1 if paint_channel(driver, workdir) else 0
    return 0


if __name__ == "__main__":
    sys.exit(main())
