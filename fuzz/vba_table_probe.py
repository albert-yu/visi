#!/usr/bin/env python3
"""
What does Excel's `ListObjects` surface actually do?
===================================================
Phase 3 of `docs/vba-macro-support.md` maps `ListObjects` onto `ExcelTable`.
Three things have to be measured rather than assumed:

* `ListObject.Name = "X"` is **not** a field write -- names are unique
  workbook-wide and a rename cascades into formula *text* across the whole
  workbook. What does Excel do to a `Sales[Amount]` formula when the table is
  renamed, and what does it do when the new name is already taken?
* `.DataBodyRange` on a table with **zero data rows**. Is it `Nothing`, an error, or an empty range?
* `.ListRows.Add` has to interact with the table's extent and with the
  header/totals flags.

    source fuzz/venv/bin/activate
    maturin develop -m visi-python/Cargo.toml --release
    python fuzz/vba_table_probe.py
    python fuzz/vba_table_probe.py --empty     # the zero-data-row fixture, alone

`--empty` runs against a separate workbook on purpose. A table whose `ref`
covers only its header row is exactly the shape that can make Excel show a
*repair* dialog on open -- which is modal, so it hangs the AppleScript bridge
the same way a compile error does. Keeping it out of the main fixture means a
bad guess there costs one run rather than all of them.

Same traps as the other VBA probes: a compile error is not catchable by the
`On Error` wrapper, so a batch that returns nothing is a compile error rather
than a slow run.
"""

import argparse
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    from openpyxl.worksheet.table import Table
except ImportError:
    sys.exit("openpyxl is required: source fuzz/venv/bin/activate && pip install -r fuzz/requirements.txt")

try:
    import visi_core
except ImportError:
    sys.exit(
        "the visi_core bindings are required: "
        "maturin develop -m visi-python/Cargo.toml --release"
    )

from fuzz_vba import HARNESS_TEMPLATE, ExcelDriver

PREAMBLE = [
    "Dim ws As Worksheet, wb As Workbook, lo As ListObject",
    "Set wb = ThisWorkbook",
    'Set ws = wb.Worksheets("Sheet1")',
    "Dim v, s, c",
]

READ_CASES = [
    # --- the objects
    "TypeName(ws.ListObjects)",
    "TypeName(ws.ListObjects(1))",
    'TypeName(ws.ListObjects("Sales"))',
    "CStr(ws.ListObjects.Count)",
    "ws.ListObjects(1).Name",
    # --- the ranges. `Sales` is A1:C4 with a header row and no totals row.
    "ws.ListObjects(1).Range.Address",
    "ws.ListObjects(1).HeaderRowRange.Address",
    "ws.ListObjects(1).DataBodyRange.Address",
    "TypeName(ws.ListObjects(1).TotalsRowRange)",
    "CStr(ws.ListObjects(1).ShowTotals)",
    "CStr(ws.ListObjects(1).ShowHeaders)",
    # --- columns
    "TypeName(ws.ListObjects(1).ListColumns)",
    "CStr(ws.ListObjects(1).ListColumns.Count)",
    "ws.ListObjects(1).ListColumns(1).Name",
    "ws.ListObjects(1).ListColumns(3).Name",
    "ws.ListObjects(1).ListColumns(3).Range.Address",
    "ws.ListObjects(1).ListColumns(3).DataBodyRange.Address",
    'ws.ListObjects(1).ListColumns("Amount").Range.Address',
    "CStr(ws.ListObjects(1).ListColumns(3).Index)",
    # --- rows
    "TypeName(ws.ListObjects(1).ListRows)",
    "CStr(ws.ListObjects(1).ListRows.Count)",
    "ws.ListObjects(1).ListRows(1).Range.Address",
    # --- how a table is reached from the workbook and from a range
    'TypeName(ws.Range("A2").ListObject)',
    'ws.Range("A2").ListObject.Name',
    'TypeName(ws.Range("G1").ListObject)',
    # --- errors
    'ws.ListObjects("nope").Name',
    "ws.ListObjects(5).Name",
    "ws.ListObjects(1).ListColumns(9).Name",
    # --- the formula that references the table, before any rename
    "ws.Range(\"E1\").Formula",
    "CStr(ws.Range(\"E1\").Value)",
]

WRITE_CASES = [
    # --- ListRows.Add: what happens to the extent, and where the new row is
    "Set lo = ws.ListObjects(1)\\nlo.ListRows.Add :: lo.Range.Address & \"|\" & CStr(lo.ListRows.Count)",
    "Set lo = ws.ListObjects(1)\\nSet v = lo.ListRows.Add :: TypeName(v)",
    # --- renaming, and the cascade into formula text
    'ws.ListObjects(1).Name = "Revenue" :: ws.ListObjects(1).Name & "|" & ws.Range("E1").Formula',
    # a name that is already taken workbook-wide
    'ws.ListObjects(1).Name = "Other" :: "no error"',
    # --- renaming a column, and its cascade
    'ws.ListObjects(1).ListColumns(3).Name = "Total" :: ws.ListObjects(1).ListColumns(3).Name & "|" & ws.Range("E1").Formula',
    # --- writing through the ranges
    'ws.ListObjects(1).DataBodyRange.Cells(1, 3).Value = 999 :: CStr(ws.Range("C2").Value)',
    # --- ShowTotals, which changes the extent
    "Set lo = ws.ListObjects(1)\\nlo.ShowTotals = True :: lo.Range.Address & \"|\" & lo.TotalsRowRange.Address",
]

# A table with zero data rows.
#
# The obvious fixture -- a table whose `ref` covers only its header row --
# does **not** work: Excel treats that file as damaged and opens a modal
# repair dialog, which hangs the bridge exactly as a compile error does
# (measured, the hard way). So the fixture is a normal one-data-row table and
# the row is deleted *from VBA*, which reaches the same state by a route
# Excel itself produces.
EMPTY_CASES = [
    "ws.ListObjects(1).ListRows(1).Delete :: CStr(ws.ListObjects.Count)",
    "ws.ListObjects(1).ListRows(1).Delete :: ws.ListObjects(1).Range.Address",
    "ws.ListObjects(1).ListRows(1).Delete :: ws.ListObjects(1).HeaderRowRange.Address",
    "ws.ListObjects(1).ListRows(1).Delete :: TypeName(ws.ListObjects(1).DataBodyRange)",
    "ws.ListObjects(1).ListRows(1).Delete :: CStr(ws.ListObjects(1).DataBodyRange Is Nothing)",
    "ws.ListObjects(1).ListRows(1).Delete :: CStr(ws.ListObjects(1).ListRows.Count)",
    "ws.ListObjects(1).ListRows(1).Delete :: CStr(ws.ListObjects(1).ListColumns.Count)",
    "ws.ListObjects(1).ListRows(1).Delete :: TypeName(ws.ListObjects(1).ListColumns(1).DataBodyRange)",
    # and adding the first row back to an emptied table
    "ws.ListObjects(1).ListRows(1).Delete\\nSet lo = ws.ListObjects(1)\\nlo.ListRows.Add :: lo.Range.Address & \"|\" & lo.DataBodyRange.Address",
]


def parse_case(text):
    setup, sep, expr = text.rpartition("::")
    if not sep:
        return [], text.strip()
    return [s.strip() for s in setup.split("\\n") if s.strip()], expr.strip()


def build_module(cases):
    parts = ['Attribute VB_Name = "T"']
    for i, (setup, expr) in enumerate(cases, start=1):
        body = "\n".join(f"    {s}" for s in PREAMBLE + setup)
        parts.append(f"Private Function Gen{i}()\n{body}\n    Gen{i} = {expr}\nEnd Function")
        parts.append(HARNESS_TEMPLATE.format(i=i))
    return "\n\n".join(parts) + "\n"


def build_workbook(path):
    """`Sales` over A1:C4 (header + 3 data rows), plus a second table so the
    duplicate-name case has something to collide with, plus a formula that
    references the table by name so a rename's cascade is observable."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    rows = [
        ("Region", "Product", "Amount"),
        ("East", "Widget", 10),
        ("West", "Gadget", 20),
        ("North", "Widget", 30),
    ]
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    ws.add_table(Table(displayName="Sales", ref="A1:C4"))

    # A second table, purely so `Name = "Other"` is a real collision.
    ws["A8"], ws["B8"] = "Key", "Val"
    ws["A9"], ws["B9"] = "k", 1
    ws.add_table(Table(displayName="Other", ref="A8:B9"))

    ws["E1"] = "=SUM(Sales[Amount])"
    wb.save(path)


def build_empty_workbook(path):
    """One header row and one data row, which each case then deletes.

    Authoring the zero-row state directly makes Excel offer to repair the
    file; see `EMPTY_CASES`."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c, v in enumerate(("Region", "Product", "Amount"), start=1):
        ws.cell(row=1, column=c, value=v)
    for c, v in enumerate(("East", "Widget", 10), start=1):
        ws.cell(row=2, column=c, value=v)
    ws.add_table(Table(displayName="Hollow", ref="A1:C2"))
    wb.save(path)


def run(driver, cases, build, batch, label):
    parsed = [parse_case(c) for c in cases]
    workdir = tempfile.mkdtemp(prefix="vba_table_probe_")
    base = os.path.join(workdir, "base.xlsx")
    xlsm = os.path.join(workdir, "probe.xlsm")
    build(base)
    # Re-read what we just wrote before handing it to Excel: an unopenable
    # workbook is a modal repair dialog, which hangs the bridge.
    openpyxl.load_workbook(base)
    wbk = visi_core.Workbook.load(base)
    wbk.add_macro("T", build_module(parsed))
    wbk.save(xlsm)

    results = {}
    idx = list(range(1, len(parsed) + 1))
    for start in range(0, len(idx), batch):
        chunk = idx[start:start + batch]
        got = driver.run_batch(xlsm, chunk)
        if not got:
            print(f"cases {chunk[0]}-{chunk[-1]}: Excel returned nothing "
                  "(a compile error, or a modal dialog).", file=sys.stderr)
        results.update(got)

    width = min(max(len(c) for c in cases), 80)
    print(f"--- {label} ---")
    for i, text in enumerate(cases, start=1):
        print(f"{text:<{width}}  {results.get(i, '<missing>')}")


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--excel-path")
    ap.add_argument("--driver", choices=["auto", "applescript", "mock"], default="auto")
    ap.add_argument("--timeout", type=int, default=180)
    ap.add_argument("--batch", type=int, default=8)
    ap.add_argument("--empty", action="store_true",
                    help="only the zero-data-row fixture (issue #11's shape)")
    args = ap.parse_args()

    driver = ExcelDriver(args.excel_path, args.driver, args.timeout)
    if args.empty:
        # One case per round trip, forced: every case here deletes the
        # table's only row, so two sharing a session would have the second
        # one delete nothing and report error 9 rather than measuring
        # anything. A round trip re-opens the workbook, which is the reset.
        run(driver, EMPTY_CASES, build_empty_workbook, 1, "zero data rows")
        return 0
    # Writes go last and in one chunk: several read back what an earlier one
    # did, and a batch boundary re-opens the workbook.
    run(driver, READ_CASES, build_workbook, args.batch, "reads")
    run(driver, WRITE_CASES, build_workbook, len(WRITE_CASES), "writes")
    return 0


if __name__ == "__main__":
    sys.exit(main())
