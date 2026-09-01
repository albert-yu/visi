#!/usr/bin/env python3
"""
visi vs. Microsoft Excel Differential Fuzzing: Pivot Tables
=============================================================
Generates a random source workbook (a mix of low-cardinality categorical
columns and numeric columns, suited to exercising pivot grouping), builds a
matching pivot table configuration in both `visi` (via the `visi pivot` CLI)
and real Microsoft Excel (by driving Excel's own PivotTable object model via
AppleScript/COM automation -- Excel does not auto-generate a pivot table from
XML the way it recalculates formulas, so this is a fundamentally different
execution pipeline from `fuzz_excel.py`'s "generate formulas -> recalculate"
flow, which is why this lives in its own script), then compares the two
engines' rendered pivot output cell-for-cell.

This reuses (imports, does not duplicate) `XLSXEvaluatedReader` and
`DifferentialComparator` from `fuzz_excel.py` -- both engines materialize
pivot output as plain literal cell values, so the existing generic cell
reader/comparator works unchanged.

IMPORTANT -- the AppleScript path does NOT build the pivot table via Excel's
native AppleScript pivot dictionary. `make new pivot cache at wb` is declared
in Excel.sdef but fails with a generic "Parameter error (-50)" for every
variant tried against real Excel (bare, with properties, range object vs.
text source data, different containers) -- confirmed to be a real
functional gap in Mac Excel's AppleScript support, not a syntax mistake.
Instead, the AppleScript driver opens a macro-enabled copy of the source
workbook carrying `fuzz/BuildFuzzPivot.bas` and invokes its `BuildFuzzPivot`
macro via the `run VB macro` command, which *is* a working, standard
AppleScript command -- the macro itself does the real work through VBA's
PivotCaches.Create / CreatePivotTable, the same well-documented object model
win32com already uses directly.

Setup: none. `fuzz/pivot_macro_template.xlsm` is built automatically on first
use by `visi macro add`, which writes the module into `vbaProject.bin` at the
file-format level. This used to be a one-time manual ritual (open the VBA
editor, paste the .bas in, Save As .xlsm) because Excel for Mac's AppleScript
dictionary exposes no VBProject object, so nothing could put a macro into a
workbook programmatically -- `visi`'s own macro CRUD is what removed the
human step. See `fuzz/vba_probe.py` for the checks that establish Excel runs
visi-authored macros exactly like its own.

The generated template stays *empty* of data: openpyxl copies each iteration's
random source rows into a copy of it, so the Excel oracle's data path never
passes through visi's xlsx writer -- only the inert VBA skeleton does.

Built through the `visi_core` bindings when they're installed, falling back to
the CLI otherwise -- so `--backend bindings` needs nothing extra.

STATUS -- piloted against real Excel, works end-to-end. Every finding from
that pilot (fixed and still-open alike) is tracked as a GitHub issue rather
than duplicated here -- see fuzz/README.md's "Known caveats" section and the
repo's issue tracker for detail.

Usage:
    python3 fuzz/fuzz_pivot.py --driver mock --iterations 5
    python3 fuzz/fuzz_pivot.py --excel-path "/Applications/Microsoft Excel.app" --iterations 1 --seed 1
"""

import argparse
import os
import random
import shutil
import subprocess
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from fuzz_excel import (  # noqa: E402
    SMOKE_BANNER,
    DifferentialComparator,
    XLSXEvaluatedReader,
    smoke_check,
)


# -----------------------------------------------------------------------------
# 1. Source workbook + pivot configuration generator
# -----------------------------------------------------------------------------


class PivotFuzzGenerator:
    """Generates a random source workbook plus a matching pivot table
    configuration (as a plain dict, not XML) that both `VisiPivotDriver` and
    `ExcelPivotDriver` build a real pivot table from.

    Columns are fixed and deliberately low-cardinality where it matters for
    grouping, unlike `fuzz_excel.py`'s fully-random grid -- a pivot fuzzer
    that used high-cardinality random text for row/col fields would turn
    every group into a singleton and never exercise subtotal/grand-total
    logic at all.
    """

    BASIC_COL_NAMES = ["Cat", "Mixed", "NumStr", "Amount", "Rate", "Flag"]
    RICH_COL_NAMES = ["Cat", "Mixed", "NumStr", "Amount", "Rate", "Flag", "DateBucket", "Segment", "Amount2"]
    COL_NAMES = BASIC_COL_NAMES
    CATEGORICAL_COLS = [0, 1, 2]  # Cat, Mixed, NumStr -- candidates for row/col fields
    NUMERIC_COLS = [3, 4]  # Amount, Rate -- candidates for value fields
    FILTERABLE_COLS = [0, 1, 2, 5]  # any column can be a filter field

    CATEGORIES = ["Alpha", "Beta", "Gamma", "Delta", "Epsilon"]
    # Same values in different cases, to probe case-insensitive grouping
    # parity between visi and Excel.
    CASE_VARIANTS = ["East", "east", "WEST", "west", "North"]
    AGGREGATIONS = ["sum", "count", "count-numbers", "average", "max", "min"]

    def __init__(self, seed=None, shape="basic"):
        if seed is not None:
            random.seed(seed)
        self.shape = shape
        if shape == "rich":
            self.COL_NAMES = self.RICH_COL_NAMES
            # Keep the pass/fail pivot configuration on the same stable field
            # families as the basic fuzzer, while the richer source columns sit
            # beside them for import/export and cache-shape coverage.
            self.CATEGORICAL_COLS = [0, 1, 2]
            self.NUMERIC_COLS = [3, 4]
            self.FILTERABLE_COLS = [0, 1, 2, 5]
        else:
            self.COL_NAMES = self.BASIC_COL_NAMES
            self.CATEGORICAL_COLS = [0, 1, 2]
            self.NUMERIC_COLS = [3, 4]
            self.FILTERABLE_COLS = [0, 1, 2, 5]

    def _random_numstr(self):
        """A quoted (forced-text), possibly numeric-looking string, or a
        blank -- probes the sort/group-key numeric-vs-text ambiguity in
        visi's `sort_group_entries` (pivot.rs)."""
        roll = random.random()
        if roll < 0.25:
            return None
        if roll < 0.5:
            return f"0{random.randint(0, 9)}"  # e.g. "08"
        if roll < 0.75:
            return f".0{random.randint(0, 999)}"  # e.g. ".0394"
        return str(random.randint(-50, 50))

    def generate(self, source_path, num_rows, use_table):
        """Builds `source_path` and returns a pivot configuration dict:
        {source_range, table_name, row_fields, col_fields, value_fields,
         filter_field, grand_totals_row, grand_totals_col}.
        `table_name` is None when `use_table` is False (raw-range source).
        """
        try:
            import openpyxl
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            print("Error: 'openpyxl' is required for generating .xlsx test files.")
            print("Please run: pip install openpyxl")
            sys.exit(1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for c, name in enumerate(self.COL_NAMES, start=1):
            ws.cell(row=1, column=c, value=name)

        # Track each filterable column's actual distinct (blank-normalized)
        # values as we generate, so the filter field picked below can select
        # a real subset instead of guessing at what exists. Both visi and
        # Excel merge case-different text into a single pivot item keyed by
        # whichever casing appeared first in the source (see pivot.rs's
        # `test_case_variant_values_merge_using_globally_first_seen_casing`)
        # -- a real Excel user filtering via the UI only ever sees one
        # checkbox per merged item, under that one casing. So `distinct`
        # must canonicalize to that same first-seen casing per
        # case-insensitive key, or the filter field below could pick two
        # different casings of what's actually a single merged item (e.g.
        # both "East" and "east") as if they were independently
        # selectable -- an unrepresentable config that made real Excel and
        # visi diverge for a reason that has nothing to do with either
        # engine's pivot correctness (fuzz/fuzz_pivot.py iteration 8, seed
        # 599783: selecting "east" alone failed to match the merged
        # "East"/"east" group in real Excel, since VBA's PivotItem.Name
        # comparison is case-sensitive and the group's canonical name was
        # "East").
        canonical_casing = {}

        def canonicalize(value):
            key = value.lower()
            if key not in canonical_casing:
                canonical_casing[key] = value
            return canonical_casing[key]

        distinct = {i: set() for i in self.FILTERABLE_COLS}
        for r in range(2, num_rows + 2):
            cat = "" if random.random() < 0.1 else random.choice(self.CATEGORIES)
            ws.cell(row=r, column=1, value=cat)
            distinct[0].add(canonicalize(cat) if cat else "(blank)")

            mixed = random.choice(self.CASE_VARIANTS)
            ws.cell(row=r, column=2, value=mixed)
            distinct[1].add(canonicalize(mixed))

            numstr = self._random_numstr()
            ws.cell(row=r, column=3, value=numstr)
            distinct[2].add(numstr if numstr else "(blank)")

            ws.cell(row=r, column=4, value=random.randint(-100, 100))
            ws.cell(row=r, column=5, value=round(random.uniform(-500.0, 500.0), 4))

            flag = random.choice([True, False])
            ws.cell(row=r, column=6, value=flag)
            distinct[5].add("TRUE" if flag else "FALSE")

            if self.shape == "rich":
                date_bucket = random.choice(["2026-Q1", "2026-Q2", "2026-Q3", "2026-Q4", "(blank)"])
                segment = random.choice(["Retail", "Enterprise", "Online", "Partner"])
                ws.cell(row=r, column=7, value=None if date_bucket == "(blank)" else date_bucket)
                ws.cell(row=r, column=8, value=segment)
                ws.cell(row=r, column=9, value=random.randint(0, 500) if random.random() > 0.15 else None)

        last_col = len(self.COL_NAMES)
        source_range = f"A1:{chr(ord('A') + last_col - 1)}{num_rows + 1}"
        table_name = None
        if use_table:
            table_name = "FuzzTable"
            table = Table(displayName=table_name, ref=source_range)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(table)

        wb.save(source_path)

        pool = self.CATEGORICAL_COLS[:]
        random.shuffle(pool)
        n_row = random.randint(0, min(2, len(pool)))
        row_cols, pool = pool[:n_row], pool[n_row:]
        n_col = random.randint(0, min(2, len(pool)))
        col_cols, pool = pool[:n_col], pool[n_col:]

        row_fields = [
            {"column": self.COL_NAMES[i], "subtotal": random.random() < 0.7} for i in row_cols
        ]
        col_fields = [
            {"column": self.COL_NAMES[i], "subtotal": random.random() < 0.7} for i in col_cols
        ]

        n_value = random.randint(1, 2)
        value_fields = [
            {
                "column": self.COL_NAMES[random.choice(self.NUMERIC_COLS)],
                "agg": random.choice(self.AGGREGATIONS),
            }
            for _ in range(n_value)
        ]

        filter_field = None
        if random.random() < 0.5:
            fcol = random.choice(self.FILTERABLE_COLS)
            values = sorted(distinct[fcol])
            # Always at least one selected value. An empty selection means
            # "select nothing", which real Excel cannot represent -- it
            # refuses to hide a page field's last visible PivotItem (runtime
            # error 1004), so BuildFuzzPivot.bas falls back to leaving the
            # field unfiltered at "(All)". Emitting one made visi render an
            # empty grid against Excel's full one and report every cell of it
            # as a mismatch (iteration 5, seed 244209) -- an unrepresentable
            # config, not an engine disagreement, the same class of bogus
            # failure the `canonicalize` comment above describes. `values` is
            # never empty itself: every filterable column gets a value (or
            # "(blank)") on every one of the >= 1 generated rows.
            selected = [v for v in values if random.random() < 0.5]
            if not selected:
                selected = [random.choice(values)]
            filter_field = {"column": self.COL_NAMES[fcol], "values": selected}

        return {
            "source_range": source_range,
            # The same block as 0-based inclusive (start_row, start_col,
            # end_row, end_col), for the bindings backend, which takes indices
            # rather than A1. Derived here beside `source_range` so the two
            # cannot disagree, and so neither the driver nor visi-python needs
            # an A1 parser of its own.
            "source_bounds": (0, 0, num_rows, last_col - 1),
            "table_name": table_name,
            "row_fields": row_fields,
            "col_fields": col_fields,
            "value_fields": value_fields,
            "filter_field": filter_field,
            "grand_totals_row": random.random() < 0.7,
            "grand_totals_col": random.random() < 0.7,
        }


# -----------------------------------------------------------------------------
# 2. Execution drivers
# -----------------------------------------------------------------------------

PIVOT_NAME = "FuzzPivot"
DEST_CELL = "H1"  # two columns clear of the source block (A:F)
DEST_RC = (0, 7)  # DEST_CELL as 0-based (row, col); must agree with it
# Macro-enabled workbook carrying BuildFuzzPivot.bas, generated on first use
# by `visi macro add` (see ExcelPivotDriver._ensure_macro_template). Cached on
# disk rather than rebuilt per iteration -- it's identical every time, and the
# AppleScript round trip already dominates the per-iteration cost. Gitignored:
# it's a build artifact derived from BuildFuzzPivot.bas, which *is* checked in.
MACRO_TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pivot_macro_template.xlsm")
MACRO_SOURCE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "BuildFuzzPivot.bas")

# visi's `Count` aggregation counts any non-blank value -- matches Excel's
# `xlCount` (the *default* summary function Excel picks for a text field,
# labeled plain "Count" in the value-field-settings UI); visi's
# `CountNumbers` counts only numeric values -- matches Excel's `xlCountNums`.
# There is no separate "CountA" member in Excel's `XlConsolidationFunction`
# enum (confirmed via Excel.sdef). `BuildFuzzPivot.bas` (invoked by the
# AppleScript path) matches these same keys directly in its `Select Case`,
# so only the win32com path needs its own VBA constant-name mapping.
AGG_TO_WIN32COM_FUNCTION = {
    "sum": "xlSum",
    "count": "xlCount",
    "count-numbers": "xlCountNums",
    "average": "xlAverage",
    "max": "xlMax",
    "min": "xlMin",
}


# `VisiPivotDriver` moved to visi_driver.py, which drives visi either through
# the in-process `visi_core` bindings or through the `visi pivot` CLI.
from visi_driver import (  # noqa: E402,F401
    VisiPivotDriver,
    add_backend_arg,
    bindings_hint,
)


class ExcelPivotDriver:
    """Drives Microsoft Excel's own PivotTable object model to build a
    matching pivot table, then saves. Unlike `fuzz_excel.py`'s `ExcelDriver`
    (which only needs `calculate` over cells visi already computed), Excel
    must *construct* a live PivotTable here -- there's no XML shortcut.

    The win32com path (Windows) uses the standard, well-documented VBA
    object model directly and is straightforward.

    The AppleScript path (macOS) cannot do the same directly: `make new
    pivot cache at wb` is declared in Excel.sdef (extracted straight from
    `Microsoft Excel.app/Contents/Resources/Excel.sdef`, since the `sdef`
    CLI tool needs a full Xcode install this environment doesn't have) but
    fails with a generic "Parameter error (-50)" for every variant tried
    against real Excel -- bare, with properties, range object vs. text
    source data, different containers. That's a real functional gap in Mac
    Excel's AppleScript support (the dictionary documents a capability the
    implementation doesn't back), not a syntax mistake -- other pivot
    field/property names below (`pivot field orientation`, `orient as row
    field`, `layout form`, `layout subtotal location`, `set subtotals`,
    `range object` of a list object) were all verified against the same
    .sdef and do work once a PivotTable already exists.

    So instead, the AppleScript path builds the pivot through a macro
    (`fuzz/BuildFuzzPivot.bas`, injected into `fuzz/pivot_macro_template.xlsm`
    by `visi macro add` on first use) invoked via `run VB macro`, which *is* a
    working AppleScript command. The macro itself uses the same
    PivotCaches.Create / CreatePivotTable / PivotFields object model as the
    win32com path below, just reached through VBA instead of AppleScript's
    broken pivot-cache creation.
    """

    def __init__(self, excel_path=None, driver_type="auto", visi_path=None):
        self.excel_path = excel_path
        # Only used to build the macro template (AppleScript path); the
        # comparison itself never goes through the CLI.
        self.visi_path = visi_path
        self.driver_type = driver_type
        if driver_type == "auto":
            if sys.platform == "darwin":
                self.driver_type = "applescript"
            elif sys.platform == "win32":
                self.driver_type = "win32com"
            else:
                self.driver_type = "mock"

    def run(self, source_file, config, output_file, dest_cell=DEST_CELL):
        if self.driver_type == "mock":
            shutil.copyfile(source_file, output_file)
            print("[ExcelPivotDriver Warning] Running in mock mode (Excel not invoked).")
            return
        elif self.driver_type == "applescript":
            # Excel needs a macro-enabled (.xlsm) copy to run BuildFuzzPivot
            # from; build one alongside the requested output path, then
            # copy its bytes back so callers don't need to know about the
            # extension difference.
            macro_file = os.path.splitext(output_file)[0] + ".xlsm"
            self._prepare_macro_workbook(source_file, config, macro_file)
            self._run_applescript_macro(os.path.abspath(macro_file), config, dest_cell)
            shutil.copyfile(macro_file, output_file)
        elif self.driver_type == "win32com":
            shutil.copyfile(source_file, output_file)
            self._run_win32com(os.path.abspath(output_file), config, dest_cell)
        else:
            raise RuntimeError(f"Unsupported pivot driver type: {self.driver_type}")

    # -- AppleScript (macOS) --------------------------------------------

    def _applescript_str(self, s):
        escaped = s.replace("\\", "\\\\").replace('"', '\\"')
        return f'"{escaped}"'

    def _ensure_macro_template(self):
        """Builds `pivot_macro_template.xlsm` from `BuildFuzzPivot.bas` if it
        isn't already there, by way of `visi macro add`.

        Rebuilds whenever the .bas is newer than the .xlsm, so editing the
        macro can't silently leave a stale template in play -- that failure
        mode (macro edited, template not regenerated, mismatches blamed on the
        engine) is exactly what the old manual setup step invited.

        The template deliberately contains no data: `_prepare_macro_workbook`
        copies each iteration's rows in via openpyxl, so nothing on the Excel
        oracle's side of the comparison is round-tripped through visi's own
        xlsx writer.
        """
        if not os.path.exists(MACRO_SOURCE_PATH):
            raise RuntimeError(f"Missing {MACRO_SOURCE_PATH} -- it should be checked in.")
        if (os.path.exists(MACRO_TEMPLATE_PATH)
                and os.path.getmtime(MACRO_TEMPLATE_PATH) >= os.path.getmtime(MACRO_SOURCE_PATH)):
            return

        with open(MACRO_SOURCE_PATH) as f:
            source = f.read()

        # The module name must match the .bas's own `Attribute VB_Name` line
        # (visi writes the source verbatim and does not reconcile the two).
        # It's the *module* name; `run VB macro` invokes the procedure name,
        # `BuildFuzzPivot`, which is deliberately distinct from it.
        via = self._build_macro_template_via_bindings(source)
        if via is None:
            via = self._build_macro_template_via_cli()
        print(f"[ExcelPivotDriver] Built {os.path.basename(MACRO_TEMPLATE_PATH)} from "
              f"{os.path.basename(MACRO_SOURCE_PATH)} via {via}.")

    def _build_macro_template_via_bindings(self, source):
        """Returns a description of what it used, or None if unavailable."""
        try:
            import visi_core
        except ImportError:
            return None
        wb = visi_core.Workbook()
        wb.add_macro("Module1", source)
        wb.save(MACRO_TEMPLATE_PATH)
        return "the visi_core bindings"

    def _build_macro_template_via_cli(self):
        visi = self.visi_path or "./target/release/visi"
        if not (shutil.which(visi) or os.path.exists(visi)):
            raise RuntimeError(
                f"Building {os.path.basename(MACRO_TEMPLATE_PATH)} needs either the "
                f"visi_core bindings (`maturin develop -m visi-python/Cargo.toml --release`) "
                f"or the visi CLI, and neither is available ({visi!r} does not exist -- "
                "run `cargo build --release`, or pass --visi-path)."
            )

        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            base = os.path.join(tmp, "base.xlsx")
            wb = openpyxl.Workbook()
            wb.active.title = "Sheet1"
            wb.save(base)
            res = subprocess.run(
                [visi, "macro", "add", base, "--name", "Module1",
                 "--kind", "standard", "--source-file", MACRO_SOURCE_PATH,
                 "--output", MACRO_TEMPLATE_PATH],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
            )
        if res.returncode != 0:
            raise RuntimeError(
                f"`visi macro add` failed to build {MACRO_TEMPLATE_PATH}: {res.stderr.strip()}"
            )
        return "`visi macro add`"

    def _prepare_macro_workbook(self, source_file, config, macro_file):
        self._ensure_macro_template()
        import openpyxl
        from openpyxl.worksheet.table import Table, TableStyleInfo

        src_wb = openpyxl.load_workbook(source_file)
        src_ws = src_wb["Sheet1"]

        macro_wb = openpyxl.load_workbook(MACRO_TEMPLATE_PATH, keep_vba=True)
        macro_ws = macro_wb.active
        macro_ws.title = "Sheet1"
        for row in src_ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    macro_ws.cell(row=cell.row, column=cell.column, value=cell.value)

        if config["table_name"]:
            table = Table(displayName=config["table_name"], ref=config["source_range"])
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            macro_ws.add_table(table)

        macro_wb.save(macro_file)
        # openpyxl's `keep_vba=True` path stashes a raw copy of the VBA
        # project's zip parts in `macro_wb.vba_archive` (an in-memory
        # ZipFile) but never closes it itself. Left unclosed, it only gets
        # reclaimed by Python's *cyclic* garbage collector (Workbook's
        # worksheet<->parent references form a reference cycle, so simple
        # refcounting alone never frees it) -- and the cyclic collector may
        # clear its underlying BytesIO buffer before running the ZipFile's
        # own __del__/close() finalizer, which then raises "ValueError: I/O
        # operation on closed file" from inside that finalizer. Python
        # reports this as a harmless-but-noisy "Exception ignored while
        # calling deallocator" on stderr -- it doesn't affect the saved
        # file or the fuzzer's result, but closing it explicitly here (while
        # its BytesIO is still guaranteed live) avoids the race entirely.
        if macro_wb.vba_archive is not None:
            macro_wb.vba_archive.close()

    def _build_applescript_macro_call(self, output_file, config, dest_cell=DEST_CELL):
        app_name = self.excel_path if self.excel_path else "Microsoft Excel"
        if app_name.endswith(".app"):
            app_name = os.path.splitext(os.path.basename(app_name))[0]

        row_fields_csv = ";".join(
            f'{f["column"]}:{"1" if f["subtotal"] else "0"}' for f in config["row_fields"]
        )
        col_fields_csv = ";".join(
            f'{f["column"]}:{"1" if f["subtotal"] else "0"}' for f in config["col_fields"]
        )
        value_fields_csv = ";".join(f'{f["column"]}:{f["agg"]}' for f in config["value_fields"])
        if config["filter_field"]:
            filter_spec = config["filter_field"]["column"] + "|" + ",".join(config["filter_field"]["values"])
        else:
            filter_spec = ""
        source_is_table = "1" if config["table_name"] else "0"
        source_ref = config["table_name"] if config["table_name"] else config["source_range"]

        # Labeled command parameters in this dictionary use space-separated
        # syntax (`label value`), not `label:value` colons -- confirmed via
        # the same real-Excel trial-and-error that found `create pivot
        # table`'s working form (see class docstring); colons here produce
        # a plain syntax error, not a runtime one.
        run_macro_line = (
            f'        run VB macro "BuildFuzzPivot" '
            f'arg1 {self._applescript_str(row_fields_csv)} '
            f'arg2 {self._applescript_str(col_fields_csv)} '
            f'arg3 {self._applescript_str(value_fields_csv)} '
            f'arg4 {self._applescript_str(filter_spec)} '
            f'arg5 {self._applescript_str(dest_cell)} '
            f'arg6 {self._applescript_str("1" if config["grand_totals_row"] else "0")} '
            f'arg7 {self._applescript_str("1" if config["grand_totals_col"] else "0")} '
            f'arg8 {self._applescript_str(source_is_table)} '
            f'arg9 {self._applescript_str(source_ref)}'
        )

        lines = [
            f'tell application "{app_name}"',
            "    set display alerts to false",
            "    try",
            "        close workbooks saving no",
            "    end try",
            "    try",
            f'        set targetFile to POSIX file "{output_file}"',
            "        open targetFile",
            "        set wb to active workbook",
            run_macro_line,
            "        close wb saving no",
            "    on error errText number errNum",
            "        try",
            "            close workbooks saving no",
            "        end try",
            "        error errText number errNum",
            "    end try",
            "end tell",
        ]
        return "\n".join(lines)

    def _restart_excel(self):
        """Force-quits and relaunches Excel entirely (not just `killall` +
        hope) -- see GitHub issue #15. `run VB macro` calls degrade into a
        session-wide, config-independent "Parameter error (-50)" after
        enough consecutive AppleScript invocations against one long-lived
        Excel process (confirmed via ~20-30 back-to-back repro calls: the
        exact same pivot config that fails deterministically for the
        remainder of that session succeeds immediately, every time, right
        after Excel is fully restarted) -- it is Excel's automation bridge
        wearing out, unrelated to any particular pivot shape (same-column
        col/filter field, zero-selection filter, single-row source, or
        otherwise). `killall` alone was observed to sometimes leave the
        process listed as still running (the app may intercept SIGTERM to
        run its own quit handshake), so this escalates to SIGKILL by PID
        before relaunching.
        """
        subprocess.run(["killall", "Microsoft Excel"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(1.0)
        pgrep = subprocess.run(["pgrep", "-x", "Microsoft Excel"], stdout=subprocess.PIPE, text=True)
        for pid in pgrep.stdout.split():
            subprocess.run(["kill", "-9", pid], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        if pgrep.stdout.strip():
            time.sleep(1.0)
        subprocess.run(["open", "-a", "Microsoft Excel"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(4.0)  # give the app time to finish launching before the next AppleScript call

    def _run_applescript_macro(self, abs_output, config, dest_cell=DEST_CELL):
        script = self._build_applescript_macro_call(abs_output, config, dest_cell)
        res = None
        for attempt in range(5):
            time.sleep(0.5)
            try:
                res = subprocess.run(
                    ["osascript", "-e", script],
                    stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=20,
                )
                if res.returncode == 0:
                    break
                # A non-timeout AppleScript failure this late in a session is
                # session degradation, not a config problem --
                # a plain retry against the same stuck process just fails
                # again, so restart Excel outright before the next attempt.
                self._restart_excel()
            except subprocess.TimeoutExpired:
                self._restart_excel()
        if res is not None and res.returncode != 0:
            raise RuntimeError(f"Excel pivot AppleScript failed:\nSTDERR: {res.stderr}\nScript:\n{script}")

    # -- win32com (Windows) -----------------------------------------------

    def _run_win32com(self, abs_output, config, dest_cell=DEST_CELL):
        try:
            import win32com.client
            from win32com.client import constants as c
        except ImportError:
            raise RuntimeError("pywin32 (win32com) is required for Excel automation on Windows.")

        # COM automation against a fresh Excel.Application is occasionally
        # flaky in a way that surfaces as unrelated-looking errors (RPC
        # server unavailable, "Call was rejected by callee", a raw OLE
        # error code) rather than a clean failure -- transient, not
        # reproducible, and not an Excel/visi disagreement. Retry with a
        # fresh Application instance, mirroring fuzz_excel.py's and
        # fuzz_chart.py's win32com drivers (see fuzz_excel.py's for the
        # "'bool' object is not callable" variant this same pattern covers
        # there; fuzz_chart.py's for the class of RPC/OLE errors this
        # pattern was added here for).
        last_err = None
        for attempt in range(5):
            if attempt > 0:
                subprocess.run(
                    ["taskkill", "/F", "/IM", "EXCEL.EXE"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                time.sleep(1.0)
            try:
                self._run_win32com_once(abs_output, config, dest_cell)
                last_err = None
                break
            except Exception as e:
                last_err = e
        if last_err is not None:
            raise last_err

    def _run_win32com_once(self, abs_output, config, dest_cell=DEST_CELL):
        import win32com.client
        from win32com.client import constants as c

        # `Dispatch` (late binding) never populates `win32com.client.constants`
        # -- that module only fills in once the Excel type library has been
        # generated, which only `gencache.EnsureDispatch` triggers. Plain
        # `Dispatch` here made every `c.xl*` lookup below raise AttributeError.
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(abs_output)
            ws = wb.Sheets("Sheet1")
            if config["table_name"]:
                src_range = ws.ListObjects(config["table_name"]).Range
            else:
                src_range = ws.Range(config["source_range"])

            pc = wb.PivotCaches().Create(SourceType=c.xlDatabase, SourceData=src_range)
            pt = pc.CreatePivotTable(TableDestination=ws.Range(dest_cell), TableName=PIVOT_NAME)

            # Mirrors BuildFuzzPivot.bas's ApplyAxisFields/ApplyValueFields --
            # this driver and that macro are meant to be the same object
            # model reached two different ways (see ExcelPivotDriver's
            # class docstring), so field-application logic here should not
            # drift from the macro's.
            #
            # Per-field LayoutForm/LayoutSubtotalLocation, not
            # PivotTable.RowAxisLayout/SubtotalLocation: the macro avoids
            # the table-wide methods because they hang Mac Excel outright.
            # On Windows they don't hang -- but unlike the per-field
            # setting (a documented no-op there, see the macro's comment),
            # they actually switch real Excel into Tabular Form, which
            # replaces the "Row Labels"/"Column Labels"/"Grand Total"
            # captions visi's pivot writer always emits with the field's
            # own name. visi has no tabular-layout support to match that
            # against, so using the table-wide methods here was producing
            # a guaranteed mismatch on every row/col-labeled cell, not a
            # real engine bug.
            for f in config["row_fields"]:
                pf = pt.PivotFields(f["column"])
                pf.Orientation = c.xlRowField
                pf.LayoutForm = c.xlTabular
                pf.LayoutSubtotalLocation = c.xlAtBottom
                pf.RepeatLabels = False
                if not f["subtotal"]:
                    # VBA's `pf.Subtotals(1) = False` is an indexed
                    # property-let with no Python equivalent syntax --
                    # win32com exposes the same property as a plain 12-bool
                    # array (index 0 = "Automatic", matching VBA's 1-based
                    # Subtotals(1)), so assign the whole array instead.
                    subtotals = [False] * 12
                    pf.Subtotals = subtotals
            for f in config["col_fields"]:
                pf = pt.PivotFields(f["column"])
                pf.Orientation = c.xlColumnField
                pf.LayoutForm = c.xlTabular
                pf.LayoutSubtotalLocation = c.xlAtBottom
                pf.RepeatLabels = False
                if not f["subtotal"]:
                    subtotals = [False] * 12
                    pf.Subtotals = subtotals
            for f in config["value_fields"]:
                pf = pt.PivotFields(f["column"])
                # AddDataField, not `.Orientation = xlDataField` in a loop:
                # the Orientation-loop pattern is non-deterministic in real
                # Excel when the same source column backs two value fields.
                # Omitting Caption lets Excel derive its own default, same
                # as the Orientation path would.
                fn = getattr(c, AGG_TO_WIN32COM_FUNCTION[f["agg"]])
                pt.AddDataField(pf, Function=fn)
            if config["filter_field"]:
                col = config["filter_field"]["column"]
                values = set(config["filter_field"]["values"])
                pf = pt.PivotFields(col)
                pf.Orientation = c.xlPageField
                # An empty `values` means the config wants "select nothing",
                # but Excel refuses to let the last visible PivotItem in a
                # field be hidden (runtime error 1004) -- visi's CLI has the
                # identical gap (see VisiPivotDriver.run's comment above) and
                # handles it the same way: leave the field unfiltered
                # ("(All)") rather than attempting an unrepresentable
                # all-hidden state. Mirrors the fix applied to
                # BuildFuzzPivot.bas's ApplyFilterField for the AppleScript
                # path.
                if values:
                    for item in pf.PivotItems():
                        item.Visible = item.Name in values

            pt.MergeLabels = False
            pt.HasAutoFormat = False
            # Assigning RowGrand/ColumnGrand straight from
            # grand_totals_row/grand_totals_col is swapped because Excel's saved
            # .xlsx rendered grid and rowGrandTotals/colGrandTotals XML attributes
            # match the swapped properties.
            pt.ColumnGrand = config["grand_totals_row"]
            pt.RowGrand = config["grand_totals_col"]
            pt.RefreshTable()
            wb.Save()
            wb.Close()
        finally:
            excel.Quit()


# -----------------------------------------------------------------------------
# 3. CLI & test runner orchestrator
# -----------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Differential fuzzing test harness for visi vs Microsoft Excel pivot tables."
    )
    parser.add_argument("--excel-path", help="Path to Microsoft Excel binary or application bundle.")
    parser.add_argument(
        "--driver", choices=["auto", "applescript", "win32com", "mock"], default="auto",
        help="Excel execution driver.",
    )
    parser.add_argument("--visi-path", default="./target/release/visi", help="Path to compiled visi binary (used by the subprocess backend).")
    add_backend_arg(parser)
    parser.add_argument("--iterations", type=int, default=10, help="Number of fuzz iterations to run.")
    parser.add_argument("--rows", type=int, default=30, help="Max source data rows per iteration.")
    parser.add_argument(
        "--source-mode", choices=["table", "range", "both"], default="both",
        help="Whether the pivot source is an Excel Table, a raw range, or alternate between both.",
    )
    parser.add_argument(
        "--shape", choices=["basic", "rich"], default="basic",
        help="Input shape profile. 'rich' adds more categorical/filter/value fields and blanks.",
    )
    parser.add_argument("--seed", type=int, default=None, help="Random seed for reproducible fuzzing.")
    parser.add_argument("--output-dir", default="./fuzz_results", help="Directory to store test outputs and failure artifacts.")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    failures_dir = os.path.join(args.output_dir, "failures")
    os.makedirs(failures_dir, exist_ok=True)

    visi_driver = VisiPivotDriver(binary_path=args.visi_path, backend=args.backend)
    if args.backend == "auto" and visi_driver.backend != "bindings":
        print(bindings_hint(), file=sys.stderr)
    excel_driver = ExcelPivotDriver(excel_path=args.excel_path, driver_type=args.driver,
                                    visi_path=args.visi_path)
    comparator = DifferentialComparator()
    smoke_mode = excel_driver.driver_type == "mock"

    print("=====================================================================")
    print("      visi vs. Microsoft Excel Pivot Table Differential Fuzzer       ")
    print("=====================================================================")
    print(f" Iterations  : {args.iterations}")
    print(f" Max rows    : {args.rows}")
    print(f" Source mode : {args.source_mode}")
    print(f" Shape       : {args.shape}")
    print(f" Visi        : {visi_driver.describe()}")
    print(f" Excel Driver: {excel_driver.driver_type} ({args.excel_path or 'Default'})")
    if smoke_mode:
        print(f" {SMOKE_BANNER}")
    print("=====================================================================\n")

    passed_count = 0
    failed_count = 0
    start_time = time.time()

    for i in range(1, args.iterations + 1):
        iter_seed = (args.seed + i) if args.seed is not None else random.randint(1, 1000000)
        generator = PivotFuzzGenerator(seed=iter_seed, shape=args.shape)

        temp_dir = tempfile.mkdtemp(prefix=f"fuzz_pivot_iter_{i}_")
        source_xlsx = os.path.join(temp_dir, "source.xlsx")
        visi_out_xlsx = os.path.join(temp_dir, "visi_out.xlsx")
        excel_out_xlsx = os.path.join(temp_dir, "excel_out.xlsx")

        try:
            num_rows = random.randint(1, max(1, args.rows))
            if args.source_mode == "table":
                use_table = True
            elif args.source_mode == "range":
                use_table = False
            else:
                use_table = i % 2 == 0

            config = generator.generate(source_xlsx, num_rows=num_rows, use_table=use_table)

            dest_cell = "L1" if args.shape == "rich" else DEST_CELL
            dest_rc = (0, 11) if args.shape == "rich" else DEST_RC
            visi_driver.run(source_xlsx, config, visi_out_xlsx, PIVOT_NAME, dest_cell, dest_rc)
            if not smoke_mode:
                excel_driver.run(source_xlsx, config, excel_out_xlsx, dest_cell=dest_cell)

            visi_cells = XLSXEvaluatedReader.read_evaluated_cells(visi_out_xlsx)

            if smoke_mode:
                ok, reason = smoke_check(visi_cells)
                if ok:
                    passed_count += 1
                    print(
                        f" Iteration {i:3d}/{args.iterations} [OK] (Seed: {iter_seed},"
                        f" rows: {num_rows}, table: {use_table}, {len(visi_cells)} cells)"
                    )
                else:
                    failed_count += 1
                    print(f"\n Iteration {i:3d}/{args.iterations} [FAILED] (Seed: {iter_seed})")
                    print(f"   {reason}")
                    fail_case_dir = os.path.join(
                        failures_dir, f"pivot_smoke_iter_{i}_seed_{iter_seed}"
                    )
                    shutil.copytree(temp_dir, fail_case_dir, dirs_exist_ok=True)
                    print(f"   Saved reproducing files to: {fail_case_dir}\n")
                continue

            excel_cells = XLSXEvaluatedReader.read_evaluated_cells(excel_out_xlsx)

            is_match, mismatches = comparator.compare(visi_cells, excel_cells)

            if is_match:
                passed_count += 1
                print(f" Iteration {i:3d}/{args.iterations} [PASSED] (Seed: {iter_seed}, rows: {num_rows}, table: {use_table})")
            else:
                failed_count += 1
                print(f"\n Iteration {i:3d}/{args.iterations} [FAILED] (Seed: {iter_seed})")
                print(f"   Found {len(mismatches)} cell mismatch(es):")
                for m in mismatches[:5]:
                    print(f"   - Cell {m['key'][1]} on {m['key'][0]}: visi={m['visi']} | Excel={m['excel']}")

                fail_case_dir = os.path.join(failures_dir, f"pivot_fail_iter_{i}_seed_{iter_seed}")
                shutil.copytree(temp_dir, fail_case_dir, dirs_exist_ok=True)
                print(f"   Saved failure reproducing files to: {fail_case_dir}\n")

        except Exception as err:
            failed_count += 1
            print(f"\n Iteration {i:3d}/{args.iterations} [ERROR]: {err}")
            fail_case_dir = os.path.join(failures_dir, f"pivot_error_iter_{i}_seed_{iter_seed}")
            if os.path.exists(temp_dir):
                shutil.copytree(temp_dir, fail_case_dir, dirs_exist_ok=True)

        finally:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)

    duration = time.time() - start_time
    print("\n=====================================================================")
    print(f" {'Smoke test' if smoke_mode else 'Fuzzing'} Completed in {duration:.2f}s")
    if smoke_mode:
        print(f" Ran    : {passed_count}/{args.iterations} without a crash")
    else:
        print(f" Passed : {passed_count}/{args.iterations}")
    print(f" Failed : {failed_count}/{args.iterations}")
    print("=====================================================================")

    if failed_count > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
