#!/usr/bin/env python3
"""
visi vs. Microsoft Excel Differential Fuzzing: VBA execution
============================================================
Phases 1 and 2 of `docs/vba-macro-support.md`. Generates a VBA procedure, runs
it in `visi`'s interpreter and in real Excel against the same workbook, and
compares four things:

    the returned value    as CStr() renders it
    the returned subtype  as TypeName() reports it
    the error number      as Err.Number reports it, when it raised
    every cell it touched  as TypeName + Value2, over the whole data grid

**The cell comparison is the Phase 2 half, and it is not optional.** A macro
that returns the right number while writing the wrong cell is exactly the
failure this feature can produce and nothing else would catch -- the return
value is compared by every earlier phase, the saved workbook by none of them.
Every case reports the grid, including the ones that never touch it, so a
macro that writes when it should not is a mismatch rather than a silence.

The grid is reported *by the harness itself*, in VBA, rather than by reading
the two saved files. That is what keeps the two engines running literally the
same code: `Harness{i}` resets the grid, calls the case, and serialises the
result, so the string Excel returns and the string `visi` returns are produced
by one implementation and compared directly. It also means a case that raises
still reports the cells it wrote before raising -- which is the measured Excel
behaviour, since nothing rolls back.

The subtype is not decoration. `1 + 1` is an `Integer` and `1 / 1` is a
`Double`; an interpreter that computes the right number with the wrong
subtype has a real bug that only shows up later, when something overflows
that should not have. `Err.Number` is likewise compared exactly -- it is a
small enumerated set, and "which error" is as much a result as "what value".

**Why the generated code is wrapped.** The procedure under test is invoked
through a `Harness` function whose `On Error GoTo` turns any runtime error
into a returned string. Without it an untrapped error pops a modal dialog
that `set display alerts to false` does not suppress, `osascript` never
returns, and Excel has to be SIGKILLed -- so a fuzzer generating random VBA
would stall on most iterations. This was established in issue #46 and the
wrapper is verified to catch errors from arbitrary call depth.

**Why the generator emits a typed AST rather than text.** The lesson already
recorded in this directory's README: a generator that splices strings
produces files Excel silently rejects, and the failure gets attributed to the
engine. Generating from a small typed grammar guarantees by construction that
every variable is `Dim`'d, no name collides with a VBA keyword, every loop
terminates, and no expression can recurse without bound.

**What is deliberately not generated.** Anything non-deterministic (`Now`,
`Rnd`, `Timer`) -- the two engines would differ by construction and every
iteration would be noise. Anything outside the Phase 2 allow-list, since both
sides would agree only that it is unsupported, and a generated `.Interior`
would additionally be a *compile* error in Excel, which hangs the bridge.
Division by an expression that could be zero is *allowed*: error 11 is a
result worth comparing.

**Cell coordinates are always literal**, never generated expressions. A
computed row index can land outside the grid, and `Cells(0, 1)` is error 1004
on both sides -- a result, but a boring one that would crowd out the
interesting cases.

Usage:
    python3 fuzz/fuzz_vba.py --driver mock --iterations 50   # visi only, no Excel
    python3 fuzz/fuzz_vba.py --iterations 20 --seed 1
    python3 fuzz/fuzz_vba.py --iterations 100 --batch 25     # 25 cases per Excel round trip
"""

import argparse
import math
import os
import random
import shutil
import subprocess
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

EXCEL_APP = "Microsoft Excel"

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: source fuzz/venv/bin/activate && pip install -r fuzz/requirements.txt")

try:
    import visi_core
except ImportError:
    sys.exit(
        "the visi_core bindings are required: "
        "maturin develop -m visi-python/Cargo.toml --release"
    )


# -----------------------------------------------------------------------------
# 1. Generation
# -----------------------------------------------------------------------------

# Variable names chosen to avoid VBA's keywords entirely -- `name`, `date`,
# `line`, `error` and friends are all keywords in some position.
VAR_NAMES = ["va", "vb", "vc", "vd", "ve"]

# Locals the host statements need, declared in every generated procedure.
# Kept disjoint from VAR_NAMES and from the loop counters `vi`/`vn`: a name in
# both lists is `Dim`'d twice, which is a **compile** error in Excel ("Duplicate
# declaration in current scope") and therefore a hang rather than a failure --
# and `visi macro check` cannot warn about it either, since Phase 0 resolves no
# names. Found the hard way, on a modal dialog. The assertion below is the
# actual fix; the rename alone would just wait to be undone.
HOST_SHEET_VAR = "wsh"
HOST_CELL_VAR = "vk"
HOST_RANGE_VAR = "vq"
_HOST_VARS = (HOST_SHEET_VAR, HOST_CELL_VAR, HOST_RANGE_VAR)
assert not set(_HOST_VARS) & set(VAR_NAMES + ["vi", "vn"]), (
    f"host locals {_HOST_VARS} collide with generated variables"
)

# Literals spanning every subtype boundary the Variant model has to get right:
# the Integer/Long edge at 32767, the Long/Double edge, and the suffixes.
NUM_LITERALS = [
    "0", "1", "2", "3", "7", "10", "-1", "-7", "255",
    "32767", "32768", "-32768", "100000", "2147483647",
    "1.5", "-2.5", "0.1", "3.75", "1E3", "0.0001",
    "1&", "1%", "2!", "3#", "&HFF", "&O17",
]
STR_LITERALS = ['""', '"a"', '"abc"', '"1"', '"12"', '"1.5"', '"  3  "', '"Z"']
BOOL_LITERALS = ["True", "False"]
SPECIAL_LITERALS = ["Empty", "Null"]

ARITH_OPS = ["+", "-", "*", "/", "\\", "Mod", "^"]
COMPARE_OPS = ["=", "<>", "<", ">", "<=", ">="]
LOGICAL_OPS = ["And", "Or", "Xor", "Eqv", "Imp"]

# One-argument intrinsics that are total enough to be worth generating.
#
# `Len` is absent on purpose. `Len(False)` -- Len applied to a *Boolean
# literal* -- is a compile error in Excel, which surfaces as the modal-dialog
# hang and poisons the whole batch. It is another instance of the Phase 0
# boundary: a check that needs the argument's static type, which the parser
# does not track. `Len` is still covered below, wrapped in `CStr` so its
# argument is always a string.
UNARY_FUNCS = [
    "CStr", "CInt", "CLng", "CDbl", "CSng", "CBool", "Abs", "Sgn", "Int",
    "Fix", "UCase", "LCase", "Trim", "TypeName", "IsNumeric",
    "IsEmpty", "IsNull", "Val", "StrReverse",
]


# The data grid every case runs against, as `(row, col, source)` with 1-based
# coordinates. Deliberately small and deliberately mixed: two numeric columns
# to aggregate, a formula cell so a write can be observed to recalculate, a
# text cell and a boolean so the "ranges skip these, direct arguments coerce
# them" split is under test, and an empty cell.
GRID_ROWS, GRID_COLS = 4, 6
GRID = [
    (1, 1, "1"), (2, 1, "2"), (3, 1, "3"), (4, 1, "4"),
    (1, 2, "10"), (2, 2, "20"), (3, 2, "30"), (4, 2, "40"),
    (1, 3, "=A1*2"), (2, 3, "=A2+B2"),
    (1, 4, "hi"), (2, 4, True),
]
# Columns E and F (5 and 6) are scratch: nothing starts there, so a generated
# write has somewhere to go that no expectation depends on.
SCRATCH_COLS = (5, 6)

# Formulas a case may write into a cell. Fixed rather than generated: whether
# the two engines agree on a *formula* is `fuzz_excel.py`'s question, and
# mixing it in here would attribute its failures to the macro layer.
WRITABLE_FORMULAS = ['"=A1+B1"', '"=SUM(A1:B2)"', '"=A1*3"', '"=COUNT(A1:B4)"']

# Worksheet functions safe to call over the grid: every one is implemented by
# both engines and none is non-deterministic.
GRID_FUNCTIONS = ["Sum", "Count", "CountA", "Min", "Max", "Average"]


class VbaGenerator:
    """Generates VBA expressions and statements from a small typed grammar.

    `depth` bounds expression nesting, which is what keeps generated source
    from growing without limit and keeps the two engines comparing the same
    thing rather than one of them giving up.
    """

    def __init__(self, seed=None):
        self.rng = random.Random(seed)

    # ---- expressions ----------------------------------------------------

    def literal(self):
        bucket = self.rng.random()
        if bucket < 0.55:
            return self.rng.choice(NUM_LITERALS)
        if bucket < 0.78:
            return self.rng.choice(STR_LITERALS)
        if bucket < 0.92:
            return self.rng.choice(BOOL_LITERALS)
        return self.rng.choice(SPECIAL_LITERALS)

    def expr(self, depth, vars_in_scope):
        if depth <= 0:
            if vars_in_scope and self.rng.random() < 0.35:
                return self.rng.choice(vars_in_scope)
            return self.literal()

        kind = self.rng.random()
        if kind < 0.34:
            op = self.rng.choice(ARITH_OPS)
            return f"({self.expr(depth - 1, vars_in_scope)} {op} {self.expr(depth - 1, vars_in_scope)})"
        if kind < 0.46:
            op = self.rng.choice(COMPARE_OPS)
            return f"({self.expr(depth - 1, vars_in_scope)} {op} {self.expr(depth - 1, vars_in_scope)})"
        if kind < 0.56:
            op = self.rng.choice(LOGICAL_OPS)
            return f"({self.expr(depth - 1, vars_in_scope)} {op} {self.expr(depth - 1, vars_in_scope)})"
        if kind < 0.64:
            return f"({self.expr(depth - 1, vars_in_scope)} & {self.expr(depth - 1, vars_in_scope)})"
        if kind < 0.72:
            return f"(Not {self.expr(depth - 1, vars_in_scope)})"
        if kind < 0.80:
            return f"(-{self.expr(depth - 1, vars_in_scope)})"
        if kind < 0.90:
            fn = self.rng.choice(UNARY_FUNCS)
            return f"{fn}({self.expr(depth - 1, vars_in_scope)})"
        if kind < 0.94:
            # See UNARY_FUNCS: Len needs a string argument to stay compilable.
            return f"Len(CStr({self.expr(depth - 1, vars_in_scope)}))"
        # Two-argument string intrinsics, which need a sane count argument.
        fn = self.rng.choice(["Left", "Right", "String", "Space", "InStr", "Mid"])
        if fn in ("Space",):
            return f"Space({self.rng.randint(0, 4)})"
        if fn == "String":
            return f'String({self.rng.randint(0, 4)}, "x")'
        if fn == "InStr":
            return f"InStr({self.expr(depth - 1, vars_in_scope)}, {self.rng.choice(STR_LITERALS)})"
        if fn == "Mid":
            return f"Mid({self.expr(depth - 1, vars_in_scope)}, {self.rng.randint(1, 4)}, {self.rng.randint(0, 4)})"
        return f"{fn}({self.expr(depth - 1, vars_in_scope)}, {self.rng.randint(0, 4)})"

    # ---- statements -----------------------------------------------------

    def statements(self, count, vars_in_scope, depth):
        out = []
        for _ in range(count):
            out.extend(self.statement(vars_in_scope, depth))
        return out

    def statement(self, vars_in_scope, depth):
        kind = self.rng.random()
        target = self.rng.choice(vars_in_scope)

        # The thresholds below are the Phase 1 mix scaled to leave ~30% for
        # the host statements at the end. Host coverage is the whole point of
        # Phase 2, and at the 5% a naive "add one more branch" gives, a
        # 300-case run generates barely a handful of them.
        if kind < 0.28:
            return [f"{target} = {self.expr(depth, vars_in_scope)}"]

        if kind < 0.39:
            return [
                f"If {self.expr(depth - 1, vars_in_scope)} Then",
                f"    {target} = {self.expr(depth - 1, vars_in_scope)}",
                "Else",
                f"    {target} = {self.expr(depth - 1, vars_in_scope)}",
                "End If",
            ]

        if kind < 0.49:
            # Bounds are literal so the loop provably terminates; a generated
            # expression could produce a limit that never ends.
            lo, hi = 1, self.rng.randint(1, 4)
            step = self.rng.choice(["", " Step 2", " Step -1"])
            if step == " Step -1":
                lo, hi = hi, 1
            return [
                f"For vi = {lo} To {hi}{step}",
                f"    {target} = {self.expr(depth - 1, vars_in_scope + ['vi'])}",
                "Next vi",
            ]

        if kind < 0.56:
            return [
                "vn = 0",
                f"Do While vn < {self.rng.randint(1, 3)}",
                "    vn = vn + 1",
                f"    {target} = {self.expr(depth - 1, vars_in_scope + ['vn'])}",
                "Loop",
            ]

        if kind < 0.63:
            subject = self.expr(depth - 1, vars_in_scope)
            return [
                f"Select Case {subject}",
                "Case 0, 1",
                f"    {target} = {self.expr(depth - 1, vars_in_scope)}",
                "Case 2 To 5",
                f"    {target} = {self.expr(depth - 1, vars_in_scope)}",
                "Case Else",
                f"    {target} = {self.expr(depth - 1, vars_in_scope)}",
                "End Select",
            ]

        if kind < 0.70:
            # A locally-trapped error, so `Err.Number` gets compared through
            # the path a real macro uses rather than only through the outer
            # harness.
            return [
                "On Error Resume Next",
                f"{target} = {self.expr(depth, vars_in_scope)}",
                "On Error GoTo 0",
            ]

        return self.host_statement(target, vars_in_scope, depth)

    # ---- the host object model ------------------------------------------

    def cell(self, scratch=False):
        """A literal `(row, col)` inside the grid, or in the scratch columns.

        Literal on purpose -- see the module docstring on why a computed
        coordinate makes for a boring case.
        """
        row = self.rng.randint(1, GRID_ROWS)
        col = self.rng.choice(SCRATCH_COLS) if scratch else self.rng.randint(1, GRID_COLS)
        return row, col

    def a1(self, row, col):
        return f"{chr(ord('A') + col - 1)}{row}"

    def host_statement(self, target, vars_in_scope, depth):
        """One statement that reads or writes the workbook.

        Every construct here is in the Phase 2 allow-list. Anything outside it
        would be an agreed-on error in both engines at best, and a compile
        error that hangs the AppleScript bridge at worst.
        """
        kind = self.rng.random()
        row, col = self.cell()
        srow, scol = self.cell(scratch=True)

        if kind < 0.22:
            # A write, then the read that proves it landed -- and, when the
            # target feeds a formula, that the formula recalculated.
            return [
                f"{HOST_SHEET_VAR}.Cells({srow}, {scol}).Value = {self.expr(depth - 1, vars_in_scope)}",
                f"{target} = {HOST_SHEET_VAR}.Cells({srow}, {scol}).Value",
            ]

        if kind < 0.34:
            # Writing over the grid's own data, which the formula cells in
            # column C read: this is the recalculation path.
            return [
                f"{HOST_SHEET_VAR}.Cells({self.rng.randint(1, 2)}, 1).Value = {self.rng.choice(NUM_LITERALS)}",
                f"{target} = {HOST_SHEET_VAR}.Range(\"C1\").Value",
            ]

        if kind < 0.44:
            return [f"{target} = {HOST_SHEET_VAR}.Cells({row}, {col}).Value"]

        if kind < 0.52:
            prop = self.rng.choice(["Value2", "Formula", "Text", "Address", "Row", "Column"])
            return [f"{target} = {HOST_SHEET_VAR}.Cells({row}, {col}).{prop}"]

        if kind < 0.60:
            r2 = self.rng.randint(row, GRID_ROWS)
            c2 = self.rng.randint(col, GRID_COLS)
            rng = f'{HOST_SHEET_VAR}.Range("{self.a1(row, col)}:{self.a1(r2, c2)}")'
            prop = self.rng.choice(["Count", "Address", "Row", "Column"])
            return [f"{target} = {rng}.{prop}"]

        if kind < 0.70:
            fn = self.rng.choice(GRID_FUNCTIONS)
            r2 = self.rng.randint(row, GRID_ROWS)
            c2 = self.rng.randint(col, GRID_COLS)
            rng = f'{HOST_SHEET_VAR}.Range("{self.a1(row, col)}:{self.a1(r2, c2)}")'
            # Both call paths, since they part company on failure: the
            # `WorksheetFunction` one raises where `Application` returns an
            # error Variant.
            via = self.rng.choice(["Application.WorksheetFunction", "Application"])
            return [f"{target} = {via}.{fn}({rng})"]

        if kind < 0.78:
            return [f"{target} = Application.WorksheetFunction.{self.rng.choice(GRID_FUNCTIONS)}"
                    f"({self.rng.choice(NUM_LITERALS)}, {self.rng.choice(NUM_LITERALS)})"]

        if kind < 0.86:
            r2 = self.rng.randint(row, GRID_ROWS)
            return [
                f'For Each {HOST_CELL_VAR} In {HOST_SHEET_VAR}.Range("{self.a1(row, col)}:{self.a1(r2, col)}")',
                f"    {target} = {HOST_CELL_VAR}.Address & \"/\" & {target}",
                f"Next {HOST_CELL_VAR}",
            ]

        if kind < 0.92:
            return [
                f"{HOST_SHEET_VAR}.Cells({srow}, {scol}).Formula = {self.rng.choice(WRITABLE_FORMULAS)}",
                f"{target} = {HOST_SHEET_VAR}.Cells({srow}, {scol}).Value",
            ]

        if kind < 0.96:
            dr, dc = self.rng.randint(0, 1), self.rng.randint(0, 1)
            return [
                f'With {HOST_SHEET_VAR}.Range("{self.a1(row, col)}")',
                f"    {target} = .Offset({dr}, {dc}).Address & \"/\" & CStr(.Value)",
                "End With",
            ]

        return [
            f"Set {HOST_RANGE_VAR} = {HOST_SHEET_VAR}.Range(\"{self.a1(row, col)}\")",
            f"{target} = CStr({HOST_RANGE_VAR} Is {HOST_SHEET_VAR}.Range(\"{self.a1(row, col)}\")) & \"/\" & TypeName({HOST_RANGE_VAR})",
        ]

    def module(self, index):
        depth = self.rng.randint(1, 3)
        n_vars = self.rng.randint(2, len(VAR_NAMES))
        vars_in_scope = VAR_NAMES[:n_vars]
        body = self.statements(self.rng.randint(1, 4), vars_in_scope, depth)
        result = self.rng.choice(vars_in_scope)

        lines = [f"Private Function Gen{index}()"]
        lines.append("    Dim " + ", ".join(vars_in_scope) + ", vi, vn")
        # Declared unconditionally: a `Dim` costs nothing and the alternative
        # is deciding after the fact which host constructs the body happened
        # to use, which is exactly the string-splicing fragility the typed
        # generator exists to avoid.
        lines.append(
            f"    Dim {HOST_SHEET_VAR} As Worksheet, "
            f"{HOST_CELL_VAR} As Range, {HOST_RANGE_VAR} As Range"
        )
        lines.append(f'    Set {HOST_SHEET_VAR} = ThisWorkbook.Worksheets("Data")')
        for v in vars_in_scope:
            lines.append(f"    {v} = {self.rng.choice(NUM_LITERALS)}")
        lines.extend("    " + s for s in body)
        lines.append(f"    Gen{index} = {result}")
        lines.append("End Function")
        return "\n".join(lines)


# The value-only wrapper.
#
# **Keep this self-contained.** `vba_expr_probe.py` and `vba_host_probe.py`
# both import it and splice it into modules of their own, which define nothing
# else -- so a call to a helper that lives only in *this* file compiles here
# and is "Sub or Function not defined" there. That is a **compile** error, so
# the symptom is a modal dialog and a hung `osascript`, not a failing test.
# Learned by doing it: `GRID_HARNESS_TEMPLATE` below is what that change
# should have been.
HARNESS_TEMPLATE = """Public Function Harness{i}() As String
    On Error GoTo Failed
    Dim r
    r = Gen{i}()
    Harness{i} = "OK|" & TypeName(r) & "|" & CStr(r)
    Exit Function
Failed:
    Harness{i} = "ERR|" & CStr(Err.Number)
End Function"""

# The wrapper this fuzzer uses, which additionally resets the data grid before
# the case and serialises it afterwards. Depends on `GRID_HELPERS`, which
# `build_module` puts in the same module.
#
# The grid is reported even when the case raised: writes made before the error
# persist in Excel, which does not roll back, so discarding them here would be
# comparing against a workbook neither engine has.
GRID_HARNESS_TEMPLATE = """Public Function Harness{i}() As String
    On Error GoTo Failed
    ResetGrid
    Dim r
    r = Gen{i}()
    Harness{i} = "OK|" & TypeName(r) & "|" & CStr(r) & "|" & GridState()
    Exit Function
Failed:
    Harness{i} = "ERR|" & CStr(Err.Number) & "|" & GridState()
End Function"""

# `ResetGrid` runs before every case so the cases in a batch cannot set each
# other up; `GridState` serialises the whole grid afterwards. Both are plain
# VBA and run identically in both engines, which is the point -- the cell
# comparison is then the same string comparison as the value comparison,
# rather than a second mechanism that could itself be wrong.
#
# `CellText` exists because `CStr` is not total: `CStr(Null)` is error 94 and
# `CStr` of a Variant array is error 13, and either would report as the
# *case* failing rather than as one cell being unrenderable.
GRID_HELPERS = """Public Sub ResetGrid()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Data")
    ws.Range("A1:F4").Value = Empty
{writes}
End Sub

Private Function CellText(ByVal c As Range) As String
    On Error GoTo Unrenderable
    CellText = TypeName(c.Value) & ":" & CStr(c.Value2)
    Exit Function
Unrenderable:
    CellText = TypeName(c.Value) & ":?" & CStr(Err.Number)
End Function

Private Function GridState() As String
    Dim ws As Worksheet, s As String, r As Long, c As Long
    Set ws = ThisWorkbook.Worksheets("Data")
    For r = 1 To {rows}
        For c = 1 To {cols}
            s = s & CellText(ws.Cells(r, c)) & ","
        Next c
    Next r
    GridState = s
End Function"""


def grid_helpers():
    writes = []
    for row, col, value in GRID:
        if value is True:
            literal = "True"
        elif isinstance(value, str) and value.startswith("="):
            writes.append(f'    ws.Cells({row}, {col}).Formula = "{value}"')
            continue
        elif isinstance(value, str) and not value.lstrip("-").replace(".", "", 1).isdigit():
            literal = f'"{value}"'
        else:
            literal = str(value)
        writes.append(f"    ws.Cells({row}, {col}).Value = {literal}")
    return GRID_HELPERS.format(
        writes="\n".join(writes), rows=GRID_ROWS, cols=GRID_COLS
    )


def check_no_duplicate_dims(source):
    """Raise if any procedure declares the same name twice.

    A compile error, not a run-time one, so Excel answers it with a modal
    dialog: `osascript` never returns, the whole batch is lost, and the only
    symptom is a hang. `visi macro check` cannot warn about it either -- Phase
    0 resolves no names, by design -- so nothing between the generator and the
    modal dialog would catch it.

    This is the general form of a bug that actually happened: a host local
    named `vc` collided with the generated variable `vc`. The name constants
    are disjoint now, but "the generator emits compilable source **by
    construction**" is the property worth enforcing rather than that one
    instance of breaking it. Cheap, and it fails in Python where a failure is
    readable.
    """
    proc, declared = None, set()
    for line in source.splitlines():
        stripped = line.strip()
        lowered = stripped.lower()
        if lowered.startswith(("private function", "public function", "private sub", "public sub")):
            proc, declared = stripped, set()
        elif lowered.startswith("dim "):
            for part in stripped[4:].split(","):
                name = part.strip().split()[0] if part.strip() else ""
                if not name:
                    continue
                if name.lower() in declared:
                    raise AssertionError(
                        f"generated source declares {name!r} twice in {proc!r}; "
                        "Excel would answer this with a modal compile error"
                    )
                declared.add(name.lower())
    return source


def build_module(cases):
    """One module holding every case in a batch, plus its harnesses.

    Batching matters: the AppleScript round trip dominates the cost by three
    orders of magnitude, so 25 cases in one workbook run in roughly the time
    one case would.
    """
    parts = ['Attribute VB_Name = "M"', grid_helpers()]
    for i, src in cases:
        parts.append(src)
        parts.append(GRID_HARNESS_TEMPLATE.format(i=i))
    return check_no_duplicate_dims("\n\n".join(parts) + "\n")


def build_workbook(path):
    """The workbook both engines run against: one sheet named `Data`.

    `ResetGrid` fills it, so the cells here only have to exist -- but the
    sheet has to be named, since an unqualified `Worksheets("Data")` is how
    every generated case reaches it.
    """
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.save(path)


# -----------------------------------------------------------------------------
# 2. The two engines
# -----------------------------------------------------------------------------


def visi_result(source, proc, workbook=None, harness=False):
    """`OK|TypeName|CStr|grid` or `ERR|number|grid`, the harness's own shape.

    `harness` says the procedure *is* one of the `Harness{i}` wrappers, whose
    return value is already that string -- so it is handed back rather than
    described. Without this the result comes back wrapped one level too deep
    (`OK|String|OK|Double|0.1|...`) and every case reports as a mismatch, for
    a reason that looks nothing like the formatting bug it is.

    `workbook` is a path to run *against*, which is what gives the macro a
    host object model at all. Without one this is the Phase 1 form: source
    text and nothing to touch, which is still the right call for
    `vba_expr_probe.py` and for a case that never reaches a workbook.

    A fresh `Workbook` is loaded per case rather than reused. That is not
    caution -- `ResetGrid` would handle carry-over -- it is what makes a case
    reproducible in isolation from its saved `source.bas`.
    """
    try:
        if workbook is None:
            type_name, value = visi_core.run_macro(source, proc)
        else:
            wb = visi_core.Workbook.load(workbook)
            wb.add_macro("M", source)
            type_name, value, _mutated = wb.run_macro(proc)
    except visi_core.VbaRuntimeError as e:
        return f"ERR|{getattr(e, 'number', '?')}"
    except visi_core.VbaSyntaxError as e:
        return f"SYNTAX|{e}"
    except visi_core.VisiError as e:
        return f"SYNTAX|{type(e).__name__}: {e}"
    if value is None:
        # CStr(Null) is itself error 94 in VBA, which is what the harness
        # would have reported.
        return "ERR|94"
    if harness:
        return value
    return f"OK|{type_name}|{value}"


# -----------------------------------------------------------------------------
# 2b. Comparison
# -----------------------------------------------------------------------------

# Same tolerance `fuzz_excel.py`'s `DifferentialComparator` uses, and for the
# same reason: the two engines do the same arithmetic in different orders and
# the last bit or two of a Double is not a disagreement worth reporting.
FLOAT_REL_TOL = 1e-7
FLOAT_ABS_TOL = 1e-7


def _numeric(text):
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def fields_match(mine, theirs):
    """Whether two harness results agree, field by field.

    Split on the harness's own separators rather than compared as one string,
    so that a Double differing in its last bit does not report as a mismatch
    while a genuinely different *cell* still does. Everything non-numeric is
    compared exactly -- a TypeName, an error number and an address are all
    small closed sets where "close" means nothing.
    """
    if mine == theirs:
        return True
    a, b = mine.split("|"), theirs.split("|")
    if len(a) != len(b):
        return False
    for x, y in zip(a, b):
        if x == y:
            continue
        # The grid field is itself a comma-separated list of `Type:Value`.
        xs, ys = x.split(","), y.split(",")
        if len(xs) != len(ys):
            return False
        for xi, yi in zip(xs, ys):
            if xi == yi:
                continue
            xt, _, xv = xi.rpartition(":")
            yt, _, yv = yi.rpartition(":")
            if xt != yt:
                return False
            xn, yn = _numeric(xv), _numeric(yv)
            if xn is None or yn is None:
                return False
            if not math.isclose(xn, yn, rel_tol=FLOAT_REL_TOL, abs_tol=FLOAT_ABS_TOL):
                return False
    return True


# The child process `ExcelDriver._run_win32com_batch` launches (see that
# method's docstring for why this has to be a separate process rather than
# an in-process win32com call). `argv[1]` is the .xlsm path, `argv[2:]` are
# the `Harness{i}` indices to run. Each result prints as `i=result`,
# matching the AppleScript driver's own `i=result` line format so both
# paths share one parser. `Application.Run` returns a function's value
# directly -- no AppleScript-style string accumulation across calls is
# needed, since each call here is its own subprocess round trip anyway.
_WIN32COM_VBA_RUNNER = """
import sys
import win32com.client

xlsm_path = sys.argv[1]
indices = [int(a) for a in sys.argv[2:]]

excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
# msoAutomationSecurityForceDisable would silently skip running any macro
# at all; msoAutomationSecurityLow (1) runs macros without the "Enable
# Content" prompt that would otherwise hang this exactly like a compile
# error does.
excel.AutomationSecurity = 1
try:
    wb = excel.Workbooks.Open(xlsm_path)
    try:
        for i in indices:
            try:
                result = excel.Run("Harness{}".format(i))
            except Exception as e:
                result = "ERR|COM:{}".format(e)
            print("{}={}".format(i, result))
    finally:
        wb.Close(False)
finally:
    excel.Quit()
"""


class ExcelDriver:
    def __init__(self, excel_path=None, driver_type="auto", timeout=60):
        self.excel_path = excel_path
        self.timeout = timeout
        self.driver_type = driver_type
        if driver_type == "auto":
            if sys.platform == "darwin":
                self.driver_type = "applescript"
            elif sys.platform == "win32":
                self.driver_type = "win32com"
            else:
                self.driver_type = "mock"
        self.restarts = 0

    def app_name(self):
        name = self.excel_path or EXCEL_APP
        if name.endswith(".app"):
            name = os.path.splitext(os.path.basename(name))[0]
        return name

    def restart(self):
        """Force-quit and relaunch, not just `killall` and hope.

        The same escalation `fuzz_pivot.py::_restart_excel` documents, and for
        the same reason: `run VB macro` degrades into a session-wide
        "Parameter error (-50)" after enough consecutive AppleScript calls
        against one long-lived Excel, and `killall` alone was observed to
        leave the process running (the app intercepts SIGTERM for its own quit
        handshake). One batch here is many `run VB macro` calls, so this fires
        both on failure and, via `--restart-every`, before the bridge has a
        chance to wear out.
        """
        self.restarts += 1
        subprocess.run(["killall", EXCEL_APP], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(1.0)
        pgrep = subprocess.run(["pgrep", "-x", EXCEL_APP], stdout=subprocess.PIPE, text=True)
        for pid in pgrep.stdout.split():
            subprocess.run(["kill", "-9", pid], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(1.0)
        subprocess.run(["open", "-a", EXCEL_APP], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(4.0)

    def restart_windows(self):
        """`taskkill` every EXCEL.EXE. Nothing to relaunch -- the next batch's
        `gencache.EnsureDispatch("Excel.Application")` starts a fresh one, the
        same as fuzz_excel.py's/fuzz_chart.py's/fuzz_pivot.py's win32com
        drivers already do on their own retries.
        """
        self.restarts += 1
        subprocess.run(
            ["taskkill", "/F", "/IM", "EXCEL.EXE", "/T"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        time.sleep(1.0)

    def _run_win32com_batch(self, xlsm, indices):
        """Runs one batch through win32com, in a *child process*.

        A generated case with a compile error (an undefined name, a
        duplicate `Dim`) pops a modal dialog that `DisplayAlerts = False`
        does not suppress -- `excel.Run(...)` for that call never returns,
        exactly the failure mode `fuzz_vba.py`'s own module docstring
        describes for the AppleScript path. AppleScript survives this
        because `osascript` is a separate process `subprocess.run(...,
        timeout=...)` can kill out from under a hung Excel; a bare
        in-process win32com call has no equivalent timeout (COM calls
        block the calling thread with no clean way to interrupt one from
        another Python thread in the same apartment). So the actual COM
        work happens in a child `python -u -c` process here too, and a
        timeout kills *that*, then `restart_windows` cleans up the Excel
        process it leaves behind orphaned and hung.
        """
        indices_args = [str(i) for i in indices]
        res = subprocess.run(
            [sys.executable, "-u", "-c", _WIN32COM_VBA_RUNNER, xlsm, *indices_args],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
            timeout=self.timeout,
        )
        out = {}
        for line in res.stdout.splitlines():
            if "=" in line:
                k, _, v = line.partition("=")
                if k.strip().isdigit():
                    out[int(k.strip())] = v.rstrip("\r")
        return res.returncode, out

    def run_batch(self, xlsm, indices):
        """Returns {index: result-string}, or {} if Excel could not be asked."""
        if self.driver_type == "win32com":
            for attempt in range(2):
                try:
                    returncode, out = self._run_win32com_batch(xlsm, indices)
                except subprocess.TimeoutExpired:
                    # A hang here is a compile error in generated source
                    # (Excel goes modal) or a wedged COM automation session --
                    # the same two causes the AppleScript path's timeout
                    # handles. Either way the batch is lost; restart and
                    # retry once.
                    self.restart_windows()
                    if attempt == 0:
                        continue
                    return {}
                if returncode == 0:
                    return out
                self.restart_windows()
                if attempt == 0:
                    continue
                return {}
            return {}
        if self.driver_type == "mock":
            return {}
        # `linefeed`, not "\n": AppleScript string literals do not carry a
        # newline escape, and the whole batch silently returned nothing until
        # this was fixed.
        calls = "\n".join(
            f'    set acc to acc & "{i}=" & (run VB macro "Harness{i}") & linefeed'
            for i in indices
        )
        script = "\n".join([
            f'tell application "{self.app_name()}"',
            "    set display alerts to false",
            "    try",
            "        close workbooks saving no",
            "    end try",
            f'    open POSIX file "{os.path.abspath(xlsm)}"',
            "    set wb to active workbook",
            '    set acc to ""',
            calls,
            "    close wb saving no",
            "    return acc",
            "end tell",
        ])
        for attempt in range(2):
            try:
                res = subprocess.run(
                    ["osascript", "-e", script],
                    stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                    text=True, timeout=self.timeout,
                )
            except subprocess.TimeoutExpired:
                # A hang here is a compile error in generated source (Excel
                # goes modal) or session degradation. Either way the batch is
                # lost; restart and retry once.
                self.restart()
                if attempt == 0:
                    continue
                return {}
            if res.returncode == 0:
                out = {}
                for line in res.stdout.splitlines():
                    if "=" in line:
                        k, _, v = line.partition("=")
                        if k.strip().isdigit():
                            # Only the trailing CR is stripped. `.strip()`
                            # here silently ate the padding on results like
                            # `OK|String|  3  ` and reported a divergence that
                            # was the harness's own doing.
                            out[int(k.strip())] = v.rstrip("\r")
                return out
            self.restart()
            if attempt == 0:
                continue
            return {}
        return {}


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--excel-path")
    ap.add_argument("--driver", choices=["auto", "applescript", "win32com", "mock"], default="auto")
    ap.add_argument("--iterations", type=int, default=20)
    ap.add_argument("--batch", type=int, default=20,
                    help="Cases per Excel round trip (default 20). The round trip dominates cost.")
    ap.add_argument("--seed", type=int, default=None)
    ap.add_argument("--timeout", type=int, default=60)
    ap.add_argument("--restart-every", type=int, default=4,
                    help="Restart Excel every N batches, before its automation "
                         "bridge degrades (0 to only restart on failure).")
    ap.add_argument("--output-dir", default="./fuzz_results")
    args = ap.parse_args()

    gen = VbaGenerator(args.seed)
    driver = ExcelDriver(args.excel_path, args.driver, args.timeout)
    failures_dir = os.path.join(args.output_dir, "failures")
    os.makedirs(failures_dir, exist_ok=True)

    print("=" * 69)
    print("    visi vs. Microsoft Excel VBA Execution Differential Fuzzer    ".center(69))
    print("=" * 69)
    print(f" Cases       : {args.iterations} in batches of {args.batch}")
    print(f" Excel driver: {driver.driver_type} ({args.excel_path or 'default'})")
    if driver.driver_type == "mock":
        print(" MOCK DRIVER -- visi runs, Excel is not consulted, nothing is compared.")
    print("=" * 69 + "\n")

    passed = failed = skipped = 0
    workdir = tempfile.mkdtemp(prefix="vba_exec_fuzz_")
    start = time.time()

    try:
        for batch_no, batch_start in enumerate(range(0, args.iterations, args.batch)):
            if (
                args.restart_every
                and batch_no
                and batch_no % args.restart_every == 0
                and driver.driver_type != "mock"
            ):
                if driver.driver_type == "win32com":
                    driver.restart_windows()
                else:
                    driver.restart()
            n = min(args.batch, args.iterations - batch_start)
            cases = [(batch_start + i + 1, gen.module(batch_start + i + 1)) for i in range(n)]
            source = build_module(cases)
            indices = [i for i, _ in cases]

            base = os.path.join(workdir, "base.xlsx")
            build_workbook(base)

            excel = {}
            if driver.driver_type != "mock":
                xlsm = os.path.join(workdir, f"batch_{batch_start}.xlsm")
                wb = visi_core.Workbook.load(base)
                wb.add_macro("M", source)
                wb.save(xlsm)
                excel = driver.run_batch(xlsm, indices)

            for i, _ in cases:
                # `Harness{i}`, not `Gen{i}`: the harness is what resets the
                # grid and serialises it, so running the inner procedure
                # directly would compare a value against a value-plus-grid.
                mine = visi_result(source, f"Harness{i}", workbook=base, harness=True)
                theirs = excel.get(i)
                if theirs is None:
                    skipped += 1
                    continue
                if fields_match(mine, theirs):
                    passed += 1
                    continue
                failed += 1
                print(f" case {i:<5} [MISMATCH]")
                print(f"   visi : {mine}")
                print(f"   excel: {theirs}")
                out = os.path.join(failures_dir, f"vba_exec_case_{i}")
                os.makedirs(out, exist_ok=True)
                with open(os.path.join(out, "source.bas"), "w") as f:
                    f.write(source)
                with open(os.path.join(out, "verdicts.txt"), "w") as f:
                    f.write(f"procedure: Gen{i}\nvisi:  {mine}\nexcel: {theirs}\n")
                print(f"   saved: {out}")
    finally:
        shutil.rmtree(workdir, ignore_errors=True)

    total = passed + failed
    print("\n" + "=" * 69)
    print(f" Completed in {time.time() - start:.1f}s ({driver.restarts} Excel restarts)")
    print(f" Agreed   : {passed}/{total}" if total else " Agreed   : n/a")
    print(f" Mismatch : {failed}")
    if skipped:
        print(f" Skipped  : {skipped} (Excel gave no answer)")
    print("=" * 69)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
