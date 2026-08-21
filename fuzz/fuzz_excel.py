#!/usr/bin/env python3
"""
visi vs. Microsoft Excel Differential Fuzzing Test Harness
===========================================================
Generates random .xlsx workbooks containing values and formulas, evaluates them
using both `visi` and actual Microsoft Excel, and performs a cell-by-cell semantic
comparison of evaluated results.

Usage:
    python3 fuzz/fuzz_excel.py --excel-path "/Applications/Microsoft Excel.app" --iterations 20
"""

import argparse
import datetime
import io
import math
import os
import random
import re
import shutil
import string
import subprocess
import sys
import tempfile
import time
import zipfile
import xml.etree.ElementTree as ET

# Namespace definitions for OpenXML spreadsheet reading
NS = {
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
}

# -----------------------------------------------------------------------------
# 1. Formula & Data Generator
# -----------------------------------------------------------------------------

class ExcelFuzzGenerator:
    """Generates random data grids and formula trees for Excel compatibility testing."""

    FUNCTIONS_SINGLE_NUM = [
        "ABS", "INT", "SQRT", "ROUND", "ROUNDUP", "ROUNDDOWN", "TRUNC",
        "GAUSS", "PHI", "FISHER", "FISHERINV", "GAMMALN", "GAMMA",
        "GAMMALN.PRECISE", "NORM.S.DIST", "NORM.S.INV", "ACOSH", "ACOT", "ACOTH",
        "ASINH", "ATANH", "COSH", "COT", "COTH", "CSC", "CSCH",
        "DEGREES", "EVEN", "FACT", "FACTDOUBLE", "ODD", "RADIANS",
        "SEC", "SECH", "SIGN", "SINH", "SQRTPI", "TANH",
        "ACOS", "ASIN", "ATAN", "COS", "SIN", "TAN", "EXP", "LN", "LOG10",
        "ERF", "ERF.PRECISE", "ERFC", "ERFC.PRECISE",
    ]
    FUNCTIONS_MULTI_NUM = [
        "SUM", "AVERAGE", "MIN", "MAX", "PRODUCT",
        "AVEDEV", "AVERAGEA", "DEVSQ", "GEOMEAN", "HARMEAN",
        "MEDIAN", "VAR.S", "VAR.P", "VARA", "VARPA",
        "STDEV.S", "STDEV.P", "STDEVA", "STDEVPA", "SKEW", "SKEW.P",
        # MULTINOMIAL is accurate in visi but Excel rounds some integer-valued
        # answers just below the integer (docs/excel-discrepancies.md #25),
        # which leaks through wrappers like INT(MULTINOMIAL(...)).
        "KURT", "MAXA", "MINA", "GCD", "LCM", "SUMSQ",
        "COUNT", "COUNTA", "COUNTBLANK", "SUMPRODUCT",
        "VAR", "VARP", "STDEV", "STDEVP",
    ]
    FUNCTIONS_STAT_BIVARIATE = [
        "CORREL", "PEARSON", "SLOPE", "INTERCEPT", "RSQ", "STEYX",
        "COVARIANCE.P", "COVARIANCE.S", "COVAR", "F.TEST", "FTEST",
        # CHITEST/CHISQ.TEST have several Excel-specific non-numeric-pair and
        # degree-of-freedom rules still being pinned down; keep their reduced
        # cases in Rust tests rather than letting them dominate random fuzzing.
        "SUMX2MY2", "SUMX2PY2", "SUMXMY2",
    ]
    # Two-plain-numeric-argument math functions -- fits the same recursive
    # sub-expression substitution as the "binary" fn_type in gen_expr, just
    # rendered as a function call instead of an infix operator.
    FUNCTIONS_TWO_NUM = ["ATAN2", "LOG", "MOD", "POWER", "QUOTIENT", "PERCENTOF"]
    FUNCTIONS_LOGIC = ["IF", "AND", "OR", "NOT"]
    FUNCTIONS_TEXT = [
        "CONCATENATE", "LEFT", "RIGHT", "LEN", "UPPER", "LOWER",
        # DBCS is locale-dependent in real Excel: on this Windows oracle it is
        # a no-op for ASCII, while visi intentionally maps ASCII to full-width.
        # Keep it out of the differential generator.
        "ASC", "CLEAN", "CODE", "EXACT", "FIND", "FINDB",
        "LEFTB", "LENB", "MIDB", "REPT", "RIGHTB", "SEARCH", "SEARCHB",
        "SUBSTITUTE", "T", "TEXTAFTER", "TEXTBEFORE", "UNICHAR", "UNICODE"
    ]
    # Date/engineering/information functions used to live in
    # FUNCTIONS_DATE/FUNCTIONS_ENGINEERING/FUNCTIONS_INFO but were never
    # actually wired into gen_expr's fn_type dispatch (dead lists -- none of
    # these were ever generated). They've been folded into the bespoke
    # per-function generators below (generate_date_formula/
    # generate_engineering_formula/generate_logic_formula) instead, since
    # their wildly varying arities (DATE takes 3 args, YEAR takes 1, BASE
    # takes a string + radix, ...) don't fit gen_expr's uniform
    # arbitrary-sub-expression substitution model.
    DATE_FUNCTIONS = [
        "DATE", "DAY", "DAYS", "DAYS360", "EDATE", "EOMONTH", "HOUR", "MINUTE",
        "MONTH", "SECOND", "TIME", "WEEKDAY", "WEEKNUM", "YEAR", "YEARFRAC",
        "DATEDIF", "DATEVALUE", "TIMEVALUE", "ISOWEEKNUM",
        "NETWORKDAYS", "NETWORKDAYS.INTL", "WORKDAY", "WORKDAY.INTL",
    ]
    ENGINEERING_FUNCTIONS = [
        "BIN2DEC", "DEC2BIN", "DEC2HEX", "DEC2OCT", "DELTA", "GESTEP",
        "HEX2DEC", "OCT2DEC", "BITAND", "BITOR", "BITXOR",
        "BIN2HEX", "BIN2OCT", "HEX2BIN", "HEX2OCT", "OCT2BIN", "OCT2HEX",
        "BASE", "DECIMAL", "BITLSHIFT", "BITRSHIFT", "CONVERT",
        "COMPLEX",
        "IMABS", "IMAGINARY", "IMARGUMENT", "IMCONJUGATE", "IMCOS", "IMCOSH",
        "IMCOT", "IMCSC", "IMCSCH", "IMDIV", "IMEXP", "IMLN", "IMLOG10",
        "IMLOG2", "IMPOWER", "IMPRODUCT", "IMREAL", "IMSEC", "IMSECH",
        "IMSIN", "IMSINH", "IMSQRT", "IMSUB", "IMSUM", "IMTAN",
        # IS.CEILING is not an Excel function name in the Windows oracle
        # (it returns #NAME?); the implemented/comparable spelling is
        # ISO.CEILING, and it stays fuzzed here.
        "ISO.CEILING", "CEILING", "CEILING.MATH", "CEILING.PRECISE",
        "FLOOR", "FLOOR.MATH", "FLOOR.PRECISE",
        "COMBIN", "COMBINA", "PERMUT", "PERMUTATIONA", "MROUND",
    ]
    # LET is deliberately excluded, like the LAMBDA family above: real
    # Excel can't open an openpyxl-authored file containing an
    # `_xlfn.`-prefixed LET at all (confirmed by isolating both
    # `_xlfn.LET(a, 5, a+1)` -- no nesting, still corrupts -- and
    # `_xlfn.LET(a, _xlfn.SKEW.P(F1,F2), a+1)` in their own single-cell
    # workbooks: `open` returns without error but no workbook window
    # appears), while leaving it bare gets it silently rewritten to
    # `_xludf.LET` (Excel's "unrecognized name" marker) and `#NAME?` --
    # an unfixable-either-way authoring limitation, not a visi bug (see
    # LAMBDA_FUNCTIONS above for the same failure signature). LET is
    # generated deterministically once per LOGIC_EXTRA_FUNCTIONS entry
    # (not randomly), so leaving it in this list would corrupt or
    # mismatch on every single fuzz iteration. Its correctness is instead
    # verified by visi-core's own Rust unit tests (see
    # test_let_binds_names_in_sequence_and_rejects_duplicate_names in
    # engine/tests/new_functions.rs).
    LOGIC_EXTRA_FUNCTIONS = [
        "ISEVEN", "ISODD", "ISLOGICAL", "ISNONTEXT", "TYPE", "XOR",
        "IFERROR", "IFNA", "IFS", "SWITCH",
        # ISERR/ISNA turn the documented error-class precedence divergence
        # (docs/excel-discrepancies.md #13) into a boolean mismatch instead of
        # a tolerated both-error case, so keep them out of the random wrapper
        # pool. ISERROR still exercises error detection without caring which
        # class won.
        "ISBLANK", "ISERROR", "ISNUMBER", "ISTEXT",
        "CHOOSE",
    ]
    # Statistical distribution/percentile-rank functions: unlike
    # FUNCTIONS_MULTI_NUM/STAT_BIVARIATE these need domain-restricted scalar
    # parameters (probabilities in (0,1), positive shape/scale params,
    # integer trial counts, ...) that gen_expr's arbitrary sub-expression
    # substitution can't guarantee, so they get their own generator --
    # same rationale as generate_financial_formula.
    DISTRIBUTION_FUNCTIONS = [
        "BETA.DIST", "BETADIST", "BETA.INV", "BETAINV",
        "BINOM.DIST", "BINOMDIST", "BINOM.DIST.RANGE", "BINOM.INV", "CRITBINOM",
        "CHISQ.DIST", "CHISQ.DIST.RT", "CHIDIST", "CHISQ.INV", "CHISQ.INV.RT", "CHIINV",
        "CONFIDENCE.NORM", "CONFIDENCE", "CONFIDENCE.T",
        "EXPON.DIST", "EXPONDIST",
        "F.DIST", "F.DIST.RT", "FDIST", "F.INV", "F.INV.RT", "FINV",
        "GAMMA.DIST", "GAMMADIST", "GAMMA.INV", "GAMMAINV",
        "HYPGEOM.DIST", "HYPGEOMDIST",
        "LOGNORM.DIST", "LOGNORMDIST", "LOGNORM.INV", "LOGINV",
        "NEGBINOM.DIST", "NEGBINOMDIST",
        "NORM.DIST", "NORMDIST", "NORM.INV", "NORMINV",
        "NORM.S.DIST", "NORMSDIST", "NORM.S.INV", "NORMSINV",
        "POISSON.DIST", "POISSON", "PROB", "STANDARDIZE",
        "T.DIST", "T.DIST.2T", "T.DIST.RT", "TDIST", "T.INV", "T.INV.2T", "TINV",
        "T.TEST", "TTEST", "WEIBULL.DIST", "WEIBULL", "Z.TEST", "ZTEST",
        "LARGE", "SMALL",
        "PERCENTILE", "PERCENTILE.INC", "PERCENTILE.EXC",
        # QUARTILE.EXC excluded: visi's exclusive-quartile interpolation
        # rejects some quart/sample-size combinations Excel accepts (see
        # "docs/excel-discrepancies.md" section 10). QUARTILE.INC and the PERCENTILE.* family
        # agree and stay fuzzed.
        "QUARTILE", "QUARTILE.INC",
        "PERCENTRANK", "PERCENTRANK.INC", "PERCENTRANK.EXC",
        # FREQUENCY excluded: its bins_array coercion for non-numeric bins
        # is not understood (see "docs/excel-discrepancies.md" section 12).
        "RANK", "RANK.EQ", "RANK.AVG", "TRIMMEAN", "MODE.MULT",
    ]
    LOOKUP_FUNCTIONS = ["INDEX", "MATCH", "VLOOKUP", "HLOOKUP", "XLOOKUP"]
    # JIS is deliberately excluded for the same practical reason as DBCS in
    # FUNCTIONS_TEXT: the installed Windows Excel oracle does not recognize
    # either bare or `_xlfn.`-prefixed JIS (it returns #NAME?), while visi
    # implements the documented text conversion locally.
    #
    # FILTERXML is implemented and works through Excel's interactive Evaluate
    # path, but this harness writes formulas through openpyxl-authored OOXML;
    # that route stores FILTERXML as `_xlfn.FILTERXML(...)`, which the same
    # Windows Excel oracle opens as #NAME?. Keep FILTERXML pinned in Rust unit
    # tests rather than making every fuzz iteration fail on an authoring quirk.
    #
    # ENCODEURL is deliberately excluded: even correctly written as
    # `_xlfn.ENCODEURL(...)`, the installed real-Excel build (16.111.3)
    # returns `#NAME?` for it -- confirmed by isolating both the bare and
    # `_xlfn.`-prefixed forms in their own single-cell workbook and
    # comparing against `_xlfn.CONCAT`/`_xlfn.TEXTJOIN` (same era of
    # function, both resolve fine) in the same file. Not a visi bug or a
    # prefixing bug, just a function this particular Excel build doesn't
    # implement.
    TEXT_EXTRA_FUNCTIONS = [
        "PROPER", "TRIM", "CHAR", "TEXTJOIN", "TEXTSPLIT", "VALUE", "VALUETOTEXT",
        "N", "NA", "DOLLAR", "FIXED", "NUMBERVALUE", "ARABIC", "ROMAN", "BAHTTEXT",
        "REGEXEXTRACT", "REGEXREPLACE", "REGEXTEST", "REPLACE", "REPLACEB",
        "CONCAT", "ERROR.TYPE", "MID", "TEXT", "ADDRESS", "ARRAYTOTEXT",
    ]
    # Array/matrix-returning functions. Excel would spill these across
    # multiple cells; visi's xlsx export caches one value per cell (see
    # AGENTS.md), so each is wrapped in INDEX(...) to pin down the single
    # scalar a plain cell comparison can check, matching how a spreadsheet
    # author would consume them from one cell in practice.
    ARRAY_FUNCTIONS = [
        "MDETERM", "MINVERSE", "MMULT", "MUNIT", "SEQUENCE",
        "LINEST", "LOGEST", "GROWTH", "TREND",
        "FORECAST", "FORECAST.LINEAR",
        "SERIESSUM",
    ]
    # The FORECAST.ETS family, fuzzed against a *well-posed* series (see
    # the ETS block in create_fuzz_workbook): a regular timeline carrying an
    # exact linear trend plus an exactly repeating season.
    #
    # That restriction is deliberate and is what makes the comparison
    # meaningful. Excel fits alpha/beta/gamma with a proprietary optimizer
    # and reports them to three decimals; no independent implementation
    # reproduces those digits on noisy data, so a random series would only
    # ever be asserting "did you reimplement Microsoft's optimizer". On a
    # series the model fits perfectly the forecast is the same for any sane
    # parameter triple -- Excel returns the exact continuation -- so what
    # actually gets tested is the part that *is* well-defined: timeline
    # validation and gap filling, season-length detection, the Holt-Winters
    # recurrences, and extrapolation.
    #
    # STAT types 1-3 (alpha/beta/gamma) are excluded for the same reason:
    # they report the optimizer's chosen parameters, which stay
    # implementation-specific even when the forecast doesn't. Types 4-8
    # (MASE/SMAPE/MAE/RMSE/step) are well-defined and are included.
    # FORECAST.ETS.SEASONALITY is deliberately absent. Excel's automatic
    # season-length detection is a heuristic that does not simply report the
    # series' true period, and its answer turns on the *arrangement* of the
    # seasonal offsets rather than on their magnitude or the trend. Probed
    # directly, over 16 points with slope 2 and the same four offsets merely
    # permuted: [8, -2, -8, 2] -> 4, [11, -11, 2, -2] -> 2,
    # [2, -2, 11, -11] -> 2, and [-2, -11, 11, 2] -> 0, i.e. Excel reports
    # *no* seasonality for a series that is exactly period-4. (Trend
    # strength is not the trigger: holding the offsets fixed and sweeping
    # the slope from 0 to 4 leaves the answer at 4 throughout.) Comparing
    # against that is asserting a specific heuristic, not correctness, so
    # detection is covered by visi-core's own unit tests -- on the patterns
    # where Excel's answer *is* the true period -- and every function below
    # is handed an explicit season length instead.
    ETS_FUNCTIONS = [
        "FORECAST.ETS", "FORECAST.ETS.CONFINT",
        "FORECAST.ETS.STAT4", "FORECAST.ETS.STAT5", "FORECAST.ETS.STAT6",
        "FORECAST.ETS.STAT7", "FORECAST.ETS.STAT8",
    ]
    CONDITIONAL_AGG_FUNCTIONS = [
        "AVERAGEIF", "AVERAGEIFS", "COUNTIF", "COUNTIFS",
        "MAXIFS", "MINIFS", "SUMIF", "SUMIFS", "SUBTOTAL", "AGGREGATE",
    ]
    # Volatile/non-deterministic functions: each engine evaluates these
    # independently (different real-world instant, different RNG state), so
    # exact-value comparison would spuriously fail. RANDBETWEEN/RANDARRAY are
    # called with min==max to force a deterministic value out of an
    # inherently random function; RAND/NOW/TODAY have no such knob, so they
    # get wrapped in a range/plausibility check both engines must satisfy
    # regardless of the actual random or wall-clock value.
    VOLATILE_FUNCTIONS = ["RAND", "RANDBETWEEN", "RANDARRAY", "NOW", "TODAY"]
    DATABASE_FUNCTIONS = [
        "DSUM", "DAVERAGE", "DCOUNT", "DCOUNTA", "DGET", "DMAX", "DMIN",
        "DPRODUCT", "DSTDEV", "DSTDEVP", "DVAR", "DVARP",
    ]
    # LAMBDA family: deliberately left empty -- none of MAP/REDUCE/BYROW/
    # BYCOL/MAKEARRAY/SCAN are fuzzed against real Excel. Two independent,
    # confirmed failure modes rule the whole family out, not just the
    # dynamic-array-spilling ones:
    #   - BYROW/BYCOL/MAKEARRAY/SCAN (and a bare, uninvoked LAMBDA) return
    #     or are a dynamic-array-spilling value, and this environment's
    #     Excel AppleScript automation bridge breaks intermittently on
    #     *any* dynamic-array-spilling formula (confirmed directly with a
    #     plain `=SEQUENCE(3)`, no LAMBDA involved at all).
    #   - MAP and REDUCE looked fuzzable (their result is pinned to a
    #     scalar via INDEX(...)/its own scalar return), but real Excel
    #     silently fails to open an openpyxl-authored file containing an
    #     `_xlfn.LAMBDA(...)` formula at all -- confirmed by isolating
    #     `INDEX(MAP(F1:F5, _xlfn.LAMBDA(x, x*2+1)), 1)` (LAMBDA prefixed,
    #     MAP left bare) in its own single-cell workbook: Excel's own
    #     AppleScript `open` returns without error, but no window and no
    #     workbook actually appears (`count of workbooks` stays 0), the
    #     same silent-corruption signature `_xlfn.`-prefixing a LET
    #     variable named "r"/"c" produces (see the LET generator below).
    #     Leaving LAMBDA bare avoids the corruption but then Excel doesn't
    #     recognize it (`_xludf.LAMBDA`, `#NAME?`) -- an unfixable-either-
    #     way formula authoring limitation, not a visi bug.
    # Their expected values are instead verified by hand-calculated
    # arithmetic / Microsoft's documented SCAN example in visi-core's own
    # Rust unit tests (see engine/tests/new_functions.rs).
    LAMBDA_FUNCTIONS = []
    # CELL/INFO are fuzzed only over the narrow info_type subsets visi
    # implements. GETPIVOTDATA is deliberately absent: fuzz_excel.py creates
    # ordinary formula workbooks, not native pivot tables, so there is no
    # stable pivot destination for Excel to resolve. Pivot definitions and
    # grids are covered by fuzz_pivot.py instead.
    RANGE_INFO_FUNCTIONS = [
        "ROW", "ROWS", "COLUMN", "COLUMNS", "AREAS", "ISREF",
        "FORMULATEXT", "ISFORMULA", "HYPERLINK", "SHEETS", "SHEET", "INDIRECT", "OFFSET",
        "CELL", "INFO",
    ]
    # Dynamic-array reshaping/lookup functions. TRANSPOSE is deliberately
    # excluded: every authoring variant tried (bare, `_xlfn.`,
    # `_xlfn._xlws.`, standalone, nested in SUM/INDEX) gave real Excel
    # `#VALUE!` when the formula was written by openpyxl rather than Excel
    # itself -- TRANSPOSE has always required the legacy CSE `t="array"`
    # formula flag, which openpyxl's plain string assignment never
    # produces, so it would be a guaranteed mismatch every run regardless
    # of visi's own (hand-verified correct) logic. LOOKUP is also excluded:
    # its vector form requires an ascending-sorted lookup array or its
    # result is documented by Microsoft as unpredictable, and this
    # generator's plain-value columns are unsorted random data -- verified
    # correct against real Excel on sorted data instead, in visi-core's own
    # Rust unit tests (see engine/tests/new_functions.rs). XMATCH is fuzzed
    # in its default exact-match mode only (well-defined regardless of
    # sort order), looking up a value known to be present so both engines
    # are guaranteed a match.
    ARRAY_RESHAPE_FUNCTIONS = [
        "HSTACK", "VSTACK", "CHOOSEROWS", "CHOOSECOLS", "DROP", "TAKE",
        "EXPAND", "TOCOL", "TOROW", "WRAPROWS", "WRAPCOLS", "UNIQUE",
        "SORT", "SORTBY", "FILTER", "TRIMRANGE", "XMATCH",
    ]

    # Scalar-argument TVM/depreciation functions. Unlike the generic
    # FUNCTIONS_* lists above, financial functions can't have arbitrary
    # sub-expressions substituted into their arguments (a rate must stay
    # small and positive, a period must stay within [1, nper], etc.), so
    # they get their own generator methods below instead of feeding into
    # gen_expr's recursive substitution.
    FINANCIAL_FUNCTIONS = [
        # RATE excluded: Excel gives up with #NUM! on series where a root
        # demonstrably exists (the same call returns a rate when handed a
        # guess near it), so the comparison asserts whether Excel's
        # iteration converged from its default 0.1 rather than anything
        # about correctness. See "docs/excel-discrepancies.md" section 11.
        "PV", "FV", "PMT", "NPER", "IPMT", "PPMT", "CUMIPMT",
        "CUMPRINC", "NPV", "IRR", "MIRR", "XNPV", "XIRR", "SLN", "SYD",
        # EUROCONVERT is implemented, but Excel exposes it through the Euro
        # Currency Tools add-in; the installed oracle returns #NAME? without
        # that add-in, so visi pins it against Microsoft's published rates in
        # Rust tests instead of fuzz_excel.py.
        "DB", "DDB", "VDB", "EFFECT", "NOMINAL", "DOLLARDE", "DOLLARFR",
        "FVSCHEDULE", "RRI", "PDURATION", "ISPMT",
        # Day-count / bond-pricing functions (see finance.rs).
        "COUPDAYBS", "COUPDAYS", "COUPDAYSNC", "COUPNCD", "COUPNUM", "COUPPCD",
        "PRICE", "YIELD", "DURATION", "MDURATION",
        "DISC", "PRICEDISC", "YIELDDISC", "PRICEMAT", "YIELDMAT",
        "RECEIVED", "INTRATE", "TBILLPRICE", "TBILLYIELD", "TBILLEQ",
        "ACCRINT", "ACCRINTM", "AMORLINC",
        "ODDLPRICE", "ODDLYIELD",
    ]
    # AMORDEGRC, ODDFPRICE and ODDFYIELD are excluded as known visi gaps --
    # see "docs/excel-discrepancies.md" sections 7 and 8. AMORDEGRC's coefficient brackets and
    # end-of-life switch to straight line aren't fully reverse-engineered,
    # and Excel rejects odd-first-coupon orderings (returning #NUM!) that
    # visi accepts. The regular-coupon bond functions above are unaffected
    # and stay fuzzed.

    # Every function name below needs a bare `_xlfn.` prefix (not the
    # `_xlfn._xlws.` double-namespace some dynamic-array functions use --
    # HSTACK/VSTACK/CHOOSEROWS/CHOOSECOLS/DROP/TAKE/EXPAND/TOCOL/TOROW/
    # WRAPROWS/WRAPCOLS/UNIQUE/SORT/SORTBY/FILTER/TRIMRANGE/XMATCH, plus
    # RRI/PDURATION/FORMULATEXT/ISFORMULA/SHEETS/SHEET, are already handled
    # by their own bespoke generators and deliberately excluded here to
    # avoid double-prefixing) when written into an openpyxl-authored file,
    # or real Excel shows `#NAME?` for it -- these are exactly the
    # functions real Excel itself rewrote as `_xludf.NAME` (its own "I
    # don't recognize this name" marker) when it re-saved a workbook
    # containing them unprefixed, confirmed against the actual installed
    # Excel build (16.111.3) by round-tripping every FUNCTIONS_*/
    # DISTRIBUTION_FUNCTIONS/ENGINEERING_FUNCTIONS/DATE_FUNCTIONS/
    # TEXT_EXTRA_FUNCTIONS/LOGIC_EXTRA_FUNCTIONS/ARRAY_FUNCTIONS/
    # CONDITIONAL_AGG_FUNCTIONS/RANGE_INFO_FUNCTIONS entry through this
    # generator's own formula construction and reading back which cells
    # Excel could and couldn't recognize. Applied as a single post-
    # processing pass over every formula cell in create_fuzz_workbook
    # (see _apply_xlfn_prefixes) rather than threading prefix logic
    # through every individual generator method, since the vast majority
    # of call sites build formula text by directly interpolating the
    # function name (`f"{fn}(...)"`) with no single seam to hook.
    NEEDS_XLFN_PREFIX = frozenset([
        "ACOT", "ACOTH", "AGGREGATE", "ARABIC", "ARRAYTOTEXT", "BASE",
        "BETA.DIST", "BETA.INV", "BINOM.DIST", "BINOM.DIST.RANGE", "BINOM.INV",
        "BITAND", "BITLSHIFT", "BITOR", "BITRSHIFT", "BITXOR",
        "CEILING.MATH", "CEILING.PRECISE",
        "CHISQ.DIST", "CHISQ.DIST.RT", "CHISQ.INV", "CHISQ.INV.RT", "CHISQ.TEST",
        "COMBINA", "CONCAT", "CONFIDENCE.NORM", "CONFIDENCE.T",
        "COT", "COTH", "COVARIANCE.P", "COVARIANCE.S", "CSC", "CSCH",
        "DAYS", "DECIMAL", "ERF.PRECISE", "ERFC.PRECISE", "EXPON.DIST",
        "F.DIST", "F.DIST.RT", "F.INV", "F.INV.RT", "F.TEST",
        "FLOOR.MATH", "FLOOR.PRECISE",
        "FORECAST.ETS", "FORECAST.ETS.CONFINT", "FORECAST.ETS.SEASONALITY",
        "FORECAST.ETS.STAT", "FORECAST.LINEAR",
        "GAMMA", "GAMMA.DIST", "GAMMA.INV", "GAMMALN.PRECISE", "GAUSS",
        "HYPGEOM.DIST", "IFNA", "IFS",
        "IMCOSH", "IMCOT", "IMCSC", "IMCSCH", "IMSEC", "IMSECH", "IMSINH", "IMTAN",
        "ISOWEEKNUM",
        "LOGNORM.DIST", "LOGNORM.INV", "MAXIFS", "MINIFS",
        "MODE.MULT", "MODE.SNGL", "MUNIT",
        "NEGBINOM.DIST", "NORM.DIST", "NORM.INV", "NORM.S.DIST", "NORM.S.INV",
        "NUMBERVALUE",
        "PERCENTILE.EXC", "PERCENTILE.INC", "PERCENTOF",
        "PERCENTRANK.EXC", "PERCENTRANK.INC", "PERMUTATIONA", "PHI",
        "POISSON.DIST", "QUARTILE.EXC", "QUARTILE.INC",
        "RANDARRAY", "RANK.AVG", "RANK.EQ",
        "REGEXEXTRACT", "REGEXREPLACE", "REGEXTEST",
        "SEC", "SECH", "SEQUENCE", "SKEW.P", "STDEV.P", "STDEV.S", "SWITCH",
        "T.DIST", "T.DIST.2T", "T.DIST.RT", "T.INV", "T.INV.2T", "T.TEST",
        "TEXTAFTER", "TEXTBEFORE", "TEXTJOIN", "TEXTSPLIT",
        "UNICHAR", "UNICODE", "VALUETOTEXT", "VAR.P", "VAR.S",
        "WEIBULL.DIST", "XLOOKUP", "XOR", "Z.TEST",
    ])

    @classmethod
    def _apply_xlfn_prefixes(cls, formula):
        """Rewrites every recognized-but-unprefixed post-2007 function call
        in a formula string to carry its required `_xlfn.` prefix (see
        NEEDS_XLFN_PREFIX). The negative lookbehind skips any occurrence
        already preceded by a `.` -- i.e. already namespaced, whether as
        plain `_xlfn.NAME(` or double-namespaced `_xlfn._xlws.NAME(` --
        so this is safe to run over formulas a bespoke generator already
        prefixed by hand."""
        if formula is None or not formula.startswith("="):
            return formula
        for name in cls.NEEDS_XLFN_PREFIX:
            pattern = r'(?<![A-Za-z0-9_.])' + re.escape(name) + r'\('
            formula = re.sub(pattern, f"_xlfn.{name}(", formula)
        return formula

    def __init__(self, seed=None):
        if seed is not None:
            random.seed(seed)
        # Populated by create_fuzz_workbook() once it defines an Excel Table
        # over the value grid; used by generate_formula() to emit structured
        # references. Left unset (no structured refs generated) if a caller
        # invokes generate_formula() without going through
        # create_fuzz_workbook() first.
        self._table_name = None
        self._table_cols = []
        # Populated by create_fuzz_workbook()'s financial data block; used
        # by generate_financial_formula() for the array-argument functions
        # (NPV/IRR/MIRR/XNPV/XIRR/FVSCHEDULE). visi's parser has no `{...}`
        # array-literal syntax, so those functions must reference real
        # ranges rather than inline arrays.
        self._fin_cash_range = None
        self._fin_date_range = None
        self._fin_schedule_range = None
        # Populated by create_fuzz_workbook(); the table block's own
        # A1:...N range, reused as the "database" argument for the D*
        # database functions (generate_database_formula) since it already
        # has a header row of column-letter names plus random data rows.
        self._db_range = None
        # Populated by create_fuzz_workbook()'s ETS block; used by
        # generate_ets_formula().
        self._ets_timeline_range = None
        self._ets_values_range = None
        self._ets_next_target = 0
        self._ets_period = 0

    def _col_name(self, col_idx):
        """Converts 1-based column index to A1 column letter (1 -> A, 2 -> B, 27 -> AA)."""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _has_table(self):
        return bool(self._table_name and self._table_cols)

    def _random_structured_col_ref(self):
        """A structured reference to one table column's whole data body,
        e.g. `Sheet1[A]`. Evaluates to an array, exactly like a range
        reference (A1:A10), so it's only ever used where a range reference
        would also be valid (as a whole argument to an aggregate function).

        visi treats a structured column reference as spanning the *entire*
        column of the sheet, not just a table's declared row range -- so if
        table columns could also hold formulas, this could create a
        dependency cycle (directly, or transitively through another
        formula). create_fuzz_workbook keeps the table's columns disjoint
        from the columns formulas are written into and never puts a
        formula in a table column, which rules that out entirely rather
        than just making it unlikely.
        """
        col_name = random.choice(self._table_cols)
        return f"{self._table_name}[{col_name}]"

    def _random_structured_header_ref(self):
        """A structured reference to a single column's header text, e.g.
        `Sheet1[[#Headers],[A]]`. Evaluates to a plain scalar string, so
        (unlike a bare column reference) it's safe to use anywhere a cell
        reference or constant would be used."""
        col_name = random.choice(self._table_cols)
        return f"{self._table_name}[[#Headers],[{col_name}]]"

    def _maybe_abs_col(self, col_text):
        """Randomly apply an absolute-column marker to an A1 reference.

        Absolute markers do not change formula evaluation, but they do
        exercise the parser/import/export path for the reference forms real
        workbooks contain (`$A$1`, `A$1`, `$A1`). Structural-edit fuzzing is
        where their movement semantics matter; here the goal is cheap input
        entropy without changing dependency acyclicity.
        """
        return f"${col_text}" if random.random() < 0.25 else col_text

    def _maybe_abs_row(self, row_idx):
        return f"${row_idx}" if random.random() < 0.25 else str(row_idx)

    def _format_cell_ref(self, row_idx, col_idx):
        return f"{self._maybe_abs_col(self._col_name(col_idx))}{self._maybe_abs_row(row_idx)}"

    def _format_col_ref(self, col_idx):
        return self._maybe_abs_col(self._col_name(col_idx))

    def _random_cell_ref(self, current_row, min_col, max_col):
        """A reference to a single cell in an earlier row (or row 1 if
        current_row <= 1), used by generate_formula and the bespoke
        generators below to avoid creating dependency cycles."""
        r = random.randint(1, max(1, current_row - 1)) if current_row > 1 else 1
        c = random.randint(min_col, max_col)
        return self._format_cell_ref(r, c)

    def _random_range_ref(self, current_row, min_col, max_col):
        """A reference to a rectangular range confined to earlier rows (or
        row 1 if current_row <= 1), for the same reason as
        _random_cell_ref."""
        r1 = random.randint(1, max(1, current_row - 1)) if current_row > 1 else 1
        r2 = random.randint(r1, max(1, current_row - 1)) if current_row > 1 else 1
        c1 = random.randint(min_col, max_col)
        c2 = random.randint(c1, max_col)
        return f"{self._format_cell_ref(r1, c1)}:{self._format_cell_ref(r2, c2)}"

    def _random_table_whole_col_ref(self):
        """A whole-column reference into the formula-free table block.

        Whole-column refs over the formula block could include the formula's
        own cell and create cycles. The table block is pure input values, so
        `A:A`/`$A:$A`-style references add address entropy while staying
        acyclic and comparable.
        """
        if not self._has_table():
            return None
        col_idx = random.randint(1, len(self._table_cols))
        col = self._format_col_ref(col_idx)
        return f"{col}:{col}"

    def generate_random_value(self):
        """Generates a random cell input value (number, string, boolean, edge case)."""
        choice = random.random()
        if choice < 0.35:
            # Integers
            return random.randint(-100, 100)
        elif choice < 0.60:
            # Floating point numbers (including zero)
            if random.random() < 0.1:
                return 0.0
            return round(random.uniform(-500.0, 500.0), random.randint(0, 4))
        elif choice < 0.78:
            # Short strings. Keep the random cell-value alphabet ASCII: Windows
            # Excel's Unicode collation is locale-sensitive and not a stable
            # oracle for SORT/SORTBY text ordering (docs/excel-discrepancies.md
            # #23). Punctuation still exercises shared-string/text paths beyond
            # plain `[A-Za-z 123]`.
            if random.random() < 0.25:
                samples = ["a,b", "quote ' test", "paren(test)", "dash-test"]
                return random.choice(samples)
            # Avoid whitespace-only strings: OOXML/openpyxl can strip them to
            # empty shared strings unless xml:space="preserve", which makes
            # Excel and visi disagree about blank-vs-empty-string observability
            # in functions such as ARRAYTOTEXT (docs/excel-discrepancies.md #14).
            chars = string.ascii_letters + "123"
            return "".join(random.choice(chars) for _ in range(random.randint(1, 8)))
        elif choice < 0.88:
            # Booleans
            return random.choice([True, False])
        elif choice < 0.98:
            # Empty / None
            return None
        else:
            # Small integers for range indexes
            return random.randint(1, 10)

    def _generate_text_expr(self, fn, gen_expr, depth):
        """Generate one FUNCTIONS_TEXT call for generate_formula's recursive text arm."""
        text = lambda: gen_expr(depth + 1)
        literal = lambda s: '"' + s.replace('"', '""') + '"'
        count = lambda: str(random.randint(1, 5))

        if fn == "CONCATENATE":
            return f"CONCATENATE({text()}, {text()})"
        if fn in ("LEFT", "RIGHT", "LEFTB", "RIGHTB"):
            return f"{fn}({text()}, {count()})"
        if fn in ("ASC", "CLEAN"):
            return f"{fn}({literal('abc')})"
        if fn in ("LEN", "LENB", "UPPER", "LOWER", "CODE", "T", "UNICODE"):
            return f"{fn}({text()})"
        if fn == "MIDB":
            return f"MIDB({text()}, {random.randint(1, 3)}, {count()})"
        if fn == "EXACT":
            probe = random.choice(["abc", "ABC", "text"])
            return f"EXACT({text()}, {literal(probe)})"
        if fn in ("FIND", "FINDB"):
            haystack = random.choice(["alphabet", "abracadabra", "text value"])
            needle = random.choice(["a", "b", "t"])
            return f"{fn}({literal(needle)}, {literal(haystack)}, {random.randint(1, 2)})"
        if fn in ("SEARCH", "SEARCHB"):
            haystack = random.choice(["Alphabet", "abracadabra", "text value"])
            needle = random.choice(["a", "?", "t"])
            return f"{fn}({literal(needle)}, {literal(haystack)}, {random.randint(1, 2)})"
        if fn == "REPT":
            return f"REPT({text()}, {random.randint(0, 3)})"
        if fn == "SUBSTITUTE":
            source = random.choice(["abracadabra", "foo bar foo", "111-222"])
            old = random.choice(["a", "foo", "1"])
            new = random.choice(["X", "z", ""])
            if random.random() < 0.5:
                return f"SUBSTITUTE({literal(source)}, {literal(old)}, {literal(new)})"
            return f"SUBSTITUTE({literal(source)}, {literal(old)}, {literal(new)}, {random.randint(1, 2)})"
        if fn in ("TEXTAFTER", "TEXTBEFORE"):
            source = random.choice(["left|right", "a,b,c", "prefix--suffix"])
            delim = "|" if "|" in source else ("," if "," in source else "--")
            return f"{fn}({literal(source)}, {literal(delim)})"
        if fn == "UNICHAR":
            # Keep generated output in the console/codepage-safe ASCII range;
            # Unicode storage paths are covered elsewhere, and the Windows
            # harness prints mismatches through a legacy console encoding.
            return f"UNICHAR({random.randint(65, 90)})"

        raise AssertionError(f"FUNCTIONS_TEXT has no generator for {fn}")

    @classmethod
    def _check_text_function_generators(cls):
        """Small coverage self-check: every listed text function has a real emitter."""
        state = random.getstate()
        try:
            gen = cls()
            missing = []
            for fn in cls.FUNCTIONS_TEXT:
                formula = gen._generate_text_expr(fn, lambda _depth=0: '"abc"', 0)
                if fn not in formula.upper():
                    missing.append((fn, formula))
            if missing:
                raise AssertionError(f"FUNCTIONS_TEXT generators missing/renamed: {missing}")
        finally:
            random.setstate(state)

    def generate_formula(self, current_row, current_col, max_row, max_col, min_col=1):
        """Generates a random formula string referencing existing cells or constants."""
        def random_cell_ref():
            return self._random_cell_ref(current_row, min_col, max_col)

        def random_range_ref():
            return self._random_range_ref(current_row, min_col, max_col)

        def gen_expr(depth=0):
            if depth >= 2 or random.random() < 0.4:
                # Leaf node: cell ref, scalar constant, or (if a table is
                # present) a structured reference to a column's header text.
                # The header ref evaluates to a plain scalar string, just
                # like a cell ref would, so it can drop in anywhere.
                roll = random.random()
                if self._has_table() and roll < 0.15:
                    return self._random_structured_header_ref()
                remaining = (roll - 0.15) / 0.85 if self._has_table() else roll
                if remaining < 0.65:
                    return random_cell_ref()
                elif remaining < 0.95:
                    return str(random.randint(-50, 50))
                else:
                    # PI() is a plain zero-arg numeric constant, so it fits
                    # anywhere a scalar leaf does.
                    return "PI()"

            fn_type = random.choice(["binary", "multi_num", "single_num", "logic", "text", "stat_bivariate", "two_num"])

            if fn_type == "binary":
                op = random.choice(["+", "-", "*", "/", "^"])
                left = gen_expr(depth + 1)
                right = gen_expr(depth + 1)
                # A huge base (FACT grows past 1e9 by its argument 13) raised
                # to a very negative exponent underflows the true
                # mathematical result into subnormal territory -- Excel's `^`
                # flattens that all the way to exactly 0 where visi's
                # f64::powf keeps the (correct, nonzero) subnormal value, a
                # divergence with no engine bug behind it (see
                # "docs/excel-discrepancies.md" section 18). Not worth
                # excluding `^` outright for this rare a combination -- just
                # avoid pairing a magnitude-prone left operand with it.
                if op == "^" and left.startswith(("FACT(", "GAMMA(", "EXP(", "PERMUT(", "COMBIN(")):
                    op = random.choice(["+", "-", "*", "/"])
                # A negative base raised to a fractional exponent close to
                # zero is a coin flip in real Excel -- almost always
                # #NUM! (mathematically undefined over the reals, which is
                # what visi's `^` gives every time), but for a handful of
                # specific exponent values it instead returns a real
                # number with no describable pattern separating those from
                # the rest (see "docs/excel-discrepancies.md" section 19).
                # Both `POWER(...)` and a bare nested `^` (e.g. `(F4 ^ -4)`,
                # seed 394540's `LEFT(F4, 3) ^ (F4 ^ -4)`) reliably land in
                # that near-zero/fractional range, so keep either off the
                # right side of `^` rather than chasing which exact
                # exponents Excel's `^` happens to like.
                if op == "^" and ("^" in right or right.startswith("POWER(")):
                    op = random.choice(["+", "-", "*", "/"])
                return f"({left} {op} {right})"

            elif fn_type == "two_num":
                fn = random.choice(self.FUNCTIONS_TWO_NUM)
                a = gen_expr(depth + 1)
                b = gen_expr(depth + 1)
                # A MOD operand built from POWER (or `^`) can put the divisor
                # far above the dividend either by making the divisor huge or
                # by making the dividend tiny. Excel then loses the small
                # operand and returns 0 instead of the true (nonzero)
                # remainder -- visi keeps the mathematically correct value
                # (see "docs/excel-discrepancies.md" section 15, e.g.
                # MOD(36, POWER(-327.3, 69)) and seed 747962's
                # MOD((-5 ^ -16), (-44 * 85))). Not a new bug, just a
                # generator gap: this shape was never fully kept out of the
                # fuzzer.
                if fn == "MOD" and (
                    "POWER(" in a
                    or "^" in a
                    or "POWER(" in b
                    or "^" in b
                    or "PERCENTOF(" in a
                    or "PERCENTOF(" in b
                ):
                    a = str(random.randint(-50, 50))
                    b = str(random.randint(-50, 50) or 1)
                return f"{fn}({a}, {b})"

            elif fn_type == "multi_num":
                fn = random.choice(self.FUNCTIONS_MULTI_NUM)
                roll = random.random()
                if fn == "COUNTBLANK":
                    # Unlike every other entry in FUNCTIONS_MULTI_NUM,
                    # COUNTBLANK takes exactly one range argument -- it's
                    # not a variadic numeric aggregate. Handing it a bare
                    # scalar or a comma-separated arg list (as the general
                    # cases below do) isn't just semantically odd, it's a
                    # formula real Excel's own UI would never let you save:
                    # openpyxl writes it anyway, and the resulting file
                    # silently fails to open in real Excel (no error, no
                    # window, workbook count stays 0) rather than showing a
                    # clean parse error. Confirmed by isolating
                    # `=COUNTBLANK(1,2,3)` in its own workbook. Whole-column
                    # references are avoided here too: Excel counts blanks
                    # across all 1,048,576 rows, while visi evaluates only
                    # the imported sheet extent.
                    arg = self._random_structured_col_ref() if (self._has_table() and roll < 0.3) else random_range_ref()
                elif self._has_table() and roll < 0.22:
                    # Single-column structured reference, e.g. SUM(Sheet1[A]).
                    arg = self._random_structured_col_ref()
                elif self._has_table() and roll < 0.34:
                    # Whole-column A:A / $A:$A style reference into the
                    # formula-free table block. This covers an address form
                    # ordinary rectangular range generation never produced.
                    arg = self._random_table_whole_col_ref()
                elif roll < 0.72:
                    arg = random_range_ref()
                else:
                    arg = f"{gen_expr(depth + 1)}, {gen_expr(depth + 1)}"
                return f"{fn}({arg})"

            elif fn_type == "single_num":
                fn = random.choice(self.FUNCTIONS_SINGLE_NUM)
                arg = gen_expr(depth + 1)
                if fn in ["ROUND", "ROUNDUP", "ROUNDDOWN", "TRUNC"]:
                    digits = random.randint(0, 2)
                    return f"{fn}({arg}, {digits})"
                elif fn in ["NORM.S.DIST"]:
                    return f"{fn}({arg}, TRUE)"
                return f"{fn}({arg})"

            elif fn_type == "stat_bivariate":
                fn = random.choice(self.FUNCTIONS_STAT_BIVARIATE)
                r1 = random_range_ref()
                r2 = random_range_ref()
                return f"{fn}({r1}, {r2})"

            elif fn_type == "logic":
                if random.random() < 0.5:
                    cond = f"({gen_expr(depth + 1)} > {gen_expr(depth + 1)})"
                    val_true = gen_expr(depth + 1)
                    val_false = gen_expr(depth + 1)
                    return f"IF({cond}, {val_true}, {val_false})"
                else:
                    fn = random.choice(["AND", "OR"])
                    return f"{fn}({gen_expr(depth + 1)} > 0, {gen_expr(depth + 1)} < 100)"

            elif fn_type == "text":
                fn = random.choice(self.FUNCTIONS_TEXT)
                return self._generate_text_expr(fn, gen_expr, depth)

        return "=" + gen_expr(0)

    # -- Financial function argument helpers -----------------------------
    def _fin_rate(self, lo=0.001, hi=0.03):
        """Per-period rate. Deliberately realistic (0.1%-3%), not the
        0.5%-20% range an earlier version used: at high per-period rates
        compounded over hundreds of periods (nper goes up to 360 below),
        (1+rate)^nper explodes into territory where computing the
        amortizing payment itself loses essentially all f64 precision --
        confirmed against arbitrary-precision decimal arithmetic while
        chasing real mismatches this generator surfaced against actual
        Excel (see finance.rs's ppmt test for the worked example). That's
        a real f64 floor no closed-form or iterative rewrite escapes, not
        a bug -- so the fix here is to stop generating inputs no real
        financial instrument would ever have anyway.
        """
        return f"{round(random.uniform(lo, hi), 4)}"

    def _fin_money(self, lo=100, hi=50000, allow_negative=True):
        v = round(random.uniform(lo, hi), 2)
        if allow_negative and random.random() < 0.5:
            v = -v
        return f"{v}"

    def _fin_int(self, lo, hi):
        return str(random.randint(lo, hi))

    def _fin_type01(self):
        return str(random.choice([0, 1]))

    def _fin_money_value(self, lo=100, hi=20000, allow_negative=True):
        v = round(random.uniform(lo, hi), 2)
        if allow_negative and random.random() < 0.5:
            v = -v
        return v

    def _fin_date(self, y_lo=1995, y_hi=2035, avoid_february_month_end=False):
        """A DATE(...) literal. Bond/day-count functions below always
        derive related dates (maturity, first coupon, ...) from one of
        these via EDATE(...)/serial-day arithmetic *inside* the generated
        formula, rather than precomputing calendar math in Python -- that
        way both visi and Excel compute the derived date with their own
        (already-validated) date logic instead of risking a Python/Excel
        calendar mismatch that has nothing to do with the function under
        test."""
        y = random.randint(y_lo, y_hi)
        m = random.randint(1, 12)
        d = random.randint(1, 28)
        if avoid_february_month_end and d == 28:
            # Capped at 27, not just rewritten when the month is February.
            # ACCRINT derives its first-interest date with EDATE, so a
            # day-28 issue in *any* month can land on 28 February and turn
            # the whole quasi-coupon schedule end-of-month -- which is the
            # case ACCRINT is known to get wrong. See
            # "docs/excel-discrepancies.md" section 9.
            d = 27
        return f"DATE({y}, {m}, {d})"

    def generate_financial_formula(self, fn=None):
        """Generates a single self-contained financial-function formula
        with semantically valid inputs (small positive rates, periods
        within range, etc.) rather than composing arbitrary sub-expressions
        the way generate_formula() does -- most financial arguments have a
        specific meaning (a rate, a period count) that a random
        sub-expression would violate far too often to be a useful test.

        fn can be passed explicitly to deterministically cycle through
        FINANCIAL_FUNCTIONS (see create_fuzz_workbook) so a single fuzz
        iteration is guaranteed to exercise every function at least once,
        rather than relying on random.choice's luck across many iterations."""
        if fn is None:
            fn = random.choice(self.FINANCIAL_FUNCTIONS)

        if fn in ("PV", "FV", "PMT"):
            rate = self._fin_rate()
            nper = self._fin_int(1, 360)
            middle = self._fin_money(10, 5000)
            other = self._fin_money(0, 5000) if random.random() < 0.6 else "0"
            typ = self._fin_type01()
            return f"={fn}({rate}, {nper}, {middle}, {other}, {typ})"

        if fn == "NPER":
            rate = self._fin_rate()
            pmt = self._fin_money(10, 2000, allow_negative=False)
            pv = self._fin_money(1000, 20000, allow_negative=False)
            typ = self._fin_type01()
            return f"=NPER({rate}, -{pmt}, {pv}, 0, {typ})"

        if fn == "RATE":
            nper = self._fin_int(6, 360)
            pmt = self._fin_money(10, 2000, allow_negative=False)
            pv = self._fin_money(1000, 20000, allow_negative=False)
            typ = self._fin_type01()
            return f"=RATE({nper}, -{pmt}, {pv}, 0, {typ})"

        if fn in ("IPMT", "PPMT"):
            rate = self._fin_rate()
            nper = self._fin_int(2, 360)
            per = self._fin_int(1, int(nper))
            pv = self._fin_money(1000, 100000, allow_negative=False)
            typ = self._fin_type01()
            return f"={fn}({rate}, {per}, {nper}, {pv}, 0, {typ})"

        if fn in ("CUMIPMT", "CUMPRINC"):
            rate = self._fin_rate()
            nper = int(self._fin_int(12, 360))
            start = random.randint(1, nper)
            end = random.randint(start, nper)
            pv = self._fin_money(1000, 100000, allow_negative=False)
            typ = self._fin_type01()
            return f"={fn}({rate}, {nper}, {pv}, {start}, {end}, {typ})"

        # NPV/IRR/MIRR/XNPV/XIRR/FVSCHEDULE take an array argument. visi's
        # parser has no `{...}` array-literal syntax, so these always
        # reference the financial data block create_fuzz_workbook lays out
        # before calling this method -- there is no standalone-call path
        # for this generator, so no inline-array fallback is needed.
        if fn == "NPV":
            assert self._fin_cash_range
            return f"=NPV({self._fin_rate()}, {self._fin_cash_range})"

        if fn == "IRR":
            assert self._fin_cash_range
            return f"=IRR({self._fin_cash_range})"

        if fn == "MIRR":
            assert self._fin_cash_range
            return f"=MIRR({self._fin_cash_range}, {self._fin_rate()}, {self._fin_rate()})"

        if fn == "XNPV":
            assert self._fin_cash_range and self._fin_date_range
            return f"=XNPV({self._fin_rate()}, {self._fin_cash_range}, {self._fin_date_range})"

        if fn == "XIRR":
            assert self._fin_cash_range and self._fin_date_range
            return f"=XIRR({self._fin_cash_range}, {self._fin_date_range})"

        if fn in ("SLN",):
            cost = self._fin_money(1000, 50000, allow_negative=False)
            salvage = self._fin_money(0, 999, allow_negative=False)
            life = self._fin_int(1, 30)
            return f"=SLN({cost}, {salvage}, {life})"

        if fn == "SYD":
            cost = self._fin_money(1000, 50000, allow_negative=False)
            salvage = self._fin_money(0, 999, allow_negative=False)
            life = int(self._fin_int(1, 30))
            per = self._fin_int(1, life)
            return f"=SYD({cost}, {salvage}, {life}, {per})"

        if fn == "DB":
            cost = self._fin_money(1000, 50000, allow_negative=False)
            salvage = self._fin_money(1, 999, allow_negative=False)
            life = int(self._fin_int(1, 20))
            period = self._fin_int(1, life)
            month = self._fin_int(1, 12)
            return f"=DB({cost}, {salvage}, {life}, {period}, {month})"

        if fn == "DDB":
            cost = self._fin_money(1000, 50000, allow_negative=False)
            salvage = self._fin_money(1, 999, allow_negative=False)
            life = int(self._fin_int(1, 20))
            period = self._fin_int(1, life)
            factor = round(random.uniform(1.0, 3.0), 2)
            return f"=DDB({cost}, {salvage}, {life}, {period}, {factor})"

        if fn == "VDB":
            cost = self._fin_money(1000, 50000, allow_negative=False)
            salvage = self._fin_money(1, 999, allow_negative=False)
            life = int(self._fin_int(1, 20))
            start = random.randint(0, life - 1) if life > 1 else 0
            end = random.randint(start + 1, life)
            factor = round(random.uniform(1.0, 3.0), 2)
            no_switch = random.choice(["TRUE", "FALSE"])
            return f"=VDB({cost}, {salvage}, {life}, {start}, {end}, {factor}, {no_switch})"

        if fn == "EFFECT":
            nominal_rate = self._fin_rate()
            npery = self._fin_int(1, 12)
            return f"=EFFECT({nominal_rate}, {npery})"

        if fn == "NOMINAL":
            effect_rate = self._fin_rate()
            npery = self._fin_int(1, 12)
            return f"=NOMINAL({effect_rate}, {npery})"

        if fn in ("DOLLARDE", "DOLLARFR"):
            dollar = round(random.uniform(1.0, 100.0), random.choice([1, 2]))
            fraction = random.choice([2, 4, 8, 16, 32, 64])
            return f"={fn}({dollar}, {fraction})"

        if fn == "FVSCHEDULE":
            assert self._fin_schedule_range
            principal = self._fin_money(1000, 50000, allow_negative=False)
            return f"=FVSCHEDULE({principal}, {self._fin_schedule_range})"

        # RRI and PDURATION were both added in Excel 2013; real Excel's own
        # OOXML writer always stores post-2007 functions with an `_xlfn.`
        # prefix in the underlying formula text (invisible in the formula
        # bar) so older parsers degrade gracefully instead of choking on an
        # unknown name -- confirmed as the actual cause of a real `#NAME?`
        # mismatch this generator produced: openpyxl writes plain
        # `RRI(...)`/`PDURATION(...)` with no prefix, and Excel refuses to
        # recognize the bare name on load. visi already strips a leading
        # `_xlfn.` before dispatch (see evaluate_function in sheet.rs), so
        # writing the prefix here is what a real xlsx producer would do and
        # keeps both sides consistent.
        if fn == "RRI":
            nper = self._fin_int(1, 30)
            pv = self._fin_money(1000, 50000, allow_negative=False)
            fv = self._fin_money(1000, 100000, allow_negative=False)
            return f"=_xlfn.RRI({nper}, {pv}, {fv})"

        if fn == "PDURATION":
            rate = self._fin_rate()
            pv = self._fin_money(1000, 50000, allow_negative=False)
            fv = self._fin_money(1000, 100000, allow_negative=False)
            return f"=_xlfn.PDURATION({rate}, {pv}, {fv})"

        if fn == "ISPMT":
            rate = self._fin_rate()
            nper = int(self._fin_int(2, 360))
            per = self._fin_int(1, nper)
            pv = self._fin_money(1000, 100000, allow_negative=False)
            return f"=ISPMT({rate}, {per}, {nper}, {pv})"

        # -- Day-count / bond-pricing functions --------------------------
        # Settlement is always the anchor DATE(...) literal; every other
        # date (maturity, issue, coupon dates, ...) is derived from it via
        # EDATE(...)/serial-day arithmetic *inside* the generated formula
        # (see _fin_date) so visi and Excel compute the same derived date
        # with their own matching logic instead of racing a Python
        # reimplementation of calendar math against either engine.
        bond_rate = lambda: round(random.uniform(0.01, 0.10), 4)
        bond_basis = lambda: random.choice([0, 1, 2, 3, 4])
        # docs/excel-discrepancies.md #22: COUPDAYS basis 1 has unresolved
        # Excel coupon-period quirks around some quarterly schedules. Keep the
        # other bases fuzzed, and keep basis 1 for the neighbouring functions.
        coupdays_basis = lambda: random.choice([0, 2, 3, 4])
        bond_freq = lambda: random.choice([1, 2, 4])

        if fn in ("COUPDAYBS", "COUPDAYS", "COUPDAYSNC", "COUPNCD", "COUPNUM", "COUPPCD"):
            settlement = self._fin_date()
            freq = bond_freq()
            maturity = f"EDATE({settlement}, {12 // freq * random.randint(2, 20)})"
            if fn in ("COUPNCD", "COUPNUM", "COUPPCD"):
                return f"={fn}({settlement}, {maturity}, {freq})"
            basis = coupdays_basis() if fn == "COUPDAYS" else bond_basis()
            return f"={fn}({settlement}, {maturity}, {freq}, {basis})"

        if fn in ("PRICE", "YIELD"):
            settlement = self._fin_date()
            freq = bond_freq()
            maturity = f"EDATE({settlement}, {12 // freq * random.randint(2, 20)})"
            rate = bond_rate()
            redemption = random.choice([100, 100, 100, 105])
            if fn == "PRICE":
                yld = bond_rate()
                return f"=PRICE({settlement}, {maturity}, {rate}, {yld}, {redemption}, {freq}, {bond_basis()})"
            pr = round(random.uniform(80, 120), 2)
            return f"=YIELD({settlement}, {maturity}, {rate}, {pr}, {redemption}, {freq}, {bond_basis()})"

        if fn in ("DURATION", "MDURATION"):
            settlement = self._fin_date()
            freq = bond_freq()
            maturity = f"EDATE({settlement}, {12 // freq * random.randint(2, 20)})"
            return f"={fn}({settlement}, {maturity}, {bond_rate()}, {bond_rate()}, {freq}, {bond_basis()})"

        if fn in ("DISC", "PRICEDISC", "YIELDDISC"):
            settlement = self._fin_date()
            maturity = f"EDATE({settlement}, {random.randint(1, 24)})"
            redemption = 100
            basis = bond_basis()
            if fn == "DISC":
                pr = round(random.uniform(85, 99), 2)
                return f"=DISC({settlement}, {maturity}, {pr}, {redemption}, {basis})"
            if fn == "PRICEDISC":
                return f"=PRICEDISC({settlement}, {maturity}, {bond_rate()}, {redemption}, {basis})"
            pr = round(random.uniform(85, 99), 2)
            return f"=YIELDDISC({settlement}, {maturity}, {pr}, {redemption}, {basis})"

        if fn in ("PRICEMAT", "YIELDMAT"):
            issue = self._fin_date()
            settlement = f"EDATE({issue}, {random.randint(1, 6)})"
            maturity = f"EDATE({issue}, {random.randint(7, 36)})"
            rate = bond_rate()
            # docs/excel-discrepancies.md #24: PRICEMAT/YIELDMAT basis 0 has
            # unresolved month-end 30/360 leg quirks on issue-anchored schedules.
            basis = random.choice([1, 2, 3, 4])
            if fn == "PRICEMAT":
                return f"=PRICEMAT({settlement}, {maturity}, {issue}, {rate}, {bond_rate()}, {basis})"
            pr = round(random.uniform(85, 120), 2)
            return f"=YIELDMAT({settlement}, {maturity}, {issue}, {rate}, {pr}, {basis})"

        if fn in ("RECEIVED", "INTRATE"):
            settlement = self._fin_date()
            maturity = f"EDATE({settlement}, {random.randint(1, 24)})"
            basis = bond_basis()
            investment = self._fin_money(1000, 50000, allow_negative=False)
            if fn == "RECEIVED":
                return f"=RECEIVED({settlement}, {maturity}, {investment}, {bond_rate()}, {basis})"
            redemption = self._fin_money(1000, 50000, allow_negative=False)
            return f"=INTRATE({settlement}, {maturity}, {investment}, {redemption}, {basis})"

        if fn in ("TBILLPRICE", "TBILLYIELD", "TBILLEQ"):
            settlement = self._fin_date()
            maturity = f"({settlement} + {random.randint(30, 182)})"
            if fn == "TBILLPRICE":
                return f"=TBILLPRICE({settlement}, {maturity}, {bond_rate()})"
            if fn == "TBILLYIELD":
                pr = round(random.uniform(90, 99.9), 2)
                return f"=TBILLYIELD({settlement}, {maturity}, {pr})"
            return f"=TBILLEQ({settlement}, {maturity}, {bond_rate()})"

        if fn == "ACCRINTM":
            issue = self._fin_date()
            settlement = f"EDATE({issue}, {random.randint(1, 24)})"
            par = self._fin_money(1000, 50000, allow_negative=False)
            return f"=ACCRINTM({issue}, {settlement}, {bond_rate()}, {par}, {bond_basis()})"

        if fn == "ACCRINT":
            # Restricted to basis 0/4 (30/360) -- see the doc comment on
            # finance::accrint for why bases 1/2/3 aren't fuzzed here.
            # February month-end issue dates are excluded as a known gap;
            # see "docs/excel-discrepancies.md" section 9.
            issue = self._fin_date(avoid_february_month_end=True)
            freq = bond_freq()
            months = 12 // freq
            first_interest = f"EDATE({issue}, {months})"
            settlement = f"EDATE({issue}, {random.randint(1, 4 * months)})"
            par = self._fin_money(1000, 50000, allow_negative=False)
            calc_method = random.choice(["TRUE", "FALSE"])
            basis = random.choice([0, 4])
            return (
                f"=ACCRINT({issue}, {first_interest}, {settlement}, {bond_rate()}, "
                f"{par}, {freq}, {basis}, {calc_method})"
            )

        if fn in ("AMORLINC", "AMORDEGRC"):
            # life is kept >= 4: AMORDEGRC rejects life <= 2 outright
            # (#NUM!), and real Excel's AMORDEGRC switches early to
            # straight-line for life in (2, 4) in a way not yet
            # reverse-engineered here (see finance::amordegrc's doc
            # comment).
            #
            # period is kept at least 3 below life: real Excel's AMORDEGRC
            # also switches from declining-balance to straight-line for
            # the last couple of periods before life is exhausted (found
            # via the differential fuzzer at life=5, period=4 -- one
            # period short of the end), which this implementation doesn't
            # yet replicate. Periods comfortably before that tail (the
            # margin validated by a full life=12 sequence, periods 0-9)
            # match real Excel exactly.
            date_purchased = self._fin_date()
            first_period = f"EDATE({date_purchased}, {random.randint(1, 11)})"
            cost = self._fin_money_value(1000, 50000, allow_negative=False)
            salvage = round(random.uniform(0, cost * 0.3), 2)
            life = random.randint(4, 20)
            rate = round(1.0 / life, 4)
            period = random.randint(0, max(0, life - 3))
            return f"={fn}({cost}, {date_purchased}, {first_period}, {salvage}, {period}, {rate}, {bond_basis()})"

        if fn in ("ODDFPRICE", "ODDFYIELD"):
            # stub_days is capped to less than one regular coupon period
            # ("short" odd first coupon). visi's implementation doesn't yet
            # match real Excel's undocumented internal algorithm for a
            # "long" odd first coupon (stub longer than a full period) --
            # confirmed as a real, separate discrepancy by the differential
            # fuzzer, distinct from (and on top of) the E/day-count fixes
            # that made the short-stub case match exactly.
            issue = self._fin_date()
            freq = bond_freq()
            period_days = 360 // freq
            stub_days = random.randint(10, max(11, period_days - 15))
            first_coupon = f"({issue} + {stub_days})"
            settlement = f"({issue} + {random.randint(0, stub_days)})"
            maturity = f"EDATE({first_coupon}, {12 // freq * random.randint(2, 10)})"
            rate = bond_rate()
            redemption = random.choice([100, 100, 105])
            basis = bond_basis()
            if fn == "ODDFPRICE":
                yld = bond_rate()
                return (
                    f"=ODDFPRICE({settlement}, {maturity}, {issue}, {first_coupon}, "
                    f"{rate}, {yld}, {redemption}, {freq}, {basis})"
                )
            pr = round(random.uniform(80, 120), 2)
            return (
                f"=ODDFYIELD({settlement}, {maturity}, {issue}, {first_coupon}, "
                f"{rate}, {pr}, {redemption}, {freq}, {basis})"
            )

        if fn in ("ODDLPRICE", "ODDLYIELD"):
            # stub_days is capped to less than one regular coupon period,
            # same reasoning as ODDFPRICE/ODDFYIELD above: a "long" odd
            # last period (longer than one regular period) isn't yet
            # correctly handled here (a real, distinct discrepancy from
            # the E/day-count fixes that made the short-stub case match
            # exactly).
            last_interest = self._fin_date()
            freq = bond_freq()
            period_days = 360 // freq
            stub_days = random.randint(10, max(11, period_days - 15))
            maturity = f"({last_interest} + {stub_days})"
            settlement = f"({last_interest} + {random.randint(0, stub_days - 1)})"
            rate = bond_rate()
            redemption = random.choice([100, 100, 105])
            basis = bond_basis()
            if fn == "ODDLPRICE":
                yld = bond_rate()
                return (
                    f"=ODDLPRICE({settlement}, {maturity}, {last_interest}, "
                    f"{rate}, {yld}, {redemption}, {freq}, {basis})"
                )
            pr = round(random.uniform(80, 120), 2)
            return (
                f"=ODDLYIELD({settlement}, {maturity}, {last_interest}, "
                f"{rate}, {pr}, {redemption}, {freq}, {basis})"
            )

        raise AssertionError(f"no generator wired up for financial function {fn}")

    # -- Statistical distribution / percentile-rank function generator ---
    def generate_distribution_formula(self, fn, value_rows, min_col, max_col):
        """Self-contained formula for one DISTRIBUTION_FUNCTIONS entry, with
        domain-valid scalar parameters (probabilities in (0,1), positive
        shape/scale parameters, integer trial counts, ...) plus a real
        backward-looking range for the array-argument ones (LARGE, SMALL,
        PERCENTILE*, RANK*, FREQUENCY, ...) drawn from the plain-value rows
        of the formula block (rows 1..value_rows in [min_col, max_col],
        which create_fuzz_workbook guarantees never contain formulas)."""
        span = max(1, max_col - min_col + 1)

        def col(offset):
            return self._col_name(min_col + offset)

        rng = f"{col(0)}1:{col(min(1, span - 1))}{value_rows}"

        def p():
            return round(random.uniform(0.05, 0.95), 3)

        def pos(lo=0.5, hi=10):
            return round(random.uniform(lo, hi), 3)

        def cum():
            return random.choice(["TRUE", "FALSE"])

        def n_int(lo=1, hi=30):
            return random.randint(lo, hi)

        if fn in ("BETA.DIST",):
            return f"=BETA.DIST({p()}, {pos()}, {pos()}, {cum()})"
        if fn == "BETADIST":
            return f"=BETADIST({p()}, {pos()}, {pos()})"
        if fn in ("BETA.INV", "BETAINV"):
            return f"={fn}({p()}, {pos()}, {pos()})"
        if fn in ("BINOM.DIST", "BINOMDIST"):
            n = n_int(5, 50)
            k = random.randint(0, n)
            return f"={fn}({k}, {n}, {p()}, {cum()})"
        if fn == "BINOM.DIST.RANGE":
            n = n_int(5, 50)
            k1 = random.randint(0, n)
            k2 = random.randint(k1, n)
            return f"=BINOM.DIST.RANGE({n}, {p()}, {k1}, {k2})"
        if fn in ("BINOM.INV", "CRITBINOM"):
            return f"={fn}({n_int(5, 50)}, {p()}, {p()})"
        if fn == "CHISQ.DIST":
            return f"=CHISQ.DIST({pos(0.1, 20)}, {n_int(1, 30)}, {cum()})"
        if fn in ("CHISQ.DIST.RT", "CHIDIST"):
            return f"={fn}({pos(0.1, 20)}, {n_int(1, 30)})"
        if fn == "CHISQ.INV":
            return f"=CHISQ.INV({p()}, {n_int(1, 30)})"
        if fn in ("CHISQ.INV.RT", "CHIINV"):
            return f"={fn}({p()}, {n_int(1, 30)})"
        if fn in ("CONFIDENCE.NORM", "CONFIDENCE", "CONFIDENCE.T"):
            return f"={fn}({p()}, {pos(0.5, 20)}, {n_int(5, 100)})"
        if fn in ("EXPON.DIST", "EXPONDIST"):
            return f"={fn}({pos(0.1, 5)}, {pos(0.1, 3)}, {cum()})"
        if fn == "F.DIST":
            return f"=F.DIST({pos(0.1, 10)}, {n_int(1, 20)}, {n_int(1, 20)}, {cum()})"
        if fn in ("F.DIST.RT", "FDIST"):
            return f"={fn}({pos(0.1, 10)}, {n_int(1, 20)}, {n_int(1, 20)})"
        if fn == "F.INV":
            return f"=F.INV({p()}, {n_int(1, 20)}, {n_int(1, 20)})"
        if fn in ("F.INV.RT", "FINV"):
            return f"={fn}({p()}, {n_int(1, 20)}, {n_int(1, 20)})"
        if fn in ("GAMMA.DIST", "GAMMADIST"):
            return f"={fn}({pos(0.1, 20)}, {pos()}, {pos()}, {cum()})"
        if fn in ("GAMMA.INV", "GAMMAINV"):
            return f"={fn}({p()}, {pos()}, {pos()})"
        if fn == "HYPGEOM.DIST":
            pop_size = n_int(20, 100)
            pop_s = random.randint(1, pop_size)
            sample_size = random.randint(1, pop_size)
            sample_s = random.randint(0, min(sample_size, pop_s))
            return f"=HYPGEOM.DIST({sample_s}, {sample_size}, {pop_s}, {pop_size}, {cum()})"
        if fn == "HYPGEOMDIST":
            pop_size = n_int(20, 100)
            pop_s = random.randint(1, pop_size)
            sample_size = random.randint(1, pop_size)
            sample_s = random.randint(0, min(sample_size, pop_s))
            return f"=HYPGEOMDIST({sample_s}, {sample_size}, {pop_s}, {pop_size})"
        if fn == "LOGNORM.DIST":
            return f"=LOGNORM.DIST({pos(0.1, 20)}, {pos(0, 3)}, {pos(0.1, 3)}, {cum()})"
        if fn == "LOGNORMDIST":
            return f"=LOGNORMDIST({pos(0.1, 20)}, {pos(0, 3)}, {pos(0.1, 3)})"
        if fn in ("LOGNORM.INV", "LOGINV"):
            return f"={fn}({p()}, {pos(0, 3)}, {pos(0.1, 3)})"
        if fn == "NEGBINOM.DIST":
            return f"=NEGBINOM.DIST({n_int(0, 30)}, {n_int(1, 20)}, {p()}, {cum()})"
        if fn == "NEGBINOMDIST":
            return f"=NEGBINOMDIST({n_int(0, 30)}, {n_int(1, 20)}, {p()})"
        if fn == "NORM.DIST":
            return f"=NORM.DIST({pos(-20, 20)}, {pos(-10, 10)}, {pos(0.1, 5)}, {cum()})"
        if fn == "NORMDIST":
            return f"=NORMDIST({pos(-20, 20)}, {pos(-10, 10)}, {pos(0.1, 5)}, {cum()})"
        if fn in ("NORM.INV", "NORMINV"):
            return f"={fn}({p()}, {pos(-10, 10)}, {pos(0.1, 5)})"
        if fn == "NORM.S.DIST":
            return f"=NORM.S.DIST({pos(-4, 4)}, {cum()})"
        if fn == "NORMSDIST":
            return f"=NORMSDIST({pos(-4, 4)})"
        if fn in ("NORM.S.INV", "NORMSINV"):
            return f"={fn}({p()})"
        if fn == "POISSON.DIST":
            return f"=POISSON.DIST({n_int(0, 30)}, {pos(0.5, 15)}, {cum()})"
        if fn == "POISSON":
            return f"=POISSON({n_int(0, 30)}, {pos(0.5, 15)}, {cum()})"
        if fn == "PROB":
            data_rng = f"{col(0)}1:{col(0)}{value_rows}"
            prob_rng = f"{col(min(1, span - 1))}1:{col(min(1, span - 1))}{value_rows}"
            return f"=PROB({data_rng}, {prob_rng}, {random.randint(-20, 0)}, {random.randint(1, 20)})"
        if fn == "STANDARDIZE":
            return f"=STANDARDIZE({pos(-20, 20)}, {pos(-10, 10)}, {pos(0.1, 5)})"
        if fn == "T.DIST":
            return f"=T.DIST({pos(-5, 5)}, {n_int(1, 30)}, {cum()})"
        if fn in ("T.DIST.2T",):
            return f"=T.DIST.2T({pos(0.1, 5)}, {n_int(1, 30)})"
        if fn == "TDIST":
            return f"=TDIST({pos(0.1, 5)}, {n_int(1, 30)}, {random.choice([1, 2])})"
        if fn == "T.DIST.RT":
            return f"=T.DIST.RT({pos(-5, 5)}, {n_int(1, 30)})"
        if fn == "T.INV":
            return f"=T.INV({p()}, {n_int(1, 30)})"
        if fn in ("T.INV.2T", "TINV"):
            return f"={fn}({p()}, {n_int(1, 30)})"
        if fn in ("T.TEST", "TTEST"):
            r1 = f"{col(0)}1:{col(0)}{value_rows}"
            r2 = f"{col(min(1, span - 1))}1:{col(min(1, span - 1))}{value_rows}"
            return f"={fn}({r1}, {r2}, {random.choice([1, 2])}, {random.choice([1, 2, 3])})"
        if fn == "WEIBULL.DIST":
            return f"=WEIBULL.DIST({pos(0.1, 20)}, {pos()}, {pos()}, {cum()})"
        if fn == "WEIBULL":
            return f"=WEIBULL({pos(0.1, 20)}, {pos()}, {pos()}, {cum()})"
        if fn in ("Z.TEST", "ZTEST"):
            return f"={fn}({rng}, {random.randint(-20, 20)})"
        if fn == "LARGE":
            return f"=LARGE({rng}, {random.randint(1, 3)})"
        if fn == "SMALL":
            return f"=SMALL({rng}, {random.randint(1, 3)})"
        if fn in ("PERCENTILE", "PERCENTILE.INC"):
            return f"={fn}({rng}, {p()})"
        if fn == "PERCENTILE.EXC":
            return f"=PERCENTILE.EXC({rng}, {p()})"
        if fn in ("QUARTILE", "QUARTILE.INC"):
            return f"={fn}({rng}, {random.randint(0, 4)})"
        if fn == "QUARTILE.EXC":
            return f"=QUARTILE.EXC({rng}, {random.randint(1, 3)})"
        if fn in ("PERCENTRANK", "PERCENTRANK.INC", "PERCENTRANK.EXC"):
            return f"={fn}({rng}, {random.randint(-20, 20)})"
        if fn in ("RANK", "RANK.EQ", "RANK.AVG"):
            return f"={fn}({random.randint(-20, 20)}, {rng}, {random.choice([0, 1])})"
        if fn == "TRIMMEAN":
            return f"=TRIMMEAN({rng}, {round(random.uniform(0.05, 0.4), 2)})"
        if fn == "MODE.MULT":
            return f"=INDEX(MODE.MULT({rng}), 1)"
        if fn == "FREQUENCY":
            data_rng = f"{col(0)}1:{col(0)}{value_rows}"
            bins_rng = f"{col(min(1, span - 1))}1:{col(min(1, span - 1))}{min(3, value_rows)}"
            return f"=INDEX(FREQUENCY({data_rng}, {bins_rng}), 1)"

        raise AssertionError(f"no generator wired up for distribution function {fn}")

    # -- Lookup function generator ----------------------------------------
    def generate_lookup_formula(self, fn, value_rows, min_col, max_col):
        """Self-contained INDEX/MATCH/VLOOKUP/HLOOKUP/XLOOKUP formula against
        a real backward-looking range in the plain-value rows of the formula
        block. Exact-match modes (FALSE / match_type 0) are used throughout
        so results stay well-defined regardless of whether the source data
        happens to be sorted."""
        row_hi = max(2, value_rows)

        if fn == "MATCH":
            c = random.randint(min_col, max_col)
            r2 = random.randint(2, row_hi)
            rng = f"{self._col_name(c)}1:{self._col_name(c)}{r2}"
            return f"=MATCH({self._col_name(c)}1, {rng}, 0)"

        if fn == "INDEX":
            r2 = random.randint(2, row_hi)
            c1 = random.randint(min_col, max_col)
            c2 = random.randint(c1, max_col)
            rng = f"{self._col_name(c1)}1:{self._col_name(c2)}{r2}"
            return f"=INDEX({rng}, {random.randint(1, r2)}, {random.randint(1, c2 - c1 + 1)})"

        if fn == "VLOOKUP":
            r2 = random.randint(2, row_hi)
            c1 = random.randint(min_col, max(min_col, max_col - 1))
            c2 = random.randint(c1 + 1, max_col) if max_col > c1 else c1
            table = f"{self._col_name(c1)}1:{self._col_name(c2)}{r2}"
            return f"=VLOOKUP({self._col_name(c1)}1, {table}, {c2 - c1 + 1}, FALSE)"

        if fn == "HLOOKUP":
            r2 = random.randint(2, row_hi)
            c1 = random.randint(min_col, max(min_col, max_col - 1))
            c2 = random.randint(c1 + 1, max_col) if max_col > c1 else c1
            table = f"{self._col_name(c1)}1:{self._col_name(c2)}{r2}"
            return f"=HLOOKUP({self._col_name(c1)}1, {table}, {r2}, FALSE)"

        if fn == "XLOOKUP":
            r2 = random.randint(2, row_hi)
            c1 = random.randint(min_col, max_col)
            others = [c for c in range(min_col, max_col + 1) if c != c1]
            c2 = random.choice(others) if others else c1
            lookup_arr = f"{self._col_name(c1)}1:{self._col_name(c1)}{r2}"
            return_arr = f"{self._col_name(c2)}1:{self._col_name(c2)}{r2}"
            return f"=XLOOKUP({self._col_name(c1)}1, {lookup_arr}, {return_arr})"

        raise AssertionError(f"no generator wired up for lookup function {fn}")

    # -- Engineering / complex-number function generator -------------------
    def generate_engineering_formula(self, fn):
        """Self-contained formula for one ENGINEERING_FUNCTIONS entry.
        Base-conversion functions need digit strings valid in their source
        radix (not arbitrary numeric expressions); complex-number functions
        need "a+bi"/"a+bj" formatted strings, matching the same suffix
        within a single binary call since Excel (and visi's parse_complex)
        errors out on mixed i/j suffixes."""

        def bin_str(n=None):
            n = n or random.randint(1, 8)
            return "".join(random.choice("01") for _ in range(n))

        def hex_str(n=None):
            n = n or random.randint(1, 4)
            return "".join(random.choice("0123456789ABCDEF") for _ in range(n))

        def oct_str(n=None):
            n = n or random.randint(1, 6)
            return "".join(random.choice("01234567") for _ in range(n))

        def cplx(suf=None):
            re_ = random.randint(-9, 9)
            im_ = random.choice([x for x in range(-9, 10) if x != 0])
            suf = suf or random.choice(["i", "j"])
            return f'{re_}{"+" if im_ > 0 else ""}{im_}{suf}'

        if fn == "BIN2DEC":
            return f'=BIN2DEC("{bin_str()}")'
        if fn == "DEC2BIN":
            return f"=DEC2BIN({random.randint(-512, 511)})"
        if fn == "DEC2HEX":
            return f"=DEC2HEX({random.randint(-1000, 1000)})"
        if fn == "DEC2OCT":
            return f"=DEC2OCT({random.randint(-512, 511)})"
        if fn == "DELTA":
            return f"=DELTA({random.randint(-5, 5)}, {random.randint(-5, 5)})"
        if fn == "GESTEP":
            return f"=GESTEP({round(random.uniform(-10, 10), 2)}, {round(random.uniform(-10, 10), 2)})"
        if fn == "HEX2DEC":
            return f'=HEX2DEC("{hex_str(3)}")'
        if fn == "OCT2DEC":
            return f'=OCT2DEC("{oct_str(4)}")'
        if fn in ("BITAND", "BITOR", "BITXOR"):
            return f"={fn}({random.randint(0, 2**20)}, {random.randint(0, 2**20)})"
        if fn == "BIN2HEX":
            return f'=BIN2HEX("{bin_str()}")'
        if fn == "BIN2OCT":
            return f'=BIN2OCT("{bin_str()}")'
        if fn == "HEX2BIN":
            return f'=HEX2BIN("{hex_str(2)}")'
        if fn == "HEX2OCT":
            return f'=HEX2OCT("{hex_str(2)}")'
        if fn == "OCT2BIN":
            return f'=OCT2BIN("{oct_str(3)}")'
        if fn == "OCT2HEX":
            return f'=OCT2HEX("{oct_str(3)}")'
        if fn == "BASE":
            return f"=BASE({random.randint(0, 5000)}, {random.randint(2, 36)})"
        if fn == "DECIMAL":
            radix = random.randint(2, 36)
            digits = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:radix]
            s = "".join(random.choice(digits) for _ in range(random.randint(1, 4)))
            return f'=DECIMAL("{s}", {radix})'
        if fn in ("BITLSHIFT", "BITRSHIFT"):
            return f"={fn}({random.randint(0, 2**20)}, {random.randint(0, 10)})"
        if fn == "CONVERT":
            groups = [("C", "F"), ("C", "K"), ("m", "ft"), ("km", "mi"),
                      ("kg", "lbm"), ("g", "ozm"), ("in", "cm"), ("yd", "m")]
            u1, u2 = random.choice(groups)
            if random.random() < 0.5:
                u1, u2 = u2, u1
            return f'=CONVERT({round(random.uniform(-100, 500), 2)}, "{u1}", "{u2}")'
        # BESSELI/BESSELJ/BESSELK/BESSELY are deliberately absent from
        # ENGINEERING_FUNCTIONS (see issue #94): real Excel cannot serve as
        # an oracle for them because Excel is the inaccurate side. Arbitrated against
        # 60-significant-digit reference values (Decimal evaluation of the
        # ascending series), visi's BESSELJ is accurate to ~1e-16 relative
        # while Excel's error is 3.8e-7 at BESSELJ(2.95, 3), 1.3e-6 at
        # BESSELJ(8.72, 2) and 1.8e-6 at BESSELJ(9.59, 1) -- all far past
        # this comparator's 1e-7 tolerance. The degradation depends on the
        # order as well as the argument, so there is no argument range
        # that keeps Excel trustworthy. visi's own accuracy is pinned
        # directly instead, against those high-precision references, by
        # test_besselj_stays_accurate_where_excel_does_not in
        # visi-core/src/core/engine/tests/extended.rs.
        if fn == "COMPLEX":
            suf = random.choice(["i", "j"])
            return f'=COMPLEX({random.randint(-9, 9)}, {random.randint(-9, 9)}, "{suf}")'
        if fn in ("IMABS", "IMAGINARY", "IMARGUMENT", "IMCONJUGATE", "IMCOS", "IMCOSH", "IMCOT",
                   "IMCSC", "IMCSCH", "IMEXP", "IMLN", "IMLOG10", "IMLOG2", "IMREAL", "IMSEC",
                   "IMSECH", "IMSIN", "IMSINH", "IMSQRT", "IMTAN"):
            return f'={fn}("{cplx()}")'
        if fn == "IMPOWER":
            return f'=IMPOWER("{cplx()}", {random.randint(1, 4)})'
        if fn in ("IMDIV", "IMSUB"):
            suf = random.choice(["i", "j"])
            return f'={fn}("{cplx(suf)}", "{cplx(suf)}")'
        if fn in ("IMSUM", "IMPRODUCT"):
            suf = random.choice(["i", "j"])
            args = ", ".join(f'"{cplx(suf)}"' for _ in range(random.randint(2, 3)))
            return f"={fn}({args})"
        if fn in ("ISO.CEILING", "CEILING", "CEILING.MATH", "CEILING.PRECISE",
                   "FLOOR", "FLOOR.MATH", "FLOOR.PRECISE"):
            num = round(random.uniform(-100, 100), 2)
            sig = random.choice([1, 2, 5, 10, 0.5])
            return f"={fn}({num}, {sig})"
        if fn in ("COMBIN", "COMBINA", "PERMUT", "PERMUTATIONA"):
            n = random.randint(1, 20)
            return f"={fn}({n}, {random.randint(0, n)})"
        if fn == "MROUND":
            return f"=MROUND({round(random.uniform(1, 100), 2)}, {random.choice([2, 3, 5, 10])})"

        raise AssertionError(f"no generator wired up for engineering function {fn}")

    # -- Date function generator --------------------------------------------
    def generate_date_formula(self, fn):
        """Self-contained formula for one DATE_FUNCTIONS entry, using plain
        Excel serial-date integers (visi has no DATE-literal parsing outside
        the DATE() function itself, mirrored by the financial data block's
        own comment about serial dates)."""

        def serial(lo=40000, hi=46000):
            return random.randint(lo, hi)

        if fn == "DATE":
            return f"=DATE({random.randint(1990, 2035)}, {random.randint(1, 12)}, {random.randint(1, 28)})"
        if fn in ("DAY", "MONTH", "YEAR", "HOUR", "MINUTE", "SECOND", "WEEKDAY", "ISOWEEKNUM"):
            return f"={fn}({serial()})"
        if fn == "WEEKNUM":
            return f"=WEEKNUM({serial()}, {random.choice([1, 2, 11, 21])})"
        if fn == "DAYS":
            s1, s2 = serial(), serial()
            return f"=DAYS({max(s1, s2)}, {min(s1, s2)})"
        if fn == "DAYS360":
            s1, s2 = serial(), serial()
            return f"=DAYS360({min(s1, s2)}, {max(s1, s2)}, {random.choice(['TRUE', 'FALSE'])})"
        if fn == "EDATE":
            return f"=EDATE({serial()}, {random.randint(-24, 24)})"
        if fn == "EOMONTH":
            return f"=EOMONTH({serial()}, {random.randint(-24, 24)})"
        if fn == "TIME":
            return f"=TIME({random.randint(0, 23)}, {random.randint(0, 59)}, {random.randint(0, 59)})"
        if fn == "YEARFRAC":
            s1, s2 = serial(), serial()
            return f"=YEARFRAC({min(s1, s2)}, {max(s1, s2)}, {random.choice([0, 1, 2, 3, 4])})"
        if fn == "DATEDIF":
            # "YD" is excluded: Excel's is internally inconsistent and no
            # candidate rule fits more than 5 of 8 probed data points (see
            # "docs/excel-discrepancies.md" section 6). The other units agree.
            s1, s2 = serial(40000, 43000), serial(43001, 46000)
            unit = random.choice(["Y", "M", "D", "MD", "YM"])
            return f'=DATEDIF({s1}, {s2}, "{unit}")'
        if fn == "DATEVALUE":
            if random.choice([True, False]):
                return f"=DATEVALUE({serial()})"
            y, m, d = random.randint(2000, 2035), random.randint(1, 12), random.randint(1, 28)
            return f'=DATEVALUE("{y:04d}-{m:02d}-{d:02d}")'
        if fn == "TIMEVALUE":
            if random.choice([True, False]):
                return f"=TIMEVALUE({serial() + random.random():.12f})"
            h, mi, s = random.randint(0, 23), random.randint(0, 59), random.randint(0, 59)
            return f'=TIMEVALUE("{h:02d}:{mi:02d}:{s:02d}")'
        if fn in ("NETWORKDAYS", "NETWORKDAYS.INTL"):
            s1, s2 = serial(), serial()
            return f"={fn}({min(s1, s2)}, {max(s1, s2)})"
        if fn in ("WORKDAY", "WORKDAY.INTL"):
            return f"={fn}({serial()}, {random.randint(-60, 60)})"

        raise AssertionError(f"no generator wired up for date function {fn}")

    # -- Extra text function generator --------------------------------------
    def generate_text2_formula(self, fn, value_rows, min_col, max_col):
        """Self-contained formula for one TEXT_EXTRA_FUNCTIONS entry. Only
        ARRAYTOTEXT needs a real range (everything else takes literal
        scalars), so value_rows/min_col/max_col are only used there."""
        words = ["hello world", "Excel Formula", "  padded text  ", "MiXeD Case", "foo-bar-baz", "123abc"]

        def txt():
            return random.choice(words)

        if fn == "PROPER":
            return f'=PROPER("{txt()}")'
        if fn == "TRIM":
            return f'=TRIM("{txt()}")'
        if fn == "CHAR":
            return f"=CHAR({random.randint(33, 126)})"
        if fn == "TEXTJOIN":
            parts = ", ".join(f'"{random.choice(["a", "b", "", "c"])}"' for _ in range(4))
            return f'=TEXTJOIN("-", TRUE, {parts})'
        if fn == "TEXTSPLIT":
            return '=INDEX(TEXTSPLIT("a,b,c", ","), 1)'
        if fn == "VALUE":
            return f'=VALUE("{round(random.uniform(-1000, 1000), 2)}")'
        if fn == "VALUETOTEXT":
            return f"=VALUETOTEXT({random.randint(-100, 100)})"
        if fn == "N":
            choice = random.choice(["number", "bool", "text"])
            arg = {"number": str(random.randint(-100, 100)), "bool": random.choice(["TRUE", "FALSE"]), "text": '"hello"'}[choice]
            return f"=N({arg})"
        if fn == "NA":
            return "=NA()"
        if fn == "DOLLAR":
            return f"=DOLLAR({round(random.uniform(-10000, 10000), 2)}, {random.randint(0, 4)})"
        if fn == "FIXED":
            return f"=FIXED({round(random.uniform(-10000, 10000), 2)}, {random.randint(0, 4)}, {random.choice(['TRUE', 'FALSE'])})"
        if fn == "NUMBERVALUE":
            return f'=NUMBERVALUE("{round(random.uniform(-1000, 1000), 2)}")'
        if fn == "ARABIC":
            romans = ["III", "IX", "LVIII", "MCMXCIV", "XL", "CD"]
            return f'=ARABIC("{random.choice(romans)}")'
        if fn == "ROMAN":
            return f"=ROMAN({random.randint(1, 3999)}, {random.randint(0, 4)})"
        if fn == "BAHTTEXT":
            return f"=BAHTTEXT({round(random.uniform(0, 100000), 2)})"
        if fn == "REGEXEXTRACT":
            return '=REGEXEXTRACT("order-12345", "[0-9]+")'
        if fn == "REGEXREPLACE":
            return '=REGEXREPLACE("order-12345", "[0-9]+", "X")'
        if fn == "REGEXTEST":
            return '=REGEXTEST("order-12345", "[0-9]+")'
        if fn in ("REPLACE", "REPLACEB"):
            return f'={fn}("{txt()}", {random.randint(1, 5)}, {random.randint(1, 4)}, "NEW")'
        if fn == "CONCAT":
            return f'=CONCAT("{txt()}", "{txt()}")'
        if fn == "ERROR.TYPE":
            return "=ERROR.TYPE(1/0)"
        if fn == "MID":
            return f'=MID("{txt()}", {random.randint(1, 5)}, {random.randint(1, 6)})'
        if fn == "TEXT":
            fmts = ["0.00", "#,##0", "0%", "$#,##0.00", "yyyy-mm-dd"]
            return f'=TEXT({round(random.uniform(-10000, 10000), 4)}, "{random.choice(fmts)}")'
        if fn == "ADDRESS":
            return f"=ADDRESS({random.randint(1, 100)}, {random.randint(1, 20)}, {random.randint(1, 4)})"
        if fn == "ARRAYTOTEXT":
            rng = self._random_range_ref(value_rows + 1, min_col, max_col)
            return f"=ARRAYTOTEXT({rng})"
        if fn == "ENCODEURL":
            return '=ENCODEURL("hello world/test?a=1&b=2")'

        raise AssertionError(f"no generator wired up for text function {fn}")

    # -- Logic/information function generator --------------------------------
    def generate_logic_formula(self, fn, value_rows, min_col, max_col):
        """Self-contained formula for one LOGIC_EXTRA_FUNCTIONS entry. Where
        an operand is just "some numeric expression", reuses
        generate_formula() itself (stripping its leading '=') rather than
        re-implementing sub-expression generation -- passing
        current_row=value_rows+1 keeps any cell/range refs it emits confined
        to the plain-value rows, same as the other bespoke generators."""

        def expr():
            return self.generate_formula(value_rows + 1, 0, value_rows, max_col, min_col)[1:]

        if fn == "IFERROR":
            return f"=IFERROR({expr()}, {expr()})"
        if fn == "IFNA":
            return f"=IFNA({expr()}, {expr()})"
        if fn == "IFS":
            return f"=IFS({expr()}>0, {expr()}, TRUE, {expr()})"
        if fn == "SWITCH":
            return f"=SWITCH({random.randint(1, 3)}, 1, {expr()}, 2, {expr()}, {expr()})"
        if fn in ("ISBLANK", "ISERR", "ISERROR", "ISNA", "ISNUMBER", "ISTEXT", "ISLOGICAL", "ISNONTEXT"):
            return f"={fn}({expr()})"
        if fn in ("ISEVEN", "ISODD"):
            return f"={fn}(INT({expr()}))"
        if fn == "TYPE":
            return f"=TYPE({expr()})"
        if fn == "XOR":
            return f"=XOR({expr()}>0, {expr()}<0)"
        if fn == "CHOOSE":
            n = random.randint(2, 4)
            idx = random.randint(1, n)
            choices = ", ".join(expr() for _ in range(n))
            return f"=CHOOSE({idx}, {choices})"
        # ISOMITTED is intentionally not listed in LOGIC_EXTRA_FUNCTIONS:
        # Excel only accepts it inside LAMBDA, and this harness cannot author
        # LAMBDA formulas into workbooks Excel reliably opens (see the
        # LAMBDA_FUNCTIONS comment). Direct ISOMITTED(A1) is #VALUE! in Excel,
        # while visi's supported subset is scoped to lambda parameter lookup.

        raise AssertionError(f"no generator wired up for logic function {fn}")

    # -- Array/matrix function generator -------------------------------------
    def generate_ets_formula(self, fn):
        """One FORECAST.ETS-family call against the workbook's ETS block.

        The season length is always passed explicitly (`_ets_period`, 0 for
        a pure-trend series) rather than left to Excel's auto-detection --
        see the ETS_FUNCTIONS comment for why. `_ets_next_target` is the
        first timeline point past the end of the series, so the forecast is
        a genuine extrapolation; Excel returns #NUM! for a target inside the
        timeline.
        """
        v = self._ets_values_range
        t = self._ets_timeline_range
        target = self._ets_next_target
        season = self._ets_period
        if fn == "FORECAST.ETS":
            return f"=FORECAST.ETS({target + random.randint(0, 1)}, {v}, {t}, {season})"
        if fn == "FORECAST.ETS.CONFINT":
            # Signature puts confidence_level *before* seasonality.
            return f"=FORECAST.ETS.CONFINT({target}, {v}, {t}, 0.95, {season})"
        if fn.startswith("FORECAST.ETS.STAT"):
            stat = int(fn[-1])
            return f"=FORECAST.ETS.STAT({v}, {t}, {stat}, {season})"

        raise AssertionError(f"no generator wired up for ETS function {fn}")

    def generate_array_formula(self, fn, value_rows, min_col, max_col):
        """Self-contained formula for one ARRAY_FUNCTIONS entry. Matrix
        functions (MDETERM/MINVERSE/MMULT) use small square subranges of the
        plain-value area so shapes are always compatible; MMULT's second
        operand is placed in a disjoint column block when there's room, else
        reuses the first so the call is still valid."""
        span = max(1, max_col - min_col + 1)

        def col(offset):
            return self._col_name(min_col + offset)

        dim = max(1, min(2, span, value_rows))

        if fn == "MDETERM":
            return f"=MDETERM({col(0)}1:{col(dim - 1)}{dim})"
        if fn == "MINVERSE":
            return f"=INDEX(MINVERSE({col(0)}1:{col(dim - 1)}{dim}), 1, 1)"
        if fn == "MMULT":
            a_rng = f"{col(0)}1:{col(dim - 1)}{dim}"
            b_off = dim if dim + dim <= span else 0
            b_rng = f"{col(b_off)}1:{col(b_off + dim - 1)}{dim}"
            return f"=INDEX(MMULT({a_rng}, {b_rng}), 1, 1)"
        if fn == "MUNIT":
            return f"=INDEX(MUNIT({random.randint(1, 4)}), 1, 1)"
        if fn == "SEQUENCE":
            return f"=INDEX(SEQUENCE({random.randint(1, 4)}, {random.randint(1, 4)}, {random.randint(-5, 5)}, {random.randint(1, 3)}), 1, 1)"
        # The FORECAST.ETS family is generated by generate_ets_formula
        # instead -- it needs the dedicated regular-timeline block and its
        # signatures differ from the plain regression functions here.
        if fn in ("LINEST", "LOGEST", "GROWTH", "TREND", "FORECAST", "FORECAST.LINEAR"):
            c2_off = 1 if span > 1 else 0
            ys = f"{col(0)}1:{col(0)}{value_rows}"
            xs = f"{col(c2_off)}1:{col(c2_off)}{value_rows}"
            if fn in ("LINEST", "LOGEST"):
                return f"=INDEX({fn}({ys}, {xs}), 1)"
            if fn in ("TREND", "GROWTH"):
                return f"={fn}({ys}, {xs}, {random.randint(1, value_rows + 2)})"
            return f"={fn}({random.randint(1, value_rows + 2)}, {ys}, {xs})"
        if fn == "SERIESSUM":
            coeffs = f"{col(0)}1:{col(0)}{min(4, value_rows)}"
            return f"=SERIESSUM({round(random.uniform(0.1, 2), 2)}, {random.randint(0, 3)}, {random.randint(1, 2)}, {coeffs})"

        raise AssertionError(f"no generator wired up for array function {fn}")

    # -- Conditional-aggregation function generator --------------------------
    def generate_conditional_formula(self, fn, value_rows, min_col, max_col):
        """Self-contained formula for one CONDITIONAL_AGG_FUNCTIONS entry."""
        span = max(1, max_col - min_col + 1)

        def col(offset):
            return self._col_name(min_col + offset)

        rng = f"{col(0)}1:{col(0)}{value_rows}"
        rng2_off = 1 if span > 1 else 0
        rng2 = f"{col(rng2_off)}1:{col(rng2_off)}{value_rows}"
        crit = f'">{random.randint(-20, 20)}"'

        if fn == "AVERAGEIF":
            return f"=AVERAGEIF({rng}, {crit}, {rng2})"
        if fn == "AVERAGEIFS":
            return f"=AVERAGEIFS({rng2}, {rng}, {crit})"
        if fn == "COUNTIF":
            return f"=COUNTIF({rng}, {crit})"
        if fn == "COUNTIFS":
            return f"=COUNTIFS({rng}, {crit})"
        if fn == "SUMIF":
            return f"=SUMIF({rng}, {crit}, {rng2})"
        if fn == "SUMIFS":
            return f"=SUMIFS({rng2}, {rng}, {crit})"
        if fn == "MAXIFS":
            return f"=MAXIFS({rng2}, {rng}, {crit})"
        if fn == "MINIFS":
            return f"=MINIFS({rng2}, {rng}, {crit})"
        if fn == "SUBTOTAL":
            return f"=SUBTOTAL({random.choice([1, 2, 4, 5, 9])}, {rng})"
        if fn == "AGGREGATE":
            # Real Excel form is AGGREGATE(function_num, options, ref...).
            return f"=AGGREGATE({random.choice([1, 4, 5, 9])}, 6, {rng})"

        raise AssertionError(f"no generator wired up for conditional-aggregate function {fn}")

    # -- Volatile function generator ------------------------------------------
    def generate_volatile_formula(self, fn):
        """RANDBETWEEN/RANDARRAY force min==max for a deterministic result;
        RAND/NOW/TODAY have no such knob so they're wrapped in a
        plausibility/range check both engines must satisfy regardless of the
        actual random or wall-clock value (see VOLATILE_FUNCTIONS)."""
        if fn == "RAND":
            return "=IF(AND(RAND()>=0,RAND()<1),1,0)"
        if fn == "RANDBETWEEN":
            k = random.randint(1, 1000)
            return f"=RANDBETWEEN({k}, {k})"
        if fn == "RANDARRAY":
            k = random.randint(1, 1000)
            return f"=INDEX(RANDARRAY(1,1,{k},{k}),1,1)"
        if fn == "NOW":
            return "=IF(NOW()>40000,1,0)"
        if fn == "TODAY":
            return "=IF(TODAY()>40000,1,0)"

        raise AssertionError(f"no generator wired up for volatile function {fn}")

    def generate_database_formula(self, fn, ws, crit_col, index):
        """Writes a 2-row criteria block (header + one comparison
        criterion) at rows `2*index+1`/`2*index+2` of `crit_col`, then
        returns a D* formula. Reuses the table block's own column-letter
        headers and random data rows (self._table_cols / self._db_range)
        as the "database" argument, since it already has exactly the
        header-row-plus-data-rows shape DSUM/DGET/etc. expect. Each of the
        12 DATABASE_FUNCTIONS entries gets its own criteria row pair so
        they don't clobber each other's criteria in the final workbook."""
        field = random.choice(self._table_cols)
        header_row = 2 * index + 1
        crit_row = 2 * index + 2
        ws.cell(row=header_row, column=crit_col, value=field)
        threshold = round(random.uniform(-500, 500), 2)
        op = random.choice([">", "<", ">=", "<="])
        ws.cell(row=crit_row, column=crit_col, value=f"{op}{threshold}")
        crit_col_letter = self._col_name(crit_col)
        crit_range = f"{crit_col_letter}{header_row}:{crit_col_letter}{crit_row}"
        field_arg = f'"{field}"' if random.random() < 0.7 else str(self._table_cols.index(field) + 1)
        return f"={fn}({self._db_range}, {field_arg}, {crit_range})"

    def generate_lambda_formula(self, fn, value_rows, min_col, max_col):
        """Self-contained MAP/REDUCE formula against a real range from the
        plain-value rows of the formula block (see generate_logic_formula
        for why this reuses that area instead of gen_expr's arbitrary
        substitution). MAP's result is wrapped in INDEX to pin down a
        single scalar the same way other array-returning functions are
        tested (see ARRAY_FUNCTIONS); REDUCE already returns a scalar."""
        def col(offset):
            return self._col_name(min_col + offset)

        rng = f"{col(0)}1:{col(0)}{value_rows}"
        if fn == "MAP":
            return f"=INDEX(MAP({rng}, LAMBDA(x, x*2+1)), 1)"
        if fn == "REDUCE":
            return f"=REDUCE(0, {rng}, LAMBDA(acc,v, acc+v))"

        raise AssertionError(f"no generator wired up for lambda function {fn}")

    def generate_range_info_formula(self, fn, value_rows, min_col, max_col):
        """Self-contained formula for one RANGE_INFO_FUNCTIONS entry,
        against a real cell/range from the plain-value rows of the formula
        block. FORMULATEXT/ISFORMULA/SHEETS/SHEET are post-2007 functions real
        Excel's own OOXML writer always stores with an `_xlfn.` prefix
        (see AGENTS.md/RRI's comment above in generate_financial_formula);
        openpyxl doesn't add that prefix automatically, so it's supplied
        here -- confirmed as the actual cause of a real #NAME? mismatch
        this generator produced without it."""
        def col(offset):
            return self._col_name(min_col + offset)

        cell = f"{col(0)}1"
        rng = f"{col(0)}1:{col(0)}{value_rows}"

        if fn == "ROW":
            # ROW() against a multi-row range spills an array in real
            # Excel, same dynamic-array-automation risk noted for
            # BYROW/BYCOL/MAKEARRAY/SCAN above -- INDEX pins it to a
            # single scalar the same way ARRAY_FUNCTIONS does.
            return f"=INDEX(ROW({rng}), 1)"
        if fn == "ROWS":
            return f"=ROWS({rng})"
        if fn == "COLUMN":
            return f"=COLUMN({cell})"
        if fn == "COLUMNS":
            return f"=COLUMNS({rng})"
        if fn == "AREAS":
            return f"=AREAS({rng})"
        if fn == "ISREF":
            return f"=ISREF({cell})"
        if fn == "FORMULATEXT":
            return f"=IFERROR(_xlfn.FORMULATEXT({cell}), \"none\")"
        if fn == "ISFORMULA":
            return f"=_xlfn.ISFORMULA({cell})"
        if fn == "HYPERLINK":
            return f'=HYPERLINK("https://example.com/{random.randint(1, 1000)}", {cell})'
        if fn == "SHEETS":
            return "=_xlfn.SHEETS()"
        if fn == "SHEET":
            # With a real second sheet now in the workbook (see
            # create_fuzz_workbook's cross-sheet block), this is no longer
            # a trivially-always-1 check in either engine.
            return "=_xlfn.SHEET()"
        if fn == "INDIRECT":
            return f'=SUM(INDIRECT("{rng}"))'
        if fn == "OFFSET":
            return f"=SUM(OFFSET({cell}, 0, 0, {value_rows}, 1))"
        if fn == "CELL":
            # Avoid "contents" here: the referenced input cell may be blank,
            # and OOXML/openpyxl blank-vs-empty-string observability is already
            # documented as unsuitable for random differential fuzzing.
            info_type = random.choice(["row", "col", "address"])
            return f'=CELL("{info_type}", {cell})'
        if fn == "INFO":
            info_type = random.choice(["numfile", "release", "system"])
            return f'=INFO("{info_type}")'

        raise AssertionError(f"no generator wired up for range-info function {fn}")

    def generate_array_reshape_formula(self, fn, value_rows, min_col, max_col):
        """Self-contained formula for one ARRAY_RESHAPE_FUNCTIONS entry.
        All of these are dynamic-array worksheet functions that real
        Excel's own OOXML writer nests under a double `_xlfn._xlws.`
        prefix (confirmed directly against real Excel's export XML for
        UNIQUE/SORT/FILTER; see sheet.rs's `evaluate_function` prefix
        stripping) -- openpyxl doesn't add either prefix automatically, so
        it's supplied here. Every formula is wrapped in SUM or INDEX
        rather than left bare, since a bare dynamic-array-spilling formula
        makes this environment's Excel AppleScript automation bridge
        intermittently fail (confirmed with a plain `=SEQUENCE(3)`, see
        LAMBDA_FUNCTIONS above) -- wrapping pins the result to a single
        cell the same way MAP/REDUCE already do."""
        def col(offset):
            return self._col_name(min_col + offset)

        P = "_xlfn._xlws."
        rng = f"{col(0)}1:{col(0)}{value_rows}"
        rng2 = f"{col(1)}1:{col(1)}{value_rows}"
        wide = f"{col(0)}1:{col(1)}{value_rows}"

        if fn == "HSTACK":
            return f"=SUM({P}HSTACK({rng},{rng2}))"
        if fn == "VSTACK":
            return f"=SUM({P}VSTACK({rng},{rng2}))"
        if fn == "CHOOSEROWS":
            return f"=INDEX({P}CHOOSEROWS({rng},1,-1),1)"
        if fn == "CHOOSECOLS":
            return f"=INDEX({P}CHOOSECOLS({wide},2),1)"
        if fn == "DROP":
            return f"=SUM({P}DROP({rng},1))"
        if fn == "TAKE":
            return f"=SUM({P}TAKE({rng},2))"
        if fn == "EXPAND":
            return f"=INDEX({P}EXPAND({rng},{value_rows}+2,1,0),{value_rows}+2,1)"
        if fn == "TOCOL":
            return f"=SUM({P}TOCOL({rng}))"
        if fn == "TOROW":
            return f"=SUM({P}TOROW({rng}))"
        if fn == "WRAPROWS":
            return f"=INDEX({P}WRAPROWS({rng},2,0),1,1)"
        if fn == "WRAPCOLS":
            return f"=INDEX({P}WRAPCOLS({rng},2,0),1,1)"
        if fn == "UNIQUE":
            return f"=SUM({P}UNIQUE({rng}))"
        if fn == "SORT":
            return f"=INDEX({P}SORT({rng},1,-1),1)"
        if fn == "SORTBY":
            return f"=INDEX({P}SORTBY({rng},{rng2},-1),1)"
        if fn == "FILTER":
            # IFERROR must wrap the whole SUM, not FILTER directly: real
            # Excel implicitly reduces a dynamic-array function nested
            # inside a non-array-aware function like IFERROR to its
            # top-left cell unless the formula is CSE-entered (the same
            # root cause documented for TRANSPOSE above) -- confirmed by
            # reproducing an apparent UNIQUE mismatch this exact way
            # (excel=-47.171 i.e. just the first element, visi=-69.171
            # the correct full-array sum) and showing it disappears once
            # IFERROR moves outside SUM.
            return f"=IFERROR(SUM({P}FILTER({rng},{self._bool_range})),0)"
        if fn == "TRIMRANGE":
            return f"=SUM({P}TRIMRANGE({rng}))"
        if fn == "XMATCH":
            return f"=_xlfn.XMATCH({col(0)}1,{rng})"

        raise AssertionError(f"no generator wired up for array-reshape function {fn}")

    def create_fuzz_workbook(self, file_path, num_rows=10, num_cols=5):
        """Creates a workbook with a mixture of raw values and formulas, plus
        a real Excel Table for structured references to resolve against.

        The sheet is laid out as two disjoint blocks of `num_cols` columns
        side by side: a "table" block (pure random values, every row) and a
        "formula" block (values on top, generated formulas below, exactly
        like the original single-block layout). They're kept disjoint
        because visi treats a structured reference to a table column as
        spanning that column's *entire* height, not just the table's
        declared row range -- if a table column could also hold a formula,
        a structured reference into it could create a dependency cycle
        (directly, or transitively through another formula). Since the
        table block never contains a formula, that's impossible by
        construction rather than merely unlikely.
        """
        try:
            import openpyxl
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            print("Error: 'openpyxl' is required for generating .xlsx test files.")
            print("Please run: pip install openpyxl")
            sys.exit(1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # --- Table block: columns 1..num_cols. Row 1 holds deterministic
        # column headers ("A", "B", ...); visi has no concept of an Excel
        # Table distinct from the worksheet itself, and on import always
        # names each column by its plain spreadsheet letter
        # (col_idx_to_letters) -- using that same scheme as the real header
        # text means `Sheet1[A]` resolves to the same column in both visi
        # (worksheet/column-name lookup) and Excel (Table lookup). Every
        # other row is a plain random value -- never a formula.
        header_names = [self._col_name(c) for c in range(1, num_cols + 1)]
        for c, name in enumerate(header_names, start=1):
            ws.cell(row=1, column=c, value=name)
        for r in range(2, num_rows + 1):
            for c in range(1, num_cols + 1):
                val = self.generate_random_value()
                if val is not None:
                    ws.cell(row=r, column=c, value=val)

        table_name = ws.title
        self._table_name = table_name
        self._table_cols = header_names
        table_ref = f"A1:{self._col_name(num_cols)}{num_rows}"
        self._db_range = table_ref
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)

        # --- Formula block: a second block of num_cols columns immediately
        # to the right of the table. Top rows hold raw random values (for
        # plain cell/range refs to point at); the rest hold generated
        # formulas referencing earlier rows in this same block, optionally
        # including structured references into the table block (see
        # generate_formula / _random_structured_col_ref /
        # _random_structured_header_ref).
        value_rows = max(2, num_rows // 2)
        min_col = num_cols + 1
        max_col = num_cols * 2
        for r in range(1, value_rows + 1):
            for c in range(min_col, max_col + 1):
                val = self.generate_random_value()
                if val is not None:
                    ws.cell(row=r, column=c, value=val)

        for r in range(value_rows + 1, num_rows + 1):
            for c in range(min_col, max_col + 1):
                if random.random() < 0.85:
                    formula = self.generate_formula(r, c, num_rows, max_col, min_col=min_col)
                    ws.cell(row=r, column=c, value=formula)
                else:
                    val = self.generate_random_value()
                    if val is not None:
                        ws.cell(row=r, column=c, value=val)

        # --- Financial data block: a small "cashflow" column, an aligned
        # ascending "date" column (plain Excel serial numbers, since visi
        # has no DATE() function to build one from parts), and a "schedule"
        # column of small rates, all plain values (never formulas). These
        # back the array-argument financial functions (NPV/IRR/MIRR/XNPV/
        # XIRR/FVSCHEDULE) via real ranges, since visi's parser doesn't
        # support `{...}` array literals the way generate_financial_formula
        # falls back to when this block hasn't been laid out.
        fin_cash_col = max_col + 2
        fin_date_col = fin_cash_col + 1
        fin_schedule_col = fin_cash_col + 2
        fin_bool_col = fin_cash_col + 3
        fin_formula_col = fin_cash_col + 4

        # A conventional investment profile: one negative outlay up front
        # followed by positive returns that more than repay it. That gives
        # exactly one sign change, so IRR/XIRR/MIRR are guaranteed a unique
        # positive root.
        #
        # The cashflows used to be `-outlay` followed by five *randomly
        # signed* amounts, which routinely produced series with no real
        # rate of return at all -- and real Excel does not report #NUM! for
        # those, it returns a non-answer. Checked directly on three series
        # this harness generated: Excel's XIRR returned -0.92945409,
        # 2.98e-09 and -0.89982008, but XNPV evaluated at those very rates
        # is -184430.99, -34415.90 and -8804.04 -- nowhere near zero -- and
        # for the first two, XNPV has no sign change anywhere in
        # (-0.999, 10), i.e. no root exists to find. visi answers #NUM!,
        # which is right; comparing against Excel's output there would be
        # asserting Excel's non-convergence garbage as the expected value.
        cash_rows = 6
        outlay = round(random.uniform(5000, 50000), 2)
        ws.cell(row=1, column=fin_cash_col, value=-outlay)
        # Split a total strictly greater than the outlay across the
        # remaining periods so the series always turns a profit.
        weights = [random.uniform(0.5, 1.5) for _ in range(cash_rows - 1)]
        total_return = outlay * random.uniform(1.05, 2.5)
        scale = total_return / sum(weights)
        for i, r in enumerate(range(2, cash_rows + 1)):
            ws.cell(row=r, column=fin_cash_col, value=round(weights[i] * scale, 2))

        date_serial = random.randint(40000, 45000)
        for r in range(1, cash_rows + 1):
            date_serial += random.randint(15, 90)
            ws.cell(row=r, column=fin_date_col, value=date_serial)

        schedule_rows = 3
        for r in range(1, schedule_rows + 1):
            ws.cell(row=r, column=fin_schedule_col, value=round(random.uniform(0.01, 0.15), 4))

        # A dedicated column of genuine boolean literals (not a broadcast
        # comparison -- this engine's comparison operators don't broadcast
        # across a range, e.g. `A1:A5>10` yields a scalar, not an array,
        # a separate, pre-existing, out-of-scope limitation) for FILTER's
        # include_array, which real Excel requires to actually be boolean.
        # Alternating rather than random so it's guaranteed to contain at
        # least one TRUE and one FALSE regardless of value_rows.
        for r in range(1, value_rows + 1):
            ws.cell(row=r, column=fin_bool_col, value=(r % 2 == 0))

        self._fin_cash_range = f"{self._col_name(fin_cash_col)}1:{self._col_name(fin_cash_col)}{cash_rows}"
        self._fin_date_range = f"{self._col_name(fin_date_col)}1:{self._col_name(fin_date_col)}{cash_rows}"
        self._fin_schedule_range = (
            f"{self._col_name(fin_schedule_col)}1:{self._col_name(fin_schedule_col)}{schedule_rows}"
        )
        self._bool_range = f"{self._col_name(fin_bool_col)}1:{self._col_name(fin_bool_col)}{value_rows}"

        # financial_formula_rows is sized to len(FINANCIAL_FUNCTIONS) and the
        # column cycles through the list deterministically (rather than
        # random.choice) so a single fuzz iteration is guaranteed to
        # exercise every financial function at least once, instead of
        # leaving coverage to chance across many --iterations runs.
        financial_formula_rows = max(len(self.FINANCIAL_FUNCTIONS), num_rows)
        for r in range(1, financial_formula_rows + 1):
            fn = self.FINANCIAL_FUNCTIONS[(r - 1) % len(self.FINANCIAL_FUNCTIONS)]
            ws.cell(row=r, column=fin_formula_col, value=self.generate_financial_formula(fn))

        # --- Bespoke-generator blocks: one column per category, each row
        # cycling deterministically through that category's function list so
        # every genuinely-implemented function (per docs/missing-excel-
        # formulas.md) is exercised at least once per fuzz iteration. Laid
        # out sequentially to the right of the financial block; each
        # generator reaches back into the plain-value rows (1..value_rows)
        # of the formula block, which are guaranteed never to contain a
        # formula, so there's no risk of a dependency cycle regardless of
        # which of these columns comes first.
        #
        # DETECTLANGUAGE, TRANSLATE, and PHONETIC are genuinely implemented
        # in visi (not stubs -- see core/text.rs) but are deliberately left
        # out of this differential harness: real Excel's versions call
        # Microsoft's cloud translation/language-detection services, so
        # their output depends on network access and service state that
        # this environment doesn't control, and isn't comparable to visi's
        # local implementation even when both succeed. LET has real
        # variable-binding support in visi (see LetScope in sheet.rs) but
        # is excluded from LOGIC_EXTRA_FUNCTIONS below for an unrelated,
        # environment-specific reason -- see the comment on
        # LOGIC_EXTRA_FUNCTIONS itself.
        next_col = fin_formula_col + 2

        def emit_block(fn_list, formula_for):
            nonlocal next_col
            col = next_col
            for i, fn in enumerate(fn_list, start=1):
                ws.cell(row=i, column=col, value=formula_for(fn))
            next_col = col + 1

        emit_block(self.DISTRIBUTION_FUNCTIONS, lambda fn: self.generate_distribution_formula(fn, value_rows, min_col, max_col))
        emit_block(self.LOOKUP_FUNCTIONS, lambda fn: self.generate_lookup_formula(fn, value_rows, min_col, max_col))
        emit_block(self.ENGINEERING_FUNCTIONS, lambda fn: self.generate_engineering_formula(fn))
        emit_block(self.DATE_FUNCTIONS, lambda fn: self.generate_date_formula(fn))
        emit_block(self.TEXT_EXTRA_FUNCTIONS, lambda fn: self.generate_text2_formula(fn, value_rows, min_col, max_col))
        emit_block(self.LOGIC_EXTRA_FUNCTIONS, lambda fn: self.generate_logic_formula(fn, value_rows, min_col, max_col))
        emit_block(self.ARRAY_FUNCTIONS, lambda fn: self.generate_array_formula(fn, value_rows, min_col, max_col))

        # --- ETS block: a regular timeline plus an exactly-modelled series.
        # Both the trend and the seasonal offsets are whole numbers so the
        # series is representable exactly in binary floating point, which
        # keeps "the model fits perfectly" true to the last bit rather than
        # only to within rounding.
        ets_time_col = next_col
        ets_value_col = next_col + 1
        next_col += 2

        ets_len = 16
        # Half the time a pure trend (no seasonal term at all), half the
        # time a trend plus an exactly repeating season. Both the slope and
        # the offsets are whole numbers, so "the model fits this perfectly"
        # holds to the last bit rather than only to within rounding -- which
        # is what makes the forecast independent of the smoothing
        # parameters, and therefore comparable across two implementations
        # that optimize them differently.
        period = random.choice([0, 2, 4])
        base = random.randint(10, 200)
        slope = random.randint(1, 4)
        if period:
            # Offsets sum to zero so they don't bias the trend.
            half = [random.randint(1, 12) for _ in range(period // 2)]
            offsets = half + [-x for x in half]
            random.shuffle(offsets)
        else:
            offsets = [0]

        for i in range(ets_len):
            ws.cell(row=i + 1, column=ets_time_col, value=i + 1)
            ws.cell(
                row=i + 1,
                column=ets_value_col,
                value=base + slope * i + offsets[i % len(offsets)],
            )
        tcol = self._col_name(ets_time_col)
        vcol = self._col_name(ets_value_col)
        self._ets_timeline_range = f"{tcol}1:{tcol}{ets_len}"
        self._ets_values_range = f"{vcol}1:{vcol}{ets_len}"
        self._ets_next_target = ets_len + 1
        self._ets_period = period

        emit_block(self.ETS_FUNCTIONS, lambda fn: self.generate_ets_formula(fn))
        emit_block(self.CONDITIONAL_AGG_FUNCTIONS, lambda fn: self.generate_conditional_formula(fn, value_rows, min_col, max_col))
        emit_block(self.VOLATILE_FUNCTIONS, lambda fn: self.generate_volatile_formula(fn))
        emit_block(self.LAMBDA_FUNCTIONS, lambda fn: self.generate_lambda_formula(fn, value_rows, min_col, max_col))
        emit_block(self.RANGE_INFO_FUNCTIONS, lambda fn: self.generate_range_info_formula(fn, value_rows, min_col, max_col))
        emit_block(self.ARRAY_RESHAPE_FUNCTIONS, lambda fn: self.generate_array_reshape_formula(fn, value_rows, min_col, max_col))

        # DATABASE_FUNCTIONS needs its own per-call criteria cells (not
        # just a formula), so it gets a dedicated column and a bespoke
        # loop rather than the generic emit_block.
        db_formula_col = next_col
        db_crit_col = next_col + 1
        for i, fn in enumerate(self.DATABASE_FUNCTIONS):
            formula = self.generate_database_formula(fn, ws, db_crit_col, i)
            ws.cell(row=i + 1, column=db_formula_col, value=formula)

        # --- Cross-sheet block: a real second sheet, so quoted sheet names
        # with spaces (`'Data Sheet'!A1`) and WorkbookManager::evaluate()'s
        # 3-pass cross-sheet propagation are actually exercised end-to-end --
        # previously this generator only ever produced single-sheet
        # workbooks (#26).
        #
        # A strict one-directional chain (never a cycle, which neither
        # engine is guaranteed to resolve the same way):
        #   Sheet1 table body (plain values)
        #     -> Sheet1!{x1_cell} (already-generated formula-block cell)
        #     -> Sheet2!B1 (SUM over Sheet1's table body, plus x1_cell)
        #     -> Sheet1!{cross_col}1 (references Sheet2!B1)
        # so resolving the final cell genuinely requires Sheet1's own
        # formula pass to complete, then Sheet2 to see it, then Sheet1
        # again -- exactly the kind of chain the 3 fixed passes exist for.
        ws2 = wb.create_sheet("Data Sheet")
        for r in range(1, value_rows + 1):
            for c in range(1, num_cols + 1):
                val = self.generate_random_value()
                if val is not None:
                    ws2.cell(row=r, column=c, value=val)

        x1_cell = f"{self._col_name(min_col)}{value_rows + 1}"
        table_body = f"Sheet1!A2:{self._col_name(num_cols)}{num_rows}"
        ws2.cell(row=1, column=num_cols + 1, value=f"=SUM({table_body})+Sheet1!{x1_cell}")

        cross_sheet_col = db_crit_col + 1
        ws.cell(row=1, column=cross_sheet_col, value=f"='Data Sheet'!{self._col_name(num_cols + 1)}1*2")

        # --- Reference/value-entropy block: deterministic formulas whose
        # purpose is to make sure every iteration contains address and source
        # value forms that pure random generation may only hit rarely. The
        # operands are chosen from formula-free helper columns so the cells
        # compare the targeted feature itself, not random type/error
        # propagation.
        ref_entropy_col = cross_sheet_col + 1
        fcash = self._col_name(fin_cash_col)
        fdate = self._col_name(fin_date_col)
        fsched = self._col_name(fin_schedule_col)
        ws.cell(row=1, column=ref_entropy_col, value="=SUM($A:$A)")
        ws.cell(row=2, column=ref_entropy_col, value=f"=${fcash}$1+{fdate}$1+${fsched}1")
        ws.cell(row=3, column=ref_entropy_col, value="=COUNTA('Data Sheet'!$A:$A)")

        numeric_entropy_col = ref_entropy_col + 1
        numeric_formula_col = numeric_entropy_col + 1
        wide_int = random.randint(10_000_000, 1_000_000_000)
        tiny_float = round(random.uniform(1.0, 500.0) * 1e-9, 12)
        ws.cell(row=1, column=numeric_entropy_col, value=wide_int)
        ws.cell(row=2, column=numeric_entropy_col, value=tiny_float)
        ws.cell(row=1, column=numeric_formula_col, value=f"={self._col_name(numeric_entropy_col)}1+1")
        ws.cell(row=2, column=numeric_formula_col, value=f"={self._col_name(numeric_entropy_col)}2*1000000")

        date_entropy_col = numeric_formula_col + 1
        date_formula_col = date_entropy_col + 1
        d1 = datetime.date(random.randint(1995, 2035), random.randint(1, 12), random.randint(1, 28))
        d2 = d1 + datetime.timedelta(days=random.randint(1, 60))
        ws.cell(row=1, column=date_entropy_col, value=d1)
        ws.cell(row=2, column=date_entropy_col, value=d2)
        ws.cell(row=1, column=date_formula_col, value=f"=YEAR({self._col_name(date_entropy_col)}1)")
        ws.cell(row=2, column=date_formula_col, value=f"={self._col_name(date_entropy_col)}2-{self._col_name(date_entropy_col)}1")

        criteria_data_col = date_formula_col + 1
        criteria_formula_col = criteria_data_col + 1
        criteria_values = ["Alpha", "alphabet", "Beta", "", "A?pha"]
        for i, value in enumerate(criteria_values, start=1):
            ws.cell(row=i, column=criteria_data_col, value=value)
        crit_col = self._col_name(criteria_data_col)
        ws.cell(row=1, column=criteria_formula_col, value=f'=COUNTIF({crit_col}1:{crit_col}5,"Al*")')
        ws.cell(row=2, column=criteria_formula_col, value=f'=COUNTIF({crit_col}1:{crit_col}5,"<>")')
        ws.cell(row=3, column=criteria_formula_col, value=f'=COUNTIF({crit_col}1:{crit_col}5,"A~?pha")')

        # Final pass: add the `_xlfn.` prefix every post-2007 function in
        # NEEDS_XLFN_PREFIX needs to be recognized when the file is opened
        # by real Excel (see _apply_xlfn_prefixes). Done once here, over
        # every formula cell on every sheet, rather than in each
        # individual generator method above.
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cell.value = self._apply_xlfn_prefixes(cell.value)

        wb.save(file_path)


# -----------------------------------------------------------------------------
# 2. Execution Drivers (visi & Microsoft Excel)
# -----------------------------------------------------------------------------

# `VisiDriver` moved to visi_driver.py, which drives visi either through the
# in-process `visi_core` bindings or through the CLI. Re-exported here because
# reverse_engineer_financial.py imports it (and ExcelDriver, and
# XLSXEvaluatedReader) from this module.
from visi_driver import (  # noqa: E402,F401
    VisiDriver,
    add_backend_arg,
    bindings_available,
    bindings_hint,
    resolve_visi_binary,
)


class ExcelDriver:
    """Invokes Microsoft Excel to recalculate and save a workbook."""
    def __init__(self, excel_path=None, driver_type="auto"):
        self.excel_path = excel_path
        self.driver_type = driver_type
        if driver_type == "auto":
            if sys.platform == "darwin":
                self.driver_type = "applescript"
            elif sys.platform == "win32":
                self.driver_type = "win32com"
            else:
                self.driver_type = "mock"

    def run(self, input_file, output_file):
        # First copy input_file to output_file so Excel modifies output_file
        shutil.copyfile(input_file, output_file)
        abs_output = os.path.abspath(output_file)

        if self.driver_type == "mock":
            # For testing the harness when Excel is not available
            print("[ExcelDriver Warning] Running in mock mode (Excel not invoked).")
            return

        elif self.driver_type == "applescript":
            # macOS AppleScript Excel recalculation driver
            app_name = self.excel_path if self.excel_path else "Microsoft Excel"
            if app_name.endswith(".app"):
                app_name = os.path.splitext(os.path.basename(app_name))[0]
            script = f'''
            tell application "{app_name}"
                set display alerts to false
                try
                    close workbooks saving no
                end try
                try
                    set targetFile to POSIX file "{abs_output}"
                    open targetFile
                    calculate
                    save active workbook
                    close active workbook saving no
                on error errText number errNum
                    try
                        close workbooks saving no
                    end try
                    error errText number errNum
                end try
            end tell
            '''
            res = None
            for attempt in range(5):
                time.sleep(0.5)
                try:
                    res = subprocess.run(["osascript", "-e", script], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=15)
                    if res.returncode == 0:
                        break
                except subprocess.TimeoutExpired:
                    subprocess.run(["killall", "Microsoft Excel"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    time.sleep(1.0)
            if res is not None and res.returncode != 0:
                raise RuntimeError(f"Excel AppleScript failed:\nSTDERR: {res.stderr}")

        elif self.driver_type == "win32com":
            # Windows COM Excel driver
            try:
                import win32com.client
            except ImportError:
                raise RuntimeError("pywin32 (win32com) is required for Excel automation on Windows.")

            # COM automation against a fresh Excel.Application is occasionally
            # flaky in a way that surfaces as unrelated-looking Python errors
            # (e.g. "'bool' object is not callable" out of win32com's dynamic
            # dispatch) rather than a COM error -- transient, not
            # reproducible, and not an Excel/visi disagreement. Retry with a
            # fresh Application instance, mirroring the AppleScript driver's
            # retry loop above.
            last_err = None
            for attempt in range(5):
                if attempt > 0:
                    subprocess.run(
                        ["taskkill", "/F", "/IM", "EXCEL.EXE"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                    )
                    time.sleep(1.0)
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                try:
                    wb = excel.Workbooks.Open(abs_output)
                    excel.Calculate()
                    wb.Save()
                    wb.Close()
                    last_err = None
                    break
                except Exception as e:
                    last_err = e
                finally:
                    excel.Quit()
            if last_err is not None:
                raise last_err

        elif self.driver_type == "cli":
            # Custom CLI executable driver specified by user
            cmd = [self.excel_path, abs_output]
            res = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            if res.returncode != 0:
                raise RuntimeError(f"Excel CLI command failed:\nSTDERR: {res.stderr}")


# -----------------------------------------------------------------------------
# 3. OpenXML Evaluated Content Extractor & Evaluated Value Reader
# -----------------------------------------------------------------------------

class XLSXEvaluatedReader:
    """Parses raw OpenXML structure directly from .xlsx zip files to read cached evaluated cell values."""

    @staticmethod
    def read_evaluated_cells(file_path):
        """
        Returns a dict mapping (sheet_name, cell_ref) -> dict of:
        {
            'raw_value': str,
            'formula': str or None,
            'type': str ('n', 's', 'b', 'e', 'str'),
            'normalized_val': parsed python object
        }
        """
        if not os.path.exists(file_path):
            return {}
        with open(file_path, 'rb') as f:
            return XLSXEvaluatedReader.read_evaluated_cells_bytes(f.read(), source=file_path)

    @staticmethod
    def read_evaluated_cells_bytes(data, source="<bytes>"):
        """The same parse, from an in-memory .xlsx.

        Deliberately the identical code path -- only where the zip comes from
        differs -- so this stays an *independent* reader of what visi actually
        wrote. It is not a shortcut around visi's exporter: reading values out
        of the engine's memory instead would stop exercising
        `export_xlsx_data`, the cached `<v>` tags, shared strings, and (for
        pivots) the hand-rolled OOXML injection, which is a large share of what
        the differential harness exists to check.
        """
        results = {}

        try:
            with zipfile.ZipFile(io.BytesIO(data), 'r') as z:
                # 1. Load shared strings table if present
                shared_strings = []
                if 'xl/sharedStrings.xml' in z.namelist():
                    with z.open('xl/sharedStrings.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        for si in root.findall('.//main:si', NS):
                            # String value can be direct <t> or nested in <r><t>
                            t_elems = si.findall('.//main:t', NS)
                            text = "".join(t.text or "" for t in t_elems)
                            shared_strings.append(text)

                # 2. Iterate sheet XML files
                sheet_files = [name for name in z.namelist() if name.startswith('xl/worksheets/sheet') and name.endswith('.xml')]
                for sheet_file in sorted(sheet_files):
                    sheet_name = os.path.basename(sheet_file).replace('.xml', '')
                    with z.open(sheet_file) as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        for cell in root.findall('.//main:c', NS):
                            ref = cell.attrib.get('r')
                            cell_type = cell.attrib.get('t', 'n')  # default numeric

                            f_elem = cell.find('main:f', NS)
                            formula = f_elem.text if f_elem is not None else None

                            if cell_type == 'inlineStr':
                                # Inline strings live in <is><t>...</t></is> (or
                                # <is><r><t>...</t></r>...</is> for rich-text runs),
                                # not <v> -- real Excel writes these for some string
                                # cells instead of using the shared-strings table
                                # (observed via fuzz_pivot.py's AppleScript/VBA
                                # macro save path).
                                is_elem = cell.find('main:is', NS)
                                if is_elem is not None:
                                    t_elems = is_elem.findall('.//main:t', NS)
                                    raw_val = "".join(t.text or "" for t in t_elems)
                                else:
                                    raw_val = None
                            else:
                                v_elem = cell.find('main:v', NS)
                                raw_val = v_elem.text if v_elem is not None else None

                            normalized_val = XLSXEvaluatedReader._normalize_cell(cell_type, raw_val, shared_strings)

                            results[(sheet_name, ref)] = {
                                'cell_ref': ref,
                                'sheet': sheet_name,
                                'type': cell_type,
                                'raw_value': raw_val,
                                'formula': formula,
                                'val': normalized_val
                            }
        except Exception as e:
            print(f"Warning: Failed to read OpenXML from {source}: {e}")

        return results

    @staticmethod
    def _normalize_cell(cell_type, raw_val, shared_strings):
        if raw_val is None:
            return None

        if cell_type == 's':
            # Shared string index
            try:
                idx = int(raw_val)
                if 0 <= idx < len(shared_strings):
                    return shared_strings[idx]
                return f"[Unknown string index {idx}]"
            except ValueError:
                return raw_val

        elif cell_type == 'b':
            # Boolean
            return raw_val == '1' or raw_val.lower() == 'true'

        elif cell_type == 'e':
            # Error code (e.g. #DIV/0!, #VALUE!, #N/A, #REF!, #NUM!, #NAME?, #NULL!)
            return raw_val.upper()

        elif cell_type == 'str' or cell_type == 'inlineStr':
            # Inline string
            return raw_val

        else:
            # Numeric (int or float)
            try:
                f_val = float(raw_val)
                if f_val.is_integer():
                    return int(f_val)
                return f_val
            except ValueError:
                return raw_val


# -----------------------------------------------------------------------------
# 3b. Smoke mode (--driver mock)
# -----------------------------------------------------------------------------
#
# The mock Excel driver copies the *unevaluated* source workbook and calls it
# Excel's output. openpyxl writes no cached <v> for a formula cell, so every
# formula cell reads as None on the "Excel" side -- 360 of 530 cells in a
# default grid. Comparing against that isn't a weak oracle, it's a guaranteed
# 100% mismatch, and the harness used to report it as such: exit 1 every run,
# plus a failure-artifact directory per iteration.
#
# So mock does not get compared. What it is actually for -- and what
# fuzz/README.md has always called it -- is a smoke test of the pipeline
# (generate -> visi -> read -> report) on a machine with no Excel automation,
# and a crash hunt over a large volume of generated formulas. Both work
# without an oracle; neither works with a fake one, because the real signal is
# invisible among the guaranteed noise.

SMOKE_BANNER = (
    "[mock] No Excel: running as a pipeline smoke test. Output is checked for "
    "readability only,\n        not correctness -- there is nothing to compare "
    "against. Crashes still fail the run."
)


def smoke_check(items, what="cells"):
    """The most that can be asserted with no oracle: visi wrote an .xlsx that
    parses and has content. Catches a corrupt or empty export, which is a real
    failure mode and one the comparison path would never reach.

    `items` is whatever that fuzzer reads back -- a cell dict, or a chart list.

    Returns (ok, reason).
    """
    if not items:
        return False, f"visi produced no readable {what}"
    return True, ""


# -----------------------------------------------------------------------------
# 4. Semantic Comparison Engine
# -----------------------------------------------------------------------------

class DifferentialComparator:
    """Compares evaluated cell contents between visi output and Excel output."""

    EXCEL_ERRORS = {"#DIV/0!", "#VALUE!", "#N/A", "#REF!", "#NUM!", "#NAME?", "#NULL!", "#CALC!", "#SPILL!"}

    # Matches a floating-point literal as it appears embedded in text, e.g.
    # inside CONCATENATE's output. Used to give text results the same
    # last-digit tolerance numeric cells already get -- see
    # `_numeric_text_equal`.
    _NUM_TOKEN_RE = re.compile(r"-?\d+\.?\d*(?:[eE][+-]?\d+)?")

    def __init__(self, float_rel_tol=1e-7, float_abs_tol=1e-7, strict_error_class=False):
        self.float_rel_tol = float_rel_tol
        self.float_abs_tol = float_abs_tol
        # When several sub-expressions of one formula each raise a
        # *different* error, visi and Excel sometimes surface different
        # ones -- which error wins depends on Excel's internal evaluation
        # order and differs per operator and per function. That is a
        # documented divergence ("docs/excel-discrepancies.md" section 13),
        # and by default a disagreement where *both* engines errored is
        # counted separately rather than as a failure.
        #
        # It is deliberately still counted and reported, not silently
        # dropped: strict error-class comparison is what surfaced genuine
        # bugs like TYPE(error) and ERROR.TYPE returning the wrong thing,
        # LOG(n, 1) being #NUM! instead of #DIV/0!, and CHITEST's #N/A
        # cases. Pass --strict-error-class to make them failures again.
        self.strict_error_class = strict_error_class
        self.error_class_only = 0

    def compare(self, visi_cells, excel_cells):
        """
        Compares two cell dictionaries.
        Returns (is_match, mismatches)
        """
        all_keys = set(visi_cells.keys()).union(set(excel_cells.keys()))
        mismatches = []

        for key in sorted(all_keys):
            v_cell = visi_cells.get(key)
            e_cell = excel_cells.get(key)

            # A cell that is absent altogether is compared with the same
            # blank-equivalence rule `values_equal` applies to a cell that is
            # present but empty -- otherwise the two paths disagree with each
            # other. Excel keeps a whitespace-only source cell as an
            # empty-string cell (`<t/>` in the shared strings) where visi
            # writes no cell at all, and both mean "nothing here"; visi treats
            # blank and empty string as the same value throughout, down to
            # ISBLANK("") being TRUE. A genuinely missing *value* still fails.
            if v_cell is None and e_cell is not None:
                if not self.values_equal(None, e_cell['val']):
                    mismatches.append({
                        'key': key,
                        'reason': 'Missing in visi output',
                        'visi': None,
                        'excel': e_cell['val'],
                        'formula': e_cell.get('formula')
                    })
                continue

            if e_cell is None and v_cell is not None:
                if not self.values_equal(v_cell['val'], None):
                    mismatches.append({
                        'key': key,
                        'reason': 'Missing in Excel output',
                        'visi': v_cell['val'],
                        'excel': None,
                        'formula': v_cell.get('formula')
                    })
                continue

            v_val = v_cell['val']
            e_val = e_cell['val']
            formula = v_cell.get('formula') or e_cell.get('formula')

            if not self.values_equal(v_val, e_val):
                if not self.strict_error_class and self._both_errors(v_val, e_val):
                    self.error_class_only += 1
                    continue
                mismatches.append({
                    'key': key,
                    'reason': f"Value mismatch ({type(v_val).__name__} vs {type(e_val).__name__})",
                    'visi': v_val,
                    'excel': e_val,
                    'formula': formula
                })

        return len(mismatches) == 0, mismatches

    def _both_errors(self, v1, v2):
        """True when both sides are Excel errors that merely differ in class."""
        return (
            isinstance(v1, str)
            and isinstance(v2, str)
            and v1.upper() in self.EXCEL_ERRORS
            and v2.upper() in self.EXCEL_ERRORS
        )

    @staticmethod
    def _parse_complex(text):
        """(real, imag, suffix) for an Excel complex literal like "3+4i",
        "-2.5e-3-1.5j", "7i" or "-j"; None if `text` isn't one."""
        s = text.strip()
        if not s or s[-1] not in "ij":
            return None
        suffix = s[-1]
        body = s[:-1]
        # Split on the last +/- that isn't an exponent sign.
        split_at = None
        for i in range(len(body) - 1, 0, -1):
            if body[i] in "+-" and body[i - 1] not in "eE":
                split_at = i
                break
        try:
            if split_at is None:
                imag_str = body
                real = 0.0
            else:
                real = float(body[:split_at])
                imag_str = body[split_at:]
            if imag_str in ("", "+"):
                imag = 1.0
            elif imag_str == "-":
                imag = -1.0
            else:
                imag = float(imag_str)
        except ValueError:
            return None
        return real, imag, suffix

    def _numeric_text_equal(self, v1, v2):
        """True if v1 and v2 are identical text apart from embedded
        floating-point numbers that differ only in the last of Excel's 15
        significant digits. CONCATENATE (and `&`) can stitch two numeric
        sub-results straight into text with no separator -- e.g. two ATAN2
        calls glued together as "2.21429743558818-1.61511318944808" -- so a
        sub-ULP double-rounding disagreement that numeric cells already
        tolerate (see the numeric branch above and the IM* case below)
        would otherwise surface as a plain string mismatch here.
        """
        parts1 = self._NUM_TOKEN_RE.split(v1)
        parts2 = self._NUM_TOKEN_RE.split(v2)
        if parts1 != parts2:
            return False
        nums1 = self._NUM_TOKEN_RE.findall(v1)
        nums2 = self._NUM_TOKEN_RE.findall(v2)
        if not nums1 or len(nums1) != len(nums2):
            return False
        try:
            return all(
                math.isclose(float(a), float(b), rel_tol=self.float_rel_tol, abs_tol=self.float_abs_tol)
                for a, b in zip(nums1, nums2)
            )
        except ValueError:
            return False

    def values_equal(self, v1, v2):
        """Checks equality between two evaluated values with floating-point tolerance."""
        if v1 is None and v2 is None:
            return True
        if v1 is None or v2 is None:
            if (v1 is None and isinstance(v2, str) and not v2.strip()) or \
               (v2 is None and isinstance(v1, str) and not v1.strip()):
                return True
            return False

        # If both are numbers (float or int)
        if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
            return math.isclose(float(v1), float(v2), rel_tol=self.float_rel_tol, abs_tol=self.float_abs_tol)

        # If one is a number and the other is a numeric string (e.g. 0.0394 vs ".0394", 8 vs "08")
        if isinstance(v1, (int, float)) and isinstance(v2, str):
            try:
                return math.isclose(float(v1), float(v2), rel_tol=self.float_rel_tol, abs_tol=self.float_abs_tol)
            except ValueError:
                pass
        if isinstance(v2, (int, float)) and isinstance(v1, str):
            try:
                return math.isclose(float(v1), float(v2), rel_tol=self.float_rel_tol, abs_tol=self.float_abs_tol)
            except ValueError:
                pass

        # If both are error strings or text strings
        if isinstance(v1, str) and isinstance(v2, str):
            if v1.upper() in self.EXCEL_ERRORS or v2.upper() in self.EXCEL_ERRORS:
                return v1.upper() == v2.upper()
            if v1.strip() == v2.strip():
                return True
            # The IM* family returns its result as *text* ("3+4i"), so a
            # plain string comparison would flag a disagreement in the last
            # displayed digit -- exactly the kind of float noise the
            # numeric branch above already tolerates for ordinary numbers.
            # Compare component-wise with the same tolerance instead, but
            # only when both sides really are complex literals (and agree
            # on the i/j suffix, which is meaningful and must match).
            c1 = self._parse_complex(v1)
            c2 = self._parse_complex(v2)
            if c1 is not None and c2 is not None:
                (re1, im1, suf1), (re2, im2, suf2) = c1, c2
                return suf1 == suf2 and all(
                    math.isclose(a, b, rel_tol=self.float_rel_tol, abs_tol=self.float_abs_tol)
                    for a, b in ((re1, re2), (im1, im2))
                )
            return self._numeric_text_equal(v1, v2)

        # Booleans vs strings/numbers (e.g. True vs 1, "TRUE" vs True, "FALSE" vs False)
        if isinstance(v1, bool) or isinstance(v2, bool):
            def to_b(v):
                if isinstance(v, bool):
                    return v
                if isinstance(v, str):
                    return v.upper() in ("TRUE", "1")
                if isinstance(v, (int, float)):
                    return v != 0
                return bool(v)
            return to_b(v1) == to_b(v2)

        return str(v1) == str(v2)


# -----------------------------------------------------------------------------
# 5. CLI & Test Runner Orchestrator
# -----------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Differential fuzzing test harness for visi vs Microsoft Excel.")
    parser.add_argument("--excel-path", help="Path to Microsoft Excel binary or application bundle (e.g. '/Applications/Microsoft Excel.app').")
    parser.add_argument("--driver", choices=["auto", "applescript", "win32com", "cli", "mock"], default="auto", help="Excel execution driver.")
    parser.add_argument("--visi-path", default="./target/release/visi", help="Path to compiled visi binary (used by the subprocess backend).")
    add_backend_arg(parser)
    parser.add_argument("--iterations", type=int, default=10, help="Number of fuzz iterations to run.")
    parser.add_argument("--rows", type=int, default=10, help="Number of rows per sheet.")
    parser.add_argument("--cols", type=int, default=5, help="Number of columns per sheet.")
    parser.add_argument("--seed", type=int, default=None, help="Random seed for reproducible fuzzing.")
    parser.add_argument("--output-dir", default="./fuzz_results", help="Directory to store test outputs and failure artifacts.")
    parser.add_argument(
        "--strict-error-class",
        action="store_true",
        help=(
            "Count a disagreement where both engines errored but with different "
            "error classes as a failure. Off by default -- see "
            "docs/excel-discrepancies.md section 13."
        ),
    )
    args = parser.parse_args()

    ExcelFuzzGenerator._check_text_function_generators()

    os.makedirs(args.output_dir, exist_ok=True)
    failures_dir = os.path.join(args.output_dir, "failures")
    os.makedirs(failures_dir, exist_ok=True)

    generator = ExcelFuzzGenerator(seed=args.seed)
    visi_driver = VisiDriver(binary_path=args.visi_path, backend=args.backend)
    # Only warn when the fallback was involuntary. Asking for --backend
    # subprocess on purpose is not something to nag about.
    if args.backend == "auto" and visi_driver.backend != "bindings":
        print(bindings_hint(), file=sys.stderr)
    excel_driver = ExcelDriver(excel_path=args.excel_path, driver_type=args.driver)
    comparator = DifferentialComparator(strict_error_class=args.strict_error_class)
    smoke_mode = excel_driver.driver_type == "mock"

    print("=====================================================================")
    print("        visi vs. Microsoft Excel Differential Fuzzing Harness       ")
    print("=====================================================================")
    print(f" Iterations : {args.iterations}")
    print(f" Grid Size  : {args.rows}x{args.cols}")
    print(f" Visi       : {visi_driver.describe()}")
    print(f" Excel Driver: {excel_driver.driver_type} ({args.excel_path or 'Default'})")
    if smoke_mode:
        print(f" {SMOKE_BANNER}")
    print("=====================================================================\n")

    passed_count = 0
    failed_count = 0
    start_time = time.time()

    for i in range(1, args.iterations + 1):
        iter_seed = (args.seed + i) if args.seed is not None else random.randint(1, 1000000)
        iter_gen = ExcelFuzzGenerator(seed=iter_seed)

        temp_dir = tempfile.mkdtemp(prefix=f"fuzz_iter_{i}_")
        source_xlsx = os.path.join(temp_dir, "source.xlsx")
        visi_out_xlsx = os.path.join(temp_dir, "visi_out.xlsx")
        excel_out_xlsx = os.path.join(temp_dir, "excel_out.xlsx")

        try:
            # 1. Generate source workbook
            iter_gen.create_fuzz_workbook(source_xlsx, num_rows=args.rows, num_cols=args.cols)

            # 2. Evaluate with visi
            visi_bytes = visi_driver.run(source_xlsx, visi_out_xlsx)

            # 3. Evaluate with Excel, if there is an Excel to evaluate with
            if not smoke_mode:
                excel_driver.run(source_xlsx, excel_out_xlsx)

            # 4. Compare evaluated cell contents. visi's output is parsed from
            # the bytes it just wrote -- the same bytes now on disk in
            # visi_out_xlsx, so this is a saved read, not a different oracle.
            visi_cells = XLSXEvaluatedReader.read_evaluated_cells_bytes(
                visi_bytes, source=visi_out_xlsx
            )

            if smoke_mode:
                ok, reason = smoke_check(visi_cells)
                if ok:
                    passed_count += 1
                    print(
                        f" Iteration {i:3d}/{args.iterations} [OK] "
                        f"(Seed: {iter_seed}, {len(visi_cells)} cells)"
                    )
                else:
                    failed_count += 1
                    print(f"\n Iteration {i:3d}/{args.iterations} [FAILED] (Seed: {iter_seed})")
                    print(f"   {reason}")
                    fail_case_dir = os.path.join(failures_dir, f"smoke_iter_{i}_seed_{iter_seed}")
                    shutil.copytree(temp_dir, fail_case_dir, dirs_exist_ok=True)
                    print(f"   Saved reproducing files to: {fail_case_dir}\n")
                continue

            excel_cells = XLSXEvaluatedReader.read_evaluated_cells(excel_out_xlsx)

            is_match, mismatches = comparator.compare(visi_cells, excel_cells)

            if is_match:
                passed_count += 1
                print(f" Iteration {i:3d}/{args.iterations} [PASSED] (Seed: {iter_seed})")
            else:
                failed_count += 1
                print(f"\n Iteration {i:3d}/{args.iterations} [FAILED] (Seed: {iter_seed})")
                print(f"   Found {len(mismatches)} cell mismatch(es):")
                for m in mismatches[:5]:  # Print first 5 mismatches
                    print(f"   - Cell {m['key'][1]} on {m['key'][0]}: visi={m['visi']} | Excel={m['excel']} (Formula: {m['formula']})")

                # Save failure artifact
                fail_case_dir = os.path.join(failures_dir, f"fail_iter_{i}_seed_{iter_seed}")
                shutil.copytree(temp_dir, fail_case_dir, dirs_exist_ok=True)
                print(f"   Saved failure reproducing files to: {fail_case_dir}\n")

        except Exception as err:
            failed_count += 1
            print(f"\n Iteration {i:3d}/{args.iterations} [ERROR]: {err}")
            fail_case_dir = os.path.join(failures_dir, f"error_iter_{i}_seed_{iter_seed}")
            if os.path.exists(temp_dir):
                shutil.copytree(temp_dir, fail_case_dir, dirs_exist_ok=True)

        finally:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)

    duration = time.time() - start_time
    print("\n=====================================================================")
    print(f" {'Smoke test' if smoke_mode else 'Fuzzing'} Completed in {duration:.2f}s")
    if smoke_mode:
        print(f" Ran    : {passed_count}/{args.iterations} without a crash")
    else:
        print(f" Passed : {passed_count}/{args.iterations}")
    print(f" Failed : {failed_count}/{args.iterations}")
    if comparator.error_class_only:
        print(
            f" Tolerated: {comparator.error_class_only} cell(s) where both engines"
            " errored with different error classes"
        )
        print(
            "            (documented divergence; re-run with --strict-error-class"
            " to treat as failures)"
        )
    print("=====================================================================")

    if failed_count > 0:
        sys.exit(1)

if __name__ == "__main__":
    main()
